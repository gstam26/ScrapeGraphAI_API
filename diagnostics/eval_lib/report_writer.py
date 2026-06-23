"""
Report writer for the extraction evaluation framework (Stage 10 / RQ3).

Turns an AlignmentResult + MetricsReport into the reviewable Excel deliverable, and
handles the manual-review round-trip (read a filled "Manual Verdict" column back so
recall_auto/recall_full collapse to the analyst-resolved truth).

Sheets written:
  1. Claim Alignment    one row per (GT claim, best AI claim) + AI-only rows; carries
                        the blank, human-fillable "Manual Verdict" column.
  2. Cell Metrics       per (entity, question).
  3. Entity Summary     per entity.
  4. Overall            macro overall + one row per question.
  5. Tag Slice          recall by type/dimension (post-hoc slicing; tags never scored).
  6. Manual Review Queue manual-band GT rows + possible-GT-gap rows, most-uncertain
                        first, with a TRUE_MATCH/FALSE_MATCH/EXCLUDE dropdown.

Round-trip:
  read_manual_verdicts(path)        -> {(entity_norm, question, gt_claim_id): VERDICT}
  apply_manual_verdicts(result, v)  -> new AlignmentResult with manual rows resolved
                                       (TRUE_MATCH->auto_match, FALSE_MATCH->auto_miss,
                                        EXCLUDE->dropped). No thresholds/matching change.

This module does no scoring — it only renders and round-trips.
"""

from __future__ import annotations

import os
import sys
from dataclasses import replace

# --- repo-root bootstrap -----------------------------------------------------
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from diagnostics.eval_lib.aligner import AlignmentResult, CellAlignment
from diagnostics.eval_lib.gt_reader import GroundTruth, normalise_entity
from diagnostics.eval_lib.metrics import (
    GroupMetrics,
    MetricsReport,
    ai_only_audit_rows,
)

# Colours (aligned with the pipeline's io_excel palette).
_HEADER_FILL = "2E4057"
_HEADER_FONT = "FFFFFF"
_ALT_ROW = "F5F5F5"
_VERDICT_FILL = {
    "auto_match": "C8E6C9",   # green
    "manual": "FFE0B2",       # amber
    "auto_miss": "FFCDD2",    # red
    "null_match": "E0E0E0",   # grey
    "ai_only": "E1F5FE",      # light blue
}
_TAB_COLORS = {
    "Claim Alignment": "4CAF50",
    "Cell Metrics": "009688",
    "Entity Summary": "3F51B5",
    "Overall": "2E4057",
    "Tag Slice": "9C27B0",
    "Manual Review Queue": "FF9800",
}
_MANUAL_VERDICT = "Manual Verdict"
_VERDICT_CHOICES = '"TRUE_MATCH,FALSE_MATCH,EXCLUDE"'
_MATCHED_VERDICTS = frozenset({"auto_match", "null_match", "manual"})


# --- row builders ------------------------------------------------------------
def _score_cols(score) -> dict:
    if score is None:
        return {"claim_cosine": "", "quote_overlap": "", "combined_score": "", "method": ""}
    return {
        "claim_cosine": round(score.claim_cosine, 4),
        "quote_overlap": round(score.quote_overlap, 4),
        "combined_score": round(score.combined_score, 4),
        "method": score.method,
    }


def _alignment_rows(result: AlignmentResult, gt: GroundTruth) -> list[dict]:
    """Claim Alignment sheet: GT rows (with best AI) then AI-only rows."""
    rows: list[dict] = []
    for cell in result.cells:
        for a in cell.alignments:
            g = a.gt_claim
            ai = a.ai_claim
            row = {
                "entity": cell.entity, "question": cell.gt_question,
                "gt_claim_id": g.claim_id, "gt_claim": g.claim,
                "gt_verbatim_quote": g.verbatim_quote,
                "gt_type": g.type, "gt_dimension": g.dimension,
                "ai_claim": str(ai.value) if ai else "",
                "ai_quote": (ai.quote or "") if ai else "",
                "ai_verified": ai.verified if ai else "",
                "ai_match_type": ai.match_type if ai else "",
                "ai_semantic_score": (ai.semantic_score if ai else ""),
                **_score_cols(a.score),
                "auto_verdict": a.verdict,
                _MANUAL_VERDICT: "",
                "notes": a.note,
            }
            rows.append(row)

    # AI-only rows (raw hypotheses; categories are the metrics audit's best guess).
    for au in ai_only_audit_rows(result, gt):
        rows.append({
            "entity": au.entity, "question": au.question,
            "gt_claim_id": "", "gt_claim": "", "gt_verbatim_quote": "",
            "gt_type": "", "gt_dimension": "",
            "ai_claim": au.value, "ai_quote": au.quote,
            "ai_verified": au.verified, "ai_match_type": au.match_type,
            "ai_semantic_score": au.semantic_score if au.semantic_score is not None else "",
            "claim_cosine": "", "quote_overlap": "",
            "combined_score": au.nearest_gt_score if au.nearest_gt_score is not None else "",
            "method": au.nearest_gt_method,
            "auto_verdict": "ai_only",
            _MANUAL_VERDICT: "",
            "notes": f"[{au.category}] nearest {au.nearest_gt_claim_id}: {au.reason}",
        })
    return rows


def _group_row(g: GroupMetrics) -> dict:
    return {
        "label": g.label, "n_cells": g.n_cells, "gt_active": g.gt_active,
        "ai_distinct": g.ai_distinct,
        "recall_auto": g.recall_auto, "recall_full": g.recall_full,
        "precision_strict": g.precision_strict, "precision_distinct": g.precision_distinct,
        "f1_strict": g.f1_strict, "f1_distinct": g.f1_distinct,
        "avg_match_cosine": g.avg_match_cosine if g.avg_match_cosine is not None else "",
        "redundant": g.redundant_restatements, "out_of_scope_fp": g.out_of_scope_fp,
        "dynamic_neutral": g.dynamic_neutral,
        "possible_gt_gap": g.possible_gt_gap, "hallucinations": g.hallucinations,
    }


def _cell_rows(report: MetricsReport) -> list[dict]:
    out = []
    for c in report.cells:
        out.append({
            "entity": c.entity, "question": c.question, "type": c.question_type,
            "gt_active": c.gt_active, "ai_distinct": c.ai_distinct, "tp": c.tp,
            "recall_auto": c.recall_auto, "recall_full": c.recall_full,
            "precision_strict": c.precision_strict, "precision_distinct": c.precision_distinct,
            "f1_strict": c.f1_strict, "f1_distinct": c.f1_distinct,
            "avg_match_cosine": c.avg_match_cosine if c.avg_match_cosine is not None else "",
            "manual_band": c.manual_band, "redundant": c.redundant_restatements,
            "possible_gt_gap": c.possible_gt_gap, "hallucinations": c.hallucinations,
            "out_of_scope_fp": c.out_of_scope_fp, "dynamic_neutral": c.dynamic_neutral,
            "source_fidelity": c.source_fidelity if c.source_fidelity is not None else "",
        })
    return out


def _queue_rows(result: AlignmentResult, gt: GroundTruth) -> list[dict]:
    """Manual-band GT rows (recall resolution) + possible-GT-gap rows (completeness),
    most-uncertain first."""
    rows: list[dict] = []
    for cell in result.cells:
        for a in cell.alignments:
            if a.verdict != "manual":
                continue
            sc = _score_cols(a.score)
            rows.append({
                "review_type": "recall (GT match?)",
                "entity": cell.entity, "question": cell.gt_question,
                "gt_claim_id": a.gt_claim.claim_id, "gt_claim": a.gt_claim.claim,
                "ai_claim": str(a.ai_claim.value) if a.ai_claim else "",
                "combined_score": sc["combined_score"], "method": sc["method"],
                _MANUAL_VERDICT: "", "notes": a.note,
            })
    queue_sort = sorted(
        rows, key=lambda r: (r["combined_score"] if r["combined_score"] != "" else 1.0))

    gap_rows = []
    for au in ai_only_audit_rows(result, gt):
        if au.category.startswith("possible_gt_gap"):
            gap_rows.append({
                "review_type": "completeness (GT gap?)",
                "entity": au.entity, "question": au.question,
                "gt_claim_id": "", "gt_claim": f"(nearest {au.nearest_gt_claim_id})",
                "ai_claim": au.value,
                "combined_score": au.nearest_gt_score if au.nearest_gt_score is not None else "",
                "method": au.nearest_gt_method,
                _MANUAL_VERDICT: "", "notes": au.reason,
            })
    return queue_sort + gap_rows


# --- styling -----------------------------------------------------------------
def _style(ws, tab_color: str, verdict_col: str | None = None) -> None:
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"
    h_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    h_font = Font(name="Arial", bold=True, color=_HEADER_FONT, size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
    ws.row_dimensions[1].height = 28

    headers = [str(c.value) for c in ws[1]]
    vcol_idx = headers.index(verdict_col) + 1 if verdict_col and verdict_col in headers else None

    b_font = Font(name="Arial", size=10)
    b_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alt = PatternFill("solid", fgColor=_ALT_ROW)
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = b_font
            cell.alignment = b_align
            if r_idx % 2 == 0:
                cell.fill = alt
        if vcol_idx is not None:
            v = ws.cell(row=r_idx, column=vcol_idx).value
            fill = _VERDICT_FILL.get(str(v))
            if fill:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill)

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((max((len(l) for l in str(c.value).split("\n")), default=0)
                       for c in col if c.value is not None), default=0)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 60)


def _add_verdict_dropdown(ws, manual_col_name: str) -> None:
    headers = [str(c.value) for c in ws[1]]
    if manual_col_name not in headers:
        return
    letter = get_column_letter(headers.index(manual_col_name) + 1)
    dv = DataValidation(type="list", formula1=_VERDICT_CHOICES, allow_blank=True)
    dv.add(f"{letter}2:{letter}{ws.max_row}")
    ws.add_data_validation(dv)


# --- main writer -------------------------------------------------------------
def write_report(
    result: AlignmentResult,
    report: MetricsReport,
    gt: GroundTruth,
    out_path: str,
) -> None:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    align_df = pd.DataFrame(_alignment_rows(result, gt))
    cell_df = pd.DataFrame(_cell_rows(report))
    entity_df = pd.DataFrame([_group_row(g) for g in report.by_entity])
    overall_df = pd.DataFrame([_group_row(report.overall)] + [_group_row(g) for g in report.by_question])
    tag_df = pd.DataFrame([_group_row(g) for g in report.by_tag])
    queue_df = pd.DataFrame(_queue_rows(result, gt))

    sheets = [
        ("Claim Alignment", align_df),
        ("Cell Metrics", cell_df),
        ("Entity Summary", entity_df),
        ("Overall", overall_df),
        ("Tag Slice", tag_df),
        ("Manual Review Queue", queue_df),
    ]
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in sheets:
            (df if not df.empty else pd.DataFrame({"(empty)": []})).to_excel(
                writer, sheet_name=name, index=False)

    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        verdict_col = "auto_verdict" if ws.title == "Claim Alignment" else None
        _style(ws, _TAB_COLORS.get(ws.title, _HEADER_FILL), verdict_col=verdict_col)
        if ws.title in {"Claim Alignment", "Manual Review Queue"}:
            _add_verdict_dropdown(ws, _MANUAL_VERDICT)
    wb.save(out_path)


# --- manual-review round-trip ------------------------------------------------
def read_manual_verdicts(path: str) -> dict[tuple[str, str, str], str]:
    """Read filled '{Manual Verdict}' values from a previously-written report.

    Returns {(entity_norm, question, gt_claim_id): VERDICT} for non-blank GT rows.
    Reads both 'Claim Alignment' and 'Manual Review Queue'; the queue wins on conflict.
    """
    verdicts: dict[tuple[str, str, str], str] = {}
    xls = pd.ExcelFile(path)
    try:
        for sheet in ("Claim Alignment", "Manual Review Queue"):
            if sheet not in xls.sheet_names:
                continue
            df = pd.read_excel(xls, sheet_name=sheet)
            cols = {str(c).strip().lower(): c for c in df.columns}
            e_c, q_c = cols.get("entity"), cols.get("question")
            id_c, v_c = cols.get("gt_claim_id"), cols.get(_MANUAL_VERDICT.lower())
            if not all([e_c, q_c, id_c, v_c]):
                continue
            for _, row in df.iterrows():
                verdict = str(row.get(v_c, "")).strip().upper()
                claim_id = str(row.get(id_c, "")).strip()
                if not verdict or verdict in ("NAN", "") or not claim_id:
                    continue
                key = (normalise_entity(str(row.get(e_c, ""))), str(row.get(q_c, "")).strip(), claim_id)
                verdicts[key] = verdict
    finally:
        xls.close()
    return verdicts


def apply_manual_verdicts(
    result: AlignmentResult,
    verdicts: dict[tuple[str, str, str], str],
) -> AlignmentResult:
    """Resolve GT alignments per analyst verdicts. Returns a NEW result; thresholds and
    matching are untouched. A filled verdict on ANY row is honoured (so the analyst can
    confirm a manual row OR dispute an auto_match); a blank leaves the row unchanged, so
    unresolved manual-band rows stay manual.

      TRUE_MATCH  -> auto_match
      FALSE_MATCH -> auto_miss
      EXCLUDE     -> row dropped (out of the recall denominator)
    """
    new_cells: list[CellAlignment] = []
    for cell in result.cells:
        new_aligns = []
        for a in cell.alignments:
            verdict = verdicts.get((cell.entity_norm, cell.gt_question, a.gt_claim.claim_id))
            if verdict == "TRUE_MATCH":
                new_aligns.append(replace(a, verdict="auto_match",
                                          note=(a.note + " | manual:TRUE_MATCH").strip(" |")))
            elif verdict == "FALSE_MATCH":
                new_aligns.append(replace(a, verdict="auto_miss",
                                          note=(a.note + " | manual:FALSE_MATCH").strip(" |")))
            elif verdict == "EXCLUDE":
                continue  # drop from the cell entirely
            else:
                new_aligns.append(a)  # blank/unknown -> unchanged
        new_cells.append(replace(cell, alignments=new_aligns))
    return replace(result, cells=new_cells)
