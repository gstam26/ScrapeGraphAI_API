"""
Ground-truth template generator.

Reads the standard pipeline input workbook (entities, urls, questions sheets)
and produces an empty Excel template for manual ground-truth completion.

Row count per (entity, question) pair adapts to the question type: questions
whose instruction field mentions "list" get 5 starter rows; all others get 1.

Optionally reads a pipeline output workbook (second positional arg) to populate
a Source URLs Fetched reference sheet with unique page URLs grouped by entity.

Usage:
    python diagnostics/ground_truth_template.py samples/input3.xlsx
    python diagnostics/ground_truth_template.py samples/input3.xlsx outputs/input3.xlsx
    python diagnostics/ground_truth_template.py samples/input3.xlsx --output outputs/my_gt.xlsx
"""

import argparse
import os
import re
import sys

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

from models import ColumnSpec
from src.io_excel import read_input

_HEADER_FILL = "2E4057"
_HEADER_FONT = "FFFFFF"
_ALT_ROW = "F5F5F5"
_PREPOP_FILL = "E8EAF6"  # light indigo for pre-populated (read-only intent) cells

_TAB_COLORS = {
    "Ground Truth": "4CAF50",
    "Source URLs Fetched": "009688",
    "Instructions": "2196F3",
}

_ROWS_LIST = 5     # starter rows for questions whose instruction mentions "list"
_ROWS_DEFAULT = 1  # starter rows for all other questions


def _question_short_code(name: str) -> str:
    """First word of question name, uppercased. Words longer than 6 chars are truncated to 4."""
    word = name.split()[0].upper() if name.split() else "Q"
    return word[:4] if len(word) > 6 else word


def _rows_for_column(col: ColumnSpec) -> int:
    if col.instruction and "list" in col.instruction.lower():
        return _ROWS_LIST
    return _ROWS_DEFAULT


def _entity_slug(entity: str) -> str:
    """Replace non-alphanumeric runs with '-' for use in Claim IDs."""
    return re.sub(r"[^a-zA-Z0-9]+", "-", entity).strip("-")


def _make_ground_truth_df(entities: list[str], columns: list[ColumnSpec]) -> pd.DataFrame:
    col_names = ["Claim ID", "Entity", "Question", "Claim", "Supporting Quote", "Source URL", "Notes"]
    rows = []
    for entity in entities:
        slug = _entity_slug(entity)
        for col in columns:
            short_code = _question_short_code(col.name)
            n_rows = _rows_for_column(col)
            for i in range(1, n_rows + 1):
                rows.append({
                    "Claim ID": f"{slug}-{short_code}-{i:02d}",
                    "Entity": entity,
                    "Question": col.name,
                    "Claim": "",
                    "Supporting Quote": "",
                    "Source URL": "",
                    "Notes": "",
                })
    return pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)


def _make_source_urls_df(pipeline_output_path: str) -> pd.DataFrame | None:
    """
    Read Acquire Log from a pipeline output workbook.
    Returns a DataFrame of unique (Entity, Page URL) pairs, or None on failure.
    """
    try:
        xls = pd.ExcelFile(pipeline_output_path)
        sheet_map = {s.strip().lower(): s for s in xls.sheet_names}
        acq_sheet = sheet_map.get("acquire log")
        if acq_sheet is None:
            print(f"  ! No 'Acquire Log' sheet in {pipeline_output_path} — skipping Source URLs sheet.")
            return None
        df = pd.read_excel(xls, sheet_name=acq_sheet)
        xls.close()
    except Exception as exc:
        print(f"  ! Could not read pipeline output ({exc}) — skipping Source URLs sheet.")
        return None

    col_map = {str(c).strip().lower(): c for c in df.columns}
    entities_col = col_map.get("entities")
    url_col = col_map.get("page url")
    if entities_col is None or url_col is None:
        print("  ! Acquire Log missing 'Entities' or 'Page URL' column — skipping Source URLs sheet.")
        return None

    rows = []
    seen: set[tuple[str, str]] = set()
    for _, row in df.iterrows():
        entities_raw = str(row[entities_col]).strip() if pd.notna(row[entities_col]) else ""
        url = str(row[url_col]).strip() if pd.notna(row[url_col]) else ""
        if not url or url.lower() == "nan":
            continue
        for entity in (e.strip() for e in entities_raw.split(",") if e.strip()):
            key = (entity, url)
            if key not in seen:
                seen.add(key)
                rows.append({"Entity": entity, "Page URL": url})

    if not rows:
        return None
    return pd.DataFrame(rows, columns=["Entity", "Page URL"])


def _make_instructions_df() -> pd.DataFrame:
    guidelines = [
        "One row per distinct claim. If a question has multiple answers for one entity, each answer gets its own row with its own Claim ID. To add extra rows, copy the row format and increment the suffix manually (e.g. -06, -07).",
        "Questions whose instruction mentions 'list' have 5 starter rows. If only 2 of those slots have claims, leave the other 3 blank — blank Claim rows are ignored during scoring.",
        "Supporting Quote must be copied verbatim from the source page — do not paraphrase or summarise. Paste the exact sentence or phrase from the page.",
        "Source URL should match the Page URL exactly as it appears in the Acquire Log or the Source URLs Fetched sheet (if provided). Copy-paste rather than typing to avoid mismatches.",
        "Claim ID, Entity, and Question are pre-populated (shaded blue) — do not edit them. Only fill in Claim, Supporting Quote, Source URL, and Notes.",
        "Notes is optional. Use it for: ambiguous attribution (e.g. claim applies to a subsidiary, not the parent entity), claims that span multiple pages, conflicting sources, or anything the analyst is unsure how to categorise.",
        "To add rows beyond the starter set, copy an existing row for that entity/question pair, paste it below, and set the Claim ID suffix to the next number (e.g. if the last was -05, the new row is -06).",
        "Leave Claim blank if no answer was found for that entity/question pair on any source page.",
    ]
    return pd.DataFrame(
        [{"#": i + 1, "Guideline": g} for i, g in enumerate(guidelines)],
        columns=["#", "Guideline"],
    )


def _style_sheet(ws, tab_color: str, prepopulated_cols: set[int] | None = None) -> None:
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"

    h_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    h_font = Font(name="Arial", bold=True, color=_HEADER_FONT, size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
    ws.row_dimensions[1].height = 28

    b_font = Font(name="Arial", size=10)
    b_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alt = PatternFill("solid", fgColor=_ALT_ROW)
    prepop = PatternFill("solid", fgColor=_PREPOP_FILL)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = b_font
            cell.alignment = b_align
            if prepopulated_cols and cell.column in prepopulated_cols:
                cell.fill = prepop
            elif row_idx % 2 == 0:
                cell.fill = alt

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                longest = max((len(line) for line in str(cell.value).split("\n")), default=0)
                max_len = max(max_len, longest)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate empty ground-truth template workbook")
    parser.add_argument("input", help="Path to the pipeline input Excel file")
    parser.add_argument(
        "pipeline_output", nargs="?", default="",
        help="Optional: path to a pipeline output workbook (to extract Acquire Log URLs)",
    )
    parser.add_argument("--output", default="", help="Path for the output template (optional)")
    args = parser.parse_args()

    pipeline_input = read_input(args.input)
    entities = pipeline_input.entities
    columns = pipeline_input.columns

    if not entities:
        print("  No entities found in input — nothing to generate.")
        return
    if not columns:
        print("  No questions found in input — nothing to generate.")
        return

    output_path = args.output
    if not output_path:
        base = os.path.splitext(os.path.basename(args.input))[0]
        output_path = os.path.join("outputs", f"{base}_ground_truth_template.xlsx")

    list_cols = [col.name for col in columns if _rows_for_column(col) == _ROWS_LIST]
    total_rows = sum(len(entities) * _rows_for_column(col) for col in columns)

    print(f"\n  Ground Truth Template")
    print(f"  input       : {args.input}")
    print(f"  entities    : {len(entities)}")
    print(f"  questions   : {len(columns)}")
    print(f"  list cols   : {list_cols or '(none)'}")
    print(f"  total rows  : {total_rows}")
    if args.pipeline_output:
        print(f"  acq source  : {args.pipeline_output}")
    print(f"  output      : {output_path}\n")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    gt_df = _make_ground_truth_df(entities, columns)
    instr_df = _make_instructions_df()
    source_urls_df = _make_source_urls_df(args.pipeline_output) if args.pipeline_output else None

    sheets: list[tuple[str, pd.DataFrame]] = [("Ground Truth", gt_df)]
    if source_urls_df is not None:
        sheets.append(("Source URLs Fetched", source_urls_df))
    sheets.append(("Instructions", instr_df))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        # Columns 1 (Claim ID), 2 (Entity), 3 (Question) are pre-populated on Ground Truth.
        prepop_cols = {1, 2, 3} if ws.title == "Ground Truth" else None
        _style_sheet(ws, _TAB_COLORS.get(ws.title, _HEADER_FILL), prepopulated_cols=prepop_cols)
    wb.save(output_path)

    print(f"  Template saved → {output_path}\n")


if __name__ == "__main__":
    main()
