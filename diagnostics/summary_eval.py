"""Faithfulness-eval harness for the LLM summary layer — validates the JUDGE.

Design: brain/proposals/llm-summary-layer.md §4 (labelled-pairs pattern).
Everything here scores the Tier-2 judge against labels KNOWN BY CONSTRUCTION
or supplied by a human, so the judge earns trust before its verdicts gate a
client deliverable. Pre-registered ship bar (George, 2026-07-06, held even
if the output reads well):

    corruption-set accuracy >= 0.90
    sentence-level agreement with human labels >= 0.80
    self-agreement >= 0.90

Subcommands (all need a pipeline workbook produced with SUMMARY_ENABLED=True
and DIAGNOSTICS=True — the Summary Log sheet carries the gate-passed prose,
input claim IDs and exact prompts this harness replays):

  positives       Judge the deterministic Digest lines (faithful by
                  construction, ~1 call each). The leading "N items across
                  M themes." prefix is template arithmetic, not claim
                  content, so only the claim-citing "Top: ..." remainder is
                  judged.
  corruptions     Programmatically corrupt gate-passed summaries (swap a
                  number, swap the entity, inject an unsupported fact,
                  re-attach citations, delete the top theme's sentence) and
                  check each corruption is caught. Deletion is caught by the
                  Tier-1 coverage gate, not the judge — scored accordingly.
  self-agreement  Judge every gate-passed summary twice; report
                  sentence-level agreement (expected ~1.0 with seeding) and
                  fingerprint variance.
  label-template  Export a ~50-summary sentence-level labelling workbook for
                  George (blind: no judge verdicts included).
  label-score     Re-judge the labelled sentences (seeded, deterministic)
                  and report agreement with George's labels.
  flags           Diagnosis view: print every flagged sentence of every
                  gate-passed summary next to the claims it cites (read-only;
                  seeded re-judge, so verdicts match summary_judge.py).

A judge failure never counts in the judge's favour: failed calls score as
misses. Usage (work laptop):
    python diagnostics/summary_eval.py corruptions --workbook outputs/run.xlsx
    python diagnostics/summary_eval.py label-template --workbook run.xlsx --out labels.xlsx
    python diagnostics/summary_eval.py label-score --workbook run.xlsx --labels labels.xlsx
"""
import argparse
import os
import random
import re
import sys

import openpyxl

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from diagnostics.summary_judge import (
    build_judge_prompt,
    col_index,
    judge_summary,
    load_claim_texts,
    parse_verdicts,
    sentences_with_claims,
)
from src.summarize import (
    _join_units,
    _split_sentences,
    azure_chat,
    cited_ids,
    has_citation,
    make_client,
    mechanical_gate,
)

_THEME_HEADER_RE = re.compile(r'^Theme "(?P<label>.*)" \(\d+ claims\):$')
_MEMBER_RE = re.compile(r"^\[(C\d{4,})\] ")


# ── Workbook loading ─────────────────────────────────────────────────────────

def load_passed_summaries(wb) -> list[dict]:
    """Gate-passed records from the Summary Log:
    {entity, question, summary, input_ids, prompt}."""
    if "Summary Log" not in wb.sheetnames:
        raise SystemExit("Workbook has no Summary Log sheet — re-run the "
                         "pipeline with SUMMARY_ENABLED=True and DIAGNOSTICS=True.")
    ws = wb["Summary Log"]
    cols = {name: col_index(ws, name) for name in
            ("Entity", "Question", "Gate", "Input Claim IDs", "Prompt", "Raw Response")}
    out = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=cols["Gate"]).value or "") != "pass":
            continue
        ids = str(ws.cell(row=r, column=cols["Input Claim IDs"]).value or "")
        out.append({
            "entity": str(ws.cell(row=r, column=cols["Entity"]).value or ""),
            "question": str(ws.cell(row=r, column=cols["Question"]).value or ""),
            "summary": str(ws.cell(row=r, column=cols["Raw Response"]).value or ""),
            "input_ids": {i.strip() for i in ids.split(",") if i.strip()},
            "prompt": str(ws.cell(row=r, column=cols["Prompt"]).value or ""),
        })
    return out


def parse_prompt_themes(prompt: str) -> list[tuple[str, set[str]]]:
    """Recover [(theme_label, shown_member_ids)] from a stored summarizer
    prompt — prompt order is size-desc, so [:3] are the coverage gate's
    top themes."""
    themes: list[tuple[str, set[str]]] = []
    current: set[str] | None = None
    for line in prompt.splitlines():
        header = _THEME_HEADER_RE.match(line.strip())
        if header:
            current = set()
            themes.append((header.group("label"), current))
            continue
        member = _MEMBER_RE.match(line.strip())
        if member and current is not None:
            current.add(member.group(1))
    return themes


# ── Corruption generators (labels known by construction) ─────────────────────

def _mask_citations(text: str) -> tuple[str, list[str]]:
    found: list[str] = []

    def keep(match):
        found.append(match.group(0))
        return f"\x00{len(found) - 1}\x00"

    return re.sub(r"\[C\d{4,}\]", keep, text), found


def _unmask(text: str, found: list[str]) -> str:
    for i, original in enumerate(found):
        text = text.replace(f"\x00{i}\x00", original)
    return text


def _sentence_of_offset(text: str, offset: int) -> int:
    """1-based sentence index containing a character offset."""
    for n, sentence in enumerate(_split_sentences(text), start=1):
        start = text.find(sentence)
        if start <= offset < start + len(sentence):
            return n
    return 1


def corrupt_swap_number(record: dict, entities: list[str], claim_texts: dict) -> tuple[str, set[int]] | None:
    masked, found = _mask_citations(record["summary"])
    match = re.search(r"\b\d+\b", masked)
    if not match:
        return None
    new_number = str(int(match.group(0)) + 7)
    corrupted = masked[: match.start()] + new_number + masked[match.end():]
    corrupted = _unmask(corrupted, found)
    return corrupted, {_sentence_of_offset(corrupted, match.start())}


def corrupt_swap_entity(record: dict, entities: list[str], claim_texts: dict) -> tuple[str, set[int]] | None:
    """Swap the entity name — but ONLY in a sentence whose cited claims
    actually NAME the entity, so the swap genuinely contradicts them.

    Claims here are subject/predicate-stripped attribute values ("own-product",
    "Lisses, France"), so most sentences cite claims that never mention the
    entity. The judge is contracted to check each sentence against its cited
    claims only, so swapping the entity in those sentences leaves them faithful
    to their claims — an invalid corruption whose "should be caught" label is
    wrong by construction (diagnosed 2026-07-08: 69/71 old swap_entity cases
    were no-ops, dragging the bar to 0.756 after the judge-blindness fix).
    Returns None when no cited claim names the entity."""
    entity = record["entity"]
    others = [e for e in entities if e and e != entity]
    if not others:
        return None
    replacement = others[0]
    sentences = _split_sentences(record["summary"])
    for n, sentence in enumerate(sentences, start=1):
        if entity not in sentence:
            continue
        if not any(entity.lower() in claim_texts.get(cid, "").lower()
                   for cid in cited_ids(sentence)):
            continue  # cited claims don't name the entity — swap wouldn't lie
        sentences[n - 1] = sentence.replace(entity, replacement)
        return _join_units(sentences, record["summary"]), {n}
    return None


def corrupt_inject_fact(record: dict, entities: list[str], claim_texts: dict) -> tuple[str, set[int]] | None:
    ids = sorted(record["input_ids"])
    if not ids:
        return None
    fabricated = (f"It also completed a full acquisition of its largest "
                  f"competitor for $9 billion [{ids[0]}].")
    # Append as a NEW unit in the summary's own shape. A space only works
    # when the summary ends with sentence punctuation (s3 prose); compact
    # one-liners like "yes [C0102]" have no terminal period, so a space
    # MERGES the fabrication into the same unit — the judge then sees one
    # half-supported half-fabricated unit instead of a clean planted lie
    # (2026-07-14 CMO s5 run: inject_fact misses clustered exactly on the
    # one-line tag cells for this reason — harness artefact, not judge).
    s = record["summary"].rstrip()
    sep = " " if ("\n" not in s and s.endswith((".", "!", "?"))) else "\n"
    corrupted = s + sep + fabricated
    return corrupted, {len(_split_sentences(corrupted))}


def corrupt_reattach_citation(record: dict, entities: list[str], claim_texts: dict) -> tuple[str, set[int]] | None:
    sentences = _split_sentences(record["summary"])
    cited = [(n, cited_ids(s)) for n, s in enumerate(sentences, start=1)]
    swappable = [(n, ids) for n, ids in cited if ids]
    if len(swappable) < 2 or set(swappable[0][1]) == set(swappable[1][1]):
        return None
    (n1, ids1), (n2, ids2) = swappable[0], swappable[1]

    def swap(sentence: str, old: list[str], new: list[str]) -> str:
        # Replace this sentence's citation set with the other sentence's.
        # Strips single-ID, multi-ID and comma-chained brackets alike.
        stripped = re.sub(r"\s*\[[^\[\]]*C\d{4,}[^\[\]]*\](,\s*\[[^\[\]]*C\d{4,}[^\[\]]*\])*", "", sentence)
        tail = " [" + ", ".join(new) + "]"
        return stripped.rstrip(".").rstrip() + tail + "."

    sentences[n1 - 1] = swap(sentences[n1 - 1], ids1, ids2)
    sentences[n2 - 1] = swap(sentences[n2 - 1], ids2, ids1)
    return _join_units(sentences, record["summary"]), {n1, n2}


_JUDGE_CORRUPTIONS = {
    "swap_number": corrupt_swap_number,
    "swap_entity": corrupt_swap_entity,
    "inject_fact": corrupt_inject_fact,
    "reattach_citation": corrupt_reattach_citation,
}


def corrupt_delete_top_sentence(record: dict) -> str | None:
    """Delete the first sentence citing the largest theme. Caught by the
    Tier-1 coverage gate (mechanical), not the judge."""
    themes = parse_prompt_themes(record["prompt"])
    if not themes:
        return None
    top_ids = themes[0][1]
    sentences = _split_sentences(record["summary"])
    keep = [s for s in sentences if not (set(cited_ids(s)) & top_ids)]
    if len(keep) == len(sentences) or not keep:
        return None
    return _join_units(keep, record["summary"])


# ── Legs ─────────────────────────────────────────────────────────────────────

def digest_judgeable_text(digest_line: str) -> str | None:
    """The claim-derived portion of a Digest line, with ALL template
    arithmetic removed. Only theme labels (verbatim claims) + citations
    remain — the part that is faithful-by-construction AGAINST THE CITED
    CLAIMS, which is the only contract the judge checks. The first laptop
    eval (2026-07-07) judged the per-theme "(9 items)" counts and correctly
    called them unsupported — a harness under-strip, not a judge error."""
    top = digest_line.find("Top: ")
    if top == -1 or not has_citation(digest_line):
        return None  # below-threshold cells carry no citations to judge
    return re.sub(r"\s*\(\d+\s+items?\)", "", digest_line[top:])


def run_positives(wb, client, limit) -> tuple[int, int]:
    """Digest lines are faithful by construction — the judge must agree.

    Each line is judged as ONE unit: it is a single template line, and theme
    labels are verbatim claims that may contain abbreviation periods, so
    sentence-splitting it produces citation-less fragments that auto-flag
    (the Sebia 3-fragment miss, 2026-07-07)."""
    if "Digest" not in wb.sheetnames:
        raise SystemExit("Workbook has no Digest sheet.")
    claim_texts = load_claim_texts(wb)
    ws = wb["Digest"]
    c_entity, c_question, c_digest = (col_index(ws, n) for n in ("Entity", "Question", "Digest"))

    items = []
    for r in range(2, ws.max_row + 1):
        judgeable = digest_judgeable_text(str(ws.cell(row=r, column=c_digest).value or ""))
        if judgeable is None:
            continue
        items.append((
            str(ws.cell(row=r, column=c_entity).value or ""),
            str(ws.cell(row=r, column=c_question).value or ""),
            judgeable,
        ))
    items = items[:limit] if limit else items
    print(f"[positives] judging {len(items)} digest lines...")

    correct = 0
    for entity, question, line in items:
        ids = cited_ids(line)
        one_unit = [{
            "n": 1,
            "sentence": line,
            "cited_ids": ids,
            "claims": {cid: claim_texts.get(cid, "(claim text not found)") for cid in ids},
        }]
        resp = azure_chat(client, build_judge_prompt(entity, question, one_unit))
        verdicts = None
        if not resp.get("error") and resp.get("text") is not None:
            verdicts = parse_verdicts(resp["text"], 1)
        ok = verdicts == {1: "faithful"}
        correct += ok
        if not ok:
            print(f"  MISS {entity} / {question}: {resp.get('error') or verdicts or resp.get('text', '')[:120]!r}")
    return correct, len(items)


def run_flags(wb, client, limit) -> None:
    """Diagnosis view: re-judge gate-passed summaries (seeded — same verdicts
    as summary_judge.py) and print every flagged sentence NEXT TO the claims
    it cites, so a human can tell over-strict judging from genuinely
    unsupported prose. Read-only; nothing is written."""
    claim_texts = load_claim_texts(wb)
    records = load_passed_summaries(wb)
    records = records[:limit] if limit else records
    print(f"[flags] re-judging {len(records)} gate-passed summaries for review...\n")

    n_flagged = 0
    for record in records:
        verdicts, _, error = judge_summary(
            client, record["entity"], record["question"], record["summary"], claim_texts)
        if verdicts is None:
            print(f"! {record['entity']} / {record['question']}: judge failed ({error})")
            continue
        flagged = sorted(n for n, v in verdicts.items() if v != "faithful")
        if not flagged:
            continue
        n_flagged += 1
        items = {i["n"]: i for i in sentences_with_claims(record["summary"], claim_texts)}
        print(f"=== {record['entity']} / {record['question']} ===")
        for n in flagged:
            item = items.get(n)
            if item is None:
                continue
            print(f"  [{verdicts[n].upper()}] sentence {n}: {item['sentence']}")
            for cid, text in item["claims"].items():
                print(f"      [{cid}] {text}")
        print()
    print(f"[flags] {n_flagged}/{len(records)} summaries have >=1 flagged sentence")


def run_corruptions(wb, client, limit) -> tuple[int, int, dict]:
    claim_texts = load_claim_texts(wb)
    records = load_passed_summaries(wb)
    records = records[:limit] if limit else records
    entities = sorted({r["entity"] for r in records})

    caught = total = 0
    by_type: dict[str, list[int]] = {}
    for record in records:
        for name, generator in _JUDGE_CORRUPTIONS.items():
            result = generator(record, entities, claim_texts)
            if result is None:
                continue
            corrupted, bad_sentences = result
            total += 1
            verdicts, _, error = judge_summary(
                client, record["entity"], record["question"], corrupted, claim_texts)
            hit = verdicts is not None and any(
                verdicts.get(n) != "faithful" for n in bad_sentences)
            caught += hit
            by_type.setdefault(name, []).append(int(hit))
            if not hit:
                print(f"  MISS [{name}] {record['entity']} / {record['question']}"
                      + (f" (judge error: {error})" if error else ""))

        # Deletion — Tier-1 coverage gate is the detector (design §4).
        deleted = corrupt_delete_top_sentence(record)
        if deleted is not None:
            total += 1
            reasons, _, _ = mechanical_gate(
                deleted, record["input_ids"], parse_prompt_themes(record["prompt"])[:3])
            hit = any("top theme not cited" in reason for reason in reasons)
            caught += hit
            by_type.setdefault("delete_top_sentence (gate)", []).append(int(hit))
            if not hit:
                print(f"  MISS [delete_top_sentence] {record['entity']} / {record['question']}")
    return caught, total, by_type


def run_self_agreement(wb, client, limit) -> tuple[int, int, set]:
    claim_texts = load_claim_texts(wb)
    records = load_passed_summaries(wb)
    records = records[:limit] if limit else records
    print(f"[self-agreement] judging {len(records)} summaries twice...")

    agree = total = 0
    fingerprints: set[str] = set()
    for record in records:
        runs = []
        for _ in range(2):
            verdicts, fp, _ = judge_summary(
                client, record["entity"], record["question"], record["summary"], claim_texts)
            if fp:
                fingerprints.add(fp)
            runs.append(verdicts)
        if runs[0] is None or runs[1] is None:
            # A failed run scores as full disagreement for this summary.
            n = len(_split_sentences(record["summary"]))
            total += n
            continue
        for n in runs[0]:
            total += 1
            agree += runs[0][n] == runs[1].get(n)
    return agree, total, fingerprints


def run_label_template(wb, out_path: str, n: int) -> None:
    claim_texts = load_claim_texts(wb)
    records = load_passed_summaries(wb)
    sample = records if len(records) <= n else random.Random(42).sample(records, n)

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Labels"
    ws.append(["Entity", "Question", "Sentence #", "Sentence", "Cited Claims",
               "Label (faithful / unsupported / contradicted)"])
    for record in sample:
        for item in sentences_with_claims(record["summary"], claim_texts):
            claims = "\n".join(f"[{cid}] {text}" for cid, text in item["claims"].items())
            ws.append([record["entity"], record["question"], item["n"],
                       item["sentence"], claims, ""])
    out.save(out_path)
    print(f"Labelling template: {len(sample)} summaries, {ws.max_row - 1} sentences "
          f"-> {out_path}")
    print("Fill the Label column (sentence-level), then run label-score.")


def run_label_score(wb, client, labels_path: str) -> tuple[int, int, int]:
    claim_texts = load_claim_texts(wb)
    lwb = openpyxl.load_workbook(labels_path)
    ws = lwb["Labels"]

    # Group labelled sentences back into summaries, preserving sentence order.
    summaries: dict[tuple[str, str], list[tuple[int, str, str]]] = {}
    for r in range(2, ws.max_row + 1):
        label = str(ws.cell(row=r, column=6).value or "").strip().lower()
        key = (str(ws.cell(row=r, column=1).value or ""), str(ws.cell(row=r, column=2).value or ""))
        summaries.setdefault(key, []).append(
            (int(ws.cell(row=r, column=3).value), str(ws.cell(row=r, column=4).value or ""), label))

    agree_binary = agree_exact = total = 0
    for (entity, question), rows in summaries.items():
        rows.sort()
        summary = " ".join(sentence for _, sentence, _ in rows)
        verdicts, _, error = judge_summary(client, entity, question, summary, claim_texts)
        for n, _, label in rows:
            if label not in {"faithful", "unsupported", "contradicted"}:
                continue  # unlabelled row
            total += 1
            judge = (verdicts or {}).get(n)  # judge failure -> disagreement
            agree_exact += judge == label
            agree_binary += (judge == "faithful") == (label == "faithful")
    return agree_binary, agree_exact, total


def _bar(name: str, value: float, threshold: float) -> str:
    return f"{name}: {value:.3f} (bar >= {threshold}) -> {'PASS' if value >= threshold else 'FAIL'}"


def main() -> int:
    ap = argparse.ArgumentParser(description="Faithfulness-eval harness (judge validation)")
    ap.add_argument("command", choices=["positives", "corruptions", "self-agreement",
                                        "label-template", "label-score", "flags"])
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--limit", type=int, default=None, help="cap summaries processed")
    ap.add_argument("--out", default="adlm-outputs/summary_labels.xlsx",
                    help="label-template output path")
    ap.add_argument("--labels", default=None, help="filled labelling workbook (label-score)")
    ap.add_argument("--n", type=int, default=50, help="label-template sample size")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)

    if args.command == "label-template":
        run_label_template(wb, args.out, args.n)
        return 0

    client = make_client()

    if args.command == "flags":
        run_flags(wb, client, args.limit)
    elif args.command == "positives":
        correct, total = run_positives(wb, client, args.limit)
        accuracy = correct / total if total else 0.0
        print(f"\n[positives] {correct}/{total} judged faithful — accuracy {accuracy:.3f}")
        print("(feeds the combined corruption-set bar; run `corruptions` too)")
    elif args.command == "corruptions":
        caught, total, by_type = run_corruptions(wb, client, args.limit)
        print("\n[corruptions] per type:")
        for name, hits in sorted(by_type.items()):
            print(f"  {name}: {sum(hits)}/{len(hits)}")
        accuracy = caught / total if total else 0.0
        print(_bar("[corruptions] catch rate", accuracy, 0.90))
        print("NOTE: the pre-registered bar is over the combined labelled-by-"
              "construction set — weight this with the `positives` leg.")
    elif args.command == "self-agreement":
        agree, total, fingerprints = run_self_agreement(wb, client, args.limit)
        rate = agree / total if total else 0.0
        print("\n" + _bar("[self-agreement] sentence-level", rate, 0.90))
        print(f"system_fingerprint(s): {sorted(fingerprints) or ['(none)']}"
              " — >1 fingerprint = backend drift, rerun before concluding")
    elif args.command == "label-score":
        if not args.labels:
            raise SystemExit("label-score needs --labels <filled template.xlsx>")
        agree_binary, agree_exact, total = run_label_score(wb, client, args.labels)
        if not total:
            raise SystemExit("No labelled rows found — fill the Label column first.")
        print("\n" + _bar("[label] binary agreement (faithful vs not)",
                          agree_binary / total, 0.80))
        print(f"[label] exact 3-way agreement: {agree_exact / total:.3f} ({total} sentences)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
