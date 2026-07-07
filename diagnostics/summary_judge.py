"""Tier-2 faithfulness judge for the AI Summary sheet (post-run pass).

Design: brain/proposals/llm-summary-layer.md §4. The judge is diagnostics,
not deliverable — it runs AFTER the pipeline (semantic-verify Phase B
pattern) and only touches the Faithfulness column of the AI Summary sheet.

Per gate-passed summary (Faithfulness == "not-assessed"), one Azure call:
each sentence plus the FULL TEXT of the claims it cites (from Provenance),
strict-JSON verdict per sentence — "faithful" | "unsupported" |
"contradicted". Catches what the Tier-1 mechanical gate can't: a sentence
citing real IDs while asserting something they don't say.

Verdict written to the Summary Log's Faithfulness column (the audit surface —
the AI Summary sheet is matrix-shaped for consultants, George 2026-07-07):
  - all sentences faithful          -> "faithful"
  - >=1 flagged                     -> "N flagged sentence(s)" + the AI
    Summary matrix cell gets an orange fill and a visible
    "[faithfulness: ...]" marker line
  - judge call/parse failure        -> stays "not-assessed" — NEVER a pass
    (semantic-verify principle 1: a broken judge must not look like passing).

Needs a workbook produced with DIAGNOSTICS=True (the Summary Log carries the
gate state and raw prose this judge reads). Calls are temperature=0 +
SUMMARY_SEED via src.summarize.azure_chat, so re-runs are deterministic to
the extent the deployment honours seeding (confirmed 2026-07-07 probe);
fingerprints are printed for drift visibility.

Usage (work laptop — needs AZURE_API_KEY):
    python diagnostics/summary_judge.py --workbook outputs/run.xlsx [--out judged.xlsx] [--limit N]
"""
import argparse
import json
import os
import re
import sys

import openpyxl

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from src.extract import _strip_json_fence
from src.summarize import _split_sentences, azure_chat, make_client

JUDGE_PROMPT_VERSION = "j1"

_CITATION_RE = re.compile(r"\[(C\d{4,})\]")
_VERDICTS = {"faithful", "unsupported", "contradicted"}
NOT_ASSESSED = "not-assessed"


# ── Workbook readers (shared with diagnostics/summary_eval.py) ───────────────

def col_index(ws, header_prefix: str) -> int:
    """1-based column whose header starts with header_prefix (the AI Summary
    sheet's Summary header carries a disclaimer suffix)."""
    for cell in ws[1]:
        if str(cell.value or "").strip().lower().startswith(header_prefix.lower()):
            return cell.column
    raise ValueError(f"{ws.title!r} has no column starting with {header_prefix!r}")


def load_claim_texts(wb) -> dict[str, str]:
    """Provenance Claim ID -> claim value text (what the summarizer saw)."""
    ws = wb["Provenance"]
    c_id = col_index(ws, "Claim ID")
    c_claim = col_index(ws, "Claim")
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=c_id).value
        if cid is not None:
            out[str(cid).strip()] = str(ws.cell(row=r, column=c_claim).value or "")
    return out


def sentences_with_claims(summary: str, claim_texts: dict[str, str]) -> list[dict]:
    """[{n, sentence, cited_ids, claims: {id: text}}] per sentence."""
    out = []
    for n, sentence in enumerate(_split_sentences(summary or ""), start=1):
        ids = _CITATION_RE.findall(sentence)
        out.append({
            "n": n,
            "sentence": sentence,
            "cited_ids": ids,
            "claims": {cid: claim_texts.get(cid, "(claim text not found)") for cid in ids},
        })
    return out


# ── Judge call ────────────────────────────────────────────────────────────────

def build_judge_prompt(entity: str, question: str, items: list[dict]) -> str:
    blocks = []
    for item in items:
        lines = [f'Sentence {item["n"]}: "{item["sentence"]}"', "Cited claims:"]
        lines += [f"  [{cid}] {text}" for cid, text in item["claims"].items()]
        blocks.append("\n".join(lines))
    return (
        f"You are checking whether each numbered sentence of a summary about "
        f'{entity} (question: "{question}") is supported by the claims it '
        "cites. Judge each sentence ONLY against its own cited claims.\n\n"
        + "\n\n".join(blocks)
        + "\n\nFor each sentence return exactly one verdict:\n"
        '- "faithful": everything the sentence asserts is supported by its cited claims\n'
        '- "unsupported": the sentence asserts something its cited claims do not say\n'
        '- "contradicted": the sentence conflicts with its cited claims\n\n'
        "Respond with ONLY strict JSON mapping sentence numbers to verdicts, "
        'e.g. {"1": "faithful", "2": "unsupported"}.'
    )


def parse_verdicts(raw: str, n_sentences: int) -> dict[int, str] | None:
    """Strict parse: every sentence judged, every verdict in-vocabulary —
    anything less is a judge failure (None), never a partial pass."""
    try:
        data = json.loads(_strip_json_fence(raw))
    except Exception:
        return None
    if not isinstance(data, dict):
        return None
    verdicts: dict[int, str] = {}
    for key, value in data.items():
        try:
            n = int(key)
        except (TypeError, ValueError):
            return None
        if not isinstance(value, str) or value.strip().lower() not in _VERDICTS:
            return None
        verdicts[n] = value.strip().lower()
    if set(verdicts) != set(range(1, n_sentences + 1)):
        return None
    return verdicts


def judge_summary(
    client,
    entity: str,
    question: str,
    summary: str,
    claim_texts: dict[str, str],
) -> tuple[dict[int, str] | None, str | None, str | None]:
    """One judged summary. Returns (verdicts|None, system_fingerprint, error).
    A sentence with no citations is flagged mechanically (the Tier-1 gate
    should have caught it; the judge never sees it as judgeable)."""
    items = sentences_with_claims(summary, claim_texts)
    if not items:
        return None, None, "no sentences"
    uncited = [i for i in items if not i["cited_ids"]]
    if uncited:
        verdicts = {i["n"]: "unsupported" for i in uncited}
        verdicts.update({i["n"]: "faithful" for i in items if i["cited_ids"]})
        # Mechanical flag only — don't spend a call on a gate-escaped summary.
        return verdicts, None, None

    resp = azure_chat(client, build_judge_prompt(entity, question, items))
    if resp.get("error") or resp.get("text") is None:
        return None, resp.get("system_fingerprint"), resp.get("error") or "no response"
    verdicts = parse_verdicts(resp["text"], len(items))
    if verdicts is None:
        return None, resp.get("system_fingerprint"), f"unparseable verdict: {resp['text'][:200]!r}"
    return verdicts, resp.get("system_fingerprint"), None


def verdict_to_cell(verdicts: dict[int, str] | None) -> str:
    if verdicts is None:
        return NOT_ASSESSED
    flagged = sum(1 for v in verdicts.values() if v != "faithful")
    return "faithful" if flagged == 0 else f"{flagged} flagged sentence(s)"


# ── CLI ───────────────────────────────────────────────────────────────────────

def annotate_matrix_cell(wb, entity: str, question: str, verdict: str) -> None:
    """Flag a judged-unfaithful summary on the consultant-facing AI Summary
    matrix: orange fill + a visible marker line appended to the prose. The
    judge reads the Summary Log's Raw Response, never this cell, so the
    marker can't feed back into a re-judge."""
    from openpyxl.styles import PatternFill

    if "AI Summary" not in wb.sheetnames:
        return
    ws = wb["AI Summary"]
    entity_rows = {str(ws.cell(row=r, column=1).value or ""): r
                   for r in range(2, ws.max_row + 1)}
    question_cols = {str(cell.value or ""): cell.column for cell in ws[1]}
    r, c = entity_rows.get(entity), question_cols.get(question)
    if r is None or c is None:
        return
    cell = ws.cell(row=r, column=c)
    cell.value = f"{cell.value or ''}\n[faithfulness: {verdict} — see Summary Log]"
    cell.fill = PatternFill("solid", fgColor="FFE0B2")  # io_excel._ORANGE_FILL


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--out", default=None, help="output path (default: in place)")
    ap.add_argument("--limit", type=int, default=None, help="judge at most N summaries")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    if "Summary Log" not in wb.sheetnames:
        print("No Summary Log sheet — run the pipeline with SUMMARY_ENABLED=true "
              "and DIAGNOSTICS=True first.")
        return 1

    client = make_client()
    claim_texts = load_claim_texts(wb)

    log = wb["Summary Log"]
    cols = {name: col_index(log, name) for name in
            ("Entity", "Question", "Gate", "Faithfulness", "Raw Response")}
    # Per-sentence verdict JSON — for diagnosis (summary_eval.py flags) and
    # the label-agreement audit trail. Older workbooks lack the column.
    try:
        c_verdicts = col_index(log, "Judge Verdicts")
    except ValueError:
        c_verdicts = log.max_column + 1
        log.cell(row=1, column=c_verdicts).value = "Judge Verdicts"

    todo = [
        r for r in range(2, log.max_row + 1)
        if str(log.cell(row=r, column=cols["Gate"]).value or "") == "pass"
        and str(log.cell(row=r, column=cols["Faithfulness"]).value or "") == NOT_ASSESSED
    ]
    if args.limit is not None:
        todo = todo[: args.limit]
    print(f"Judging {len(todo)} gate-passed summaries (prompt {JUDGE_PROMPT_VERSION}, "
          f"one call each)...")

    fingerprints: set[str] = set()
    counts = {"faithful": 0, "flagged": 0, "not-assessed": 0}
    for r in todo:
        entity = str(log.cell(row=r, column=cols["Entity"]).value or "")
        question = str(log.cell(row=r, column=cols["Question"]).value or "")
        summary = str(log.cell(row=r, column=cols["Raw Response"]).value or "")
        verdicts, fp, error = judge_summary(client, entity, question, summary, claim_texts)
        if fp:
            fingerprints.add(fp)
        cell_value = verdict_to_cell(verdicts)
        log.cell(row=r, column=cols["Faithfulness"]).value = cell_value
        if verdicts is not None:
            log.cell(row=r, column=c_verdicts).value = json.dumps(verdicts, sort_keys=True)
        if cell_value == "faithful":
            counts["faithful"] += 1
        elif cell_value == NOT_ASSESSED:
            counts["not-assessed"] += 1
            print(f"  ! {entity} / {question}: judge failed ({error}) — left {NOT_ASSESSED}")
        else:
            counts["flagged"] += 1
            annotate_matrix_cell(wb, entity, question, cell_value)
            print(f"  - {entity} / {question}: {cell_value}")

    out_path = args.out or args.workbook
    wb.save(out_path)
    print(f"\nfaithful: {counts['faithful']}  flagged: {counts['flagged']}  "
          f"not-assessed: {counts['not-assessed']}")
    print(f"system_fingerprint(s) seen: {sorted(fingerprints) or ['(none)']}")
    print(f"Saved -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
