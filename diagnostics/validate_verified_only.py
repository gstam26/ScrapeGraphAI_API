"""Validate the verified-only invariants of an output workbook.

Mechanises the 2026-07-05 audit's validation criteria (gap #1) plus the
2026-07-06 enforcement checks, so a validation pass is one command instead
of a hand audit:

  1. Every Claim ID cited in Grouped Themes ("Claim IDs" column and inline
     [C####] bullet tags) resolves to a Provenance row with Verified=True.
  2. Every Grouped Themes Theme-cell hyperlink targets a Provenance row with
     Verified=True (the anchor is a verified claim).
  3. Every [C####] cited in Digest text resolves to Verified=True.
  4. Every Provenance row with Verified=False carries the orange
     analyst-review fill on its Verified cell.
  5. Optional (--baseline): Matrix diff against a baseline workbook.
     ONLY meaningful for page-set-pinned replays (build_replay_input.py) —
     across crawled runs the page sets differ and a Matrix diff confounds
     code changes with crawl drift (decision-log 2026-07-06 finding).
     Additions can still be legitimate on a replay: baseline chunks that
     returned empty or timed out were never extract-cached and re-fire live,
     so a previously-502'd chunk may now contribute data.

Exit 0 = all checks pass. Exit 1 = violations found (each printed).

Usage:
    python diagnostics/validate_verified_only.py --workbook outputs/replay_run_2026-07-06.xlsx \
        [--baseline adlm-outputs/validation_sample_run_2026-07-03.xlsx]
"""
import argparse
import re
import sys

import openpyxl

_CLAIM_ID_RE = re.compile(r"\bC\d{4,}\b")
_ORANGE = "FFE0B2"  # io_excel._ORANGE_FILL — the analyst-review flag


def _col_index(ws, header_name: str) -> int:
    for cell in ws[1]:
        if str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    raise ValueError(f"{ws.title!r} has no {header_name!r} column")


def _load_provenance(wb) -> tuple[dict, dict]:
    """Return (claim_id -> verified, excel_row -> verified) for Provenance."""
    ws = wb["Provenance"]
    c_id = _col_index(ws, "Claim ID")
    c_ver = _col_index(ws, "Verified")
    by_id: dict[str, bool] = {}
    by_row: dict[int, bool] = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=c_id).value
        if cid is None:
            continue
        verified = bool(ws.cell(row=r, column=c_ver).value)
        by_id[str(cid).strip()] = verified
        by_row[r] = verified
    return by_id, by_row


def check_workbook(path: str, baseline_path: str | None = None) -> list[str]:
    """Run all checks; return a list of violation strings (empty = pass)."""
    wb = openpyxl.load_workbook(path)
    violations: list[str] = []

    if "Provenance" not in wb.sheetnames:
        return [f"{path}: no Provenance sheet — nothing to validate"]
    by_id, by_row = _load_provenance(wb)

    # 1 + 2: Grouped Themes citations and anchors.
    if "Grouped Themes" in wb.sheetnames:
        ws = wb["Grouped Themes"]
        c_theme = _col_index(ws, "Theme")
        c_values = _col_index(ws, "Values")
        c_ids = _col_index(ws, "Claim IDs")
        for r in range(2, ws.max_row + 1):
            where = f"Grouped Themes row {r}"
            ids_text = str(ws.cell(row=r, column=c_ids).value or "")
            bullets_text = str(ws.cell(row=r, column=c_values).value or "")
            for cid in _CLAIM_ID_RE.findall(ids_text) + _CLAIM_ID_RE.findall(bullets_text):
                if cid not in by_id:
                    violations.append(f"{where}: cites {cid} which is not in Provenance")
                elif not by_id[cid]:
                    violations.append(f"{where}: cites UNVERIFIED claim {cid}")
            link = ws.cell(row=r, column=c_theme).hyperlink
            if link is None:
                violations.append(f"{where}: Theme cell has no Provenance hyperlink")
            else:
                m = re.fullmatch(r"#Provenance!A(\d+)", str(link.target or ""))
                if not m:
                    violations.append(f"{where}: unexpected Theme link target {link.target!r}")
                elif not by_row.get(int(m.group(1)), False):
                    violations.append(
                        f"{where}: Theme anchors on UNVERIFIED Provenance row {m.group(1)}"
                    )
    else:
        print("  (no Grouped Themes sheet — grouping was off/skipped; checks 1-2 skipped)")

    # 3: Digest citations.
    if "Digest" in wb.sheetnames:
        ws = wb["Digest"]
        c_digest = _col_index(ws, "Digest")
        for r in range(2, ws.max_row + 1):
            text = str(ws.cell(row=r, column=c_digest).value or "")
            for cid in _CLAIM_ID_RE.findall(text):
                if not by_id.get(cid, False):
                    violations.append(f"Digest row {r}: cites unverified/unknown claim {cid}")

    # 4: orange review flag on every Verified=False Provenance row.
    ws = wb["Provenance"]
    c_ver = _col_index(ws, "Verified")
    for r, verified in by_row.items():
        if verified:
            continue
        fill = ws.cell(row=r, column=c_ver).fill
        rgb = str(getattr(fill.start_color, "rgb", "") or "")
        if _ORANGE not in rgb:
            violations.append(
                f"Provenance row {r}: Verified=False but no orange review flag (fill {rgb!r})"
            )

    # 5: optional Matrix diff (replay runs only — see module docstring).
    if baseline_path:
        base = openpyxl.load_workbook(baseline_path)
        cur_m, base_m = wb["Matrix"], base["Matrix"]

        def _matrix_cells(m):
            headers = [str(c.value) for c in m[1]]
            out = {}
            for r in range(2, m.max_row + 1):
                entity = m.cell(row=r, column=1).value
                for c in range(2, m.max_column + 1):
                    out[(str(entity), headers[c - 1])] = str(m.cell(row=r, column=c).value or "")
            return out

        cur, prev = _matrix_cells(cur_m), _matrix_cells(base_m)
        changed = [k for k in prev if k in cur and cur[k] != prev[k]]
        missing = [k for k in prev if k not in cur]
        added = [k for k in cur if k not in prev]
        print(
            f"  Matrix vs baseline: {len(prev) - len(changed) - len(missing)} identical, "
            f"{len(changed)} changed, {len(missing)} missing, {len(added)} new cells"
        )
        for k in changed:
            print(f"    changed: {k[0]} / {k[1]}")
        for k in missing:
            violations.append(f"Matrix cell missing vs baseline: {k[0]} / {k[1]}")
        base.close()

    wb.close()
    return violations


def main() -> int:
    ap = argparse.ArgumentParser(description="Validate verified-only output invariants")
    ap.add_argument("--workbook", required=True, help="output .xlsx to validate")
    ap.add_argument("--baseline", default=None,
                    help="baseline .xlsx for a Matrix diff (pinned replays only)")
    args = ap.parse_args()

    violations = check_workbook(args.workbook, args.baseline)
    if violations:
        print(f"\nFAIL — {len(violations)} violation(s):")
        for v in violations:
            print(f"  - {v}")
        return 1
    print("\nPASS — all verified-only invariants hold.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
