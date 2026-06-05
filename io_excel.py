import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import ColumnSpec, PipelineResult

# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FILL   = "2E4057"
_HEADER_FONT   = "FFFFFF"
_ALT_ROW       = "F5F5F5"
_RED_FILL      = "FFCDD2"
_RED_FONT      = "C62828"
_ORANGE_FILL   = "FFE0B2"
_LORANGE_FILL  = "FFF3E0"   # light orange for mixed verified/unverified

_TAB_COLORS = {
    "Summary":          "2E4057",
    "Matrix":           "4CAF50",
    "Provenance":       "009688",
    "Acquire Log":      "FF9800",
    "Crawl Candidates": "FF9800",
    "Extract Log":      "F44336",
    "Verify Log":       "9C27B0",
}


# ── Public helpers (used by main.py) ─────────────────────────────────────────

def read_urls_from_excel(filepath: str) -> list[tuple[str, int]]:
    """Read entity URLs (and optional per-URL depth) from input Excel file.

    Returns list of (url, depth) tuples. depth defaults to 0 when the column
    is absent. depth > 0 enables guided crawling for that URL.
    """
    df = pd.read_excel(filepath)

    url_col = next(
        (col for col in df.columns if "url" in col.lower() or "link" in col.lower()),
        df.columns[0],
    )
    depth_col = next(
        (col for col in df.columns if col.lower() == "depth"),
        None,
    )

    urls = df[url_col].dropna().astype(str).tolist()

    if depth_col is not None:
        depths = df.loc[df[url_col].notna(), depth_col].fillna(0).astype(int).tolist()
    else:
        depths = [0] * len(urls)

    return list(zip(urls, depths))


def parse_columns(raw_columns: list[str]) -> list[ColumnSpec]:
    """Parse column specs from user input."""
    columns = []
    for raw in raw_columns:
        if ":" in raw:
            name, instruction = raw.split(":", 1)
            columns.append(ColumnSpec(name=name.strip(), instruction=instruction.strip()))
        else:
            columns.append(ColumnSpec(name=raw.strip()))
    return columns


# ── DataFrame builders ───────────────────────────────────────────────────────

def _make_df(rows: list[dict], keys: list[str], col_names: list[str]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=col_names)
    records = [{col: row.get(key, "") for col, key in zip(col_names, keys)} for row in rows]
    return pd.DataFrame(records, columns=col_names)


def _make_summary_df(diag: dict | None, result: PipelineResult) -> pd.DataFrame:
    col_names = [
        "Entity URL", "Pages Fetched", "Pages Crawled", "Total Claims Found",
        "Claims Verified", "Claims Unverified", "Cells With No Data",
        "Total Fetch Time", "Total Extract Time", "Acquire Tool Used", "Extract Tool Used",
    ]
    keys = [
        "entity_url", "pages_fetched", "pages_crawled", "total_claims_found",
        "claims_verified", "claims_unverified", "cells_with_no_data",
        "total_fetch_time", "total_extract_time", "acquire_tool_used", "extract_tool_used",
    ]
    if diag and diag.get("summary"):
        return _make_df(diag["summary"], keys, col_names)

    # Fallback when diag not available
    rows = []
    for row in result.rows:
        src = row.all_cells if row.all_cells else row.cells
        all_ev = [e for c in src for e in c.evidence]
        rows.append({
            "Entity URL": row.entity_url,
            "Pages Fetched": len(src),
            "Pages Crawled": "",
            "Total Claims Found": len(all_ev),
            "Claims Verified": sum(1 for e in all_ev if e.verified),
            "Claims Unverified": sum(1 for e in all_ev if not e.verified),
            "Cells With No Data": sum(1 for c in src if not c.evidence),
            "Total Fetch Time": "",
            "Total Extract Time": "",
            "Acquire Tool Used": "",
            "Extract Tool Used": "",
        })
    return pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)


def _match_type_str(verified: bool, score) -> str:
    if score is None or score == "":
        return "none"
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "none"
    if s >= 100:
        return "exact"
    if verified:
        return "fuzzy"
    return "none"


def _make_matrix_df(
    result: PipelineResult,
    columns: list[ColumnSpec],
) -> tuple[pd.DataFrame, dict]:
    """Returns (DataFrame, cell_fills) where cell_fills maps (row, col) → hex color."""
    col_order = ["Entity URL"] + [c.name for c in columns]
    matrix_rows = []
    fills: dict[tuple[int, int], str] = {}

    for row_data_idx, row in enumerate(result.rows):
        excel_row = row_data_idx + 2  # 1-indexed + header
        src = row.all_cells if row.all_cells else row.cells
        matrix_row: dict = {"Entity URL": row.entity_url}

        for col_excel_idx, col in enumerate(columns, start=2):
            verified_vals: list[str] = []
            unverified_vals: list[str] = []
            seen: set[str] = set()

            for cell in src:
                if cell.column != col.name:
                    continue
                for ev in cell.evidence:
                    val_str = str(ev.value).strip() if ev.value is not None else ""
                    if not val_str or val_str in seen:
                        continue
                    seen.add(val_str)
                    if ev.verified:
                        verified_vals.append(val_str)
                    else:
                        unverified_vals.append(val_str)

            if not verified_vals and not unverified_vals:
                text = "No data found"
                fills[(excel_row, col_excel_idx)] = _RED_FILL
            elif not verified_vals:
                lines = ["• " + v for v in unverified_vals] + ["(unverified)"]
                text = "\n".join(lines)
                fills[(excel_row, col_excel_idx)] = _ORANGE_FILL
            elif unverified_vals:
                lines = ["• " + v for v in verified_vals]
                lines += ["— Unverified —"] + ["• " + v for v in unverified_vals]
                text = "\n".join(lines)
                fills[(excel_row, col_excel_idx)] = _LORANGE_FILL
            else:
                text = "\n".join("• " + v for v in verified_vals)

            matrix_row[col.name] = text

        matrix_rows.append(matrix_row)

    df = pd.DataFrame(matrix_rows, columns=col_order) if matrix_rows else pd.DataFrame(columns=col_order)
    return df, fills


def _make_provenance_df(
    result: PipelineResult,
    columns: list[ColumnSpec],
    diag: dict | None,
) -> pd.DataFrame:
    col_names = [
        "Entity URL", "Source URL", "Question", "Claim", "Verbatim Quote",
        "Verified", "Verification Score", "Match Type", "Source Page Depth",
    ]

    url_to_depth: dict[str, int] = {}
    if diag:
        for r in diag.get("acquire_log", []):
            url_to_depth[r.get("page_url", "")] = r.get("depth", 0)

    rows = []
    for entity_row in result.rows:
        src = entity_row.all_cells if entity_row.all_cells else entity_row.cells
        for cell in src:
            for ev in cell.evidence:
                if ev.value is None:
                    continue
                score = ev.verification_score
                verified = ev.verified
                rows.append({
                    "Entity URL": entity_row.entity_url,
                    "Source URL": cell.source_url,
                    "Question": cell.column,
                    "Claim": str(ev.value),
                    "Verbatim Quote": ev.quote or "",
                    "Verified": verified,
                    "Verification Score": round(score, 1) if isinstance(score, float) else "",
                    "Match Type": _match_type_str(verified, score),
                    "Source Page Depth": url_to_depth.get(cell.source_url, 0),
                })

    return pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)


# ── Workbook formatting ───────────────────────────────────────────────────────

def _style_sheet(ws, tab_color: str, matrix_fills: dict | None = None) -> None:
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"

    # Header row
    h_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    h_font = Font(name="Arial", bold=True, color=_HEADER_FONT, size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
    ws.row_dimensions[1].height = 28

    # Which columns have "Verified" in header (for FALSE → orange rule)
    verified_cols: set[int] = set()
    for cell in ws[1]:
        if cell.value and "verified" in str(cell.value).lower():
            verified_cols.add(cell.column)

    # Body rows: font + alignment + alt shading
    b_font = Font(name="Arial", size=10)
    b_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alt = PatternFill("solid", fgColor=_ALT_ROW)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = b_font
            cell.alignment = b_align
            if row_idx % 2 == 0:
                cell.fill = alt

    # Apply matrix-specific fills (override alt shading)
    if matrix_fills:
        for (r, c), hex_color in matrix_fills.items():
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=hex_color)
            if hex_color == _RED_FILL:
                cell.font = Font(name="Arial", size=10, color=_RED_FONT)

    # Text-based conditional fills
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            vl = val.lower()

            if "no data found" in vl:
                cell.fill = PatternFill("solid", fgColor=_RED_FILL)
                cell.font = Font(name="Arial", size=10, color=_RED_FONT)
            elif cell.column in verified_cols and val == "FALSE":
                cell.fill = PatternFill("solid", fgColor=_ORANGE_FILL)
            elif any(kw in vl for kw in ("timed out", "timeout", "status: error")):
                cell.fill = PatternFill("solid", fgColor=_RED_FILL)

    # Column widths (max 60)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                longest = max((len(ln) for ln in str(cell.value).split("\n")), default=0)
                max_len = max(max_len, longest)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    # Row heights for matrix (multi-line bullet cells)
    if ws.title == "Matrix":
        for row_idx in range(2, ws.max_row + 1):
            max_lines = 1
            for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
                if col[0].value:
                    max_lines = max(max_lines, str(col[0].value).count("\n") + 1)
            ws.row_dimensions[row_idx].height = min(max(max_lines * 14, 16), 200)


# ── Main entry point ─────────────────────────────────────────────────────────

def write_output_excel(
    result: PipelineResult,
    columns: list[ColumnSpec],
    output_path: str,
    diag: dict | None = None,
) -> None:
    try:
        from config import DIAGNOSTICS
    except Exception:
        DIAGNOSTICS = False

    write_diag = DIAGNOSTICS and diag is not None

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    summary_df                   = _make_summary_df(diag, result)
    matrix_df, matrix_fills      = _make_matrix_df(result, columns)
    provenance_df                = _make_provenance_df(result, columns, diag)

    acq_col_keys = ["entity_url","page_url","parent_url","depth","crawl_score","above_threshold",
                    "fetch_tool","page_length","fetch_time_ms","from_cache","status","skip_reason"]
    acq_col_names = ["Entity URL","Page URL","Parent URL","Depth","Crawl Score","Above Threshold",
                     "Fetch Tool","Page Length (chars)","Fetch Time (ms)","From Cache","Status","Skip Reason"]

    cand_col_keys  = ["parent_url","candidate_url","anchor_text","url_path","crawl_score",
                      "threshold","followed","skip_reason"]
    cand_col_names = ["Parent URL","Candidate URL","Anchor Text","URL Path","Crawl Score",
                      "Threshold","Followed","Skip Reason"]

    ext_col_keys  = ["entity_url","source_url","question","extract_tool","items_extracted",
                     "extraction_time_ms","timed_out","retry_count","page_length_input","raw_answer_preview"]
    ext_col_names = ["Entity URL","Source URL","Question","Extract Tool","Items Extracted",
                     "Extraction Time (ms)","Timed Out","Retry Count","Page Length Input (chars)","Raw Answer Preview"]

    ver_col_keys  = ["entity_url","source_url","question","claim_preview","quote_preview",
                     "verified","match_type","verification_score","verifier_tool"]
    ver_col_names = ["Entity URL","Source URL","Question","Claim Preview","Quote Preview",
                     "Verified","Match Type","Verification Score","Verifier Tool"]

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Summary",    summary_df),
        ("Matrix",     matrix_df),
        ("Provenance", provenance_df),
    ]

    if write_diag:
        sheets += [
            ("Acquire Log",      _make_df(diag.get("acquire_log", []),      acq_col_keys,  acq_col_names)),
            ("Crawl Candidates", _make_df(diag.get("crawl_candidates", []), cand_col_keys, cand_col_names)),
            ("Extract Log",      _make_df(diag.get("extract_log", []),      ext_col_keys,  ext_col_names)),
            ("Verify Log",       _make_df(diag.get("verify_log", []),       ver_col_keys,  ver_col_names)),
        ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply formatting
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        fills = matrix_fills if ws.title == "Matrix" else None
        _style_sheet(ws, _TAB_COLORS.get(ws.title, _HEADER_FILL), matrix_fills=fills)
    wb.save(output_path)
