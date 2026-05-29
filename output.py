import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def format_value(value) -> str:
    if isinstance(value, list):
        cleaned = [str(item).strip() for item in value if item not in (None, "")]
        if len(cleaned) == 0:
            return ""
        if len(cleaned) == 1:
            return cleaned[0]
        return "\n".join(f"• {item}" for item in cleaned)
    return value if value is not None else ""


def write_output_excel(results: list[dict], columns: list[str], output_path: str):
    all_columns = ["URL"] + columns
    formatted = [
        {k: format_value(v) for k, v in row.items()}
        for row in results
    ]
    df = pd.DataFrame(formatted, columns=all_columns)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2E4057")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.column_dimensions["A"].width = 40
    for i in range(2, len(all_columns) + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = 30

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = left

    ws.freeze_panes = "A2"
    wb.save(output_path)
