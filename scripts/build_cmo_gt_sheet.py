"""Build the CMO ground-truth handoff pair from ONE canonical question spec.

Emits two workbooks that must never drift apart:

  1. cmo-inputs/cmo_gt_analyst_TEMPLATE.xlsx — the fillable matrix for the
     analyst (Caitlin): one row per crawlable entity, verified seed URL
     prefilled, one column per question, an answer-format guidance row, an
     Instructions tab (GT scope = website only, per 2026-07-22 decision) and
     an Excluded-companies tab (George's EDIT verdicts — do-not-research).
  2. cmo-inputs/cmo_input_v2.xlsx — the pipeline input workbook with the SAME
     question text (+ the same guidance as the extractor `instructions`
     column). The v1 baseline used the 15 verbatim client headers; v2 fixes
     typos and splits the two compound questions (independence/acquirer,
     volume/low-volume), so the eval aligns GT columns to pipeline columns
     1:1. A v2 pipeline run is therefore required before scoring against the
     analyst GT.

Question-text edits vs the client sheet (flag these back to the client):
  * "plastic circuit board (PCB)" -> "printed circuit board (PCB)" — PCB is
    printed; the extractor would chase a term that barely exists on the web.
  * Q2 split into independence (Yes/No) + acquirer (n/a when independent —
    "n/a" is a gt_convert null marker, so independent rows score as a clean
    true negative rather than a hallucination hit).
  * The USA>Canada&Mexico>… preference ranking is the client's downstream
    scoring rubric, not an extraction instruction — dropped from Q4.
  * Q14 split into stated volume (verbatim wording) + low-volume Yes/No.

The filled template converts with:
  python src/eval/gt_convert.py cmo_gt_analyst_TEMPLATE.xlsx \
      --output cmo_ground_truth.xlsx --sheet "Ground Truth" \
      --ignore-cols "Website,Notes / sources,Date checked"
(the guidance row has an empty entity cell, which gt_convert already skips).

Usage (from repo root):
    python scripts/build_cmo_gt_sheet.py
"""
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IN_DIR = "cmo-inputs"
SOURCE_WB = os.path.join(IN_DIR, "cmo_input.xlsx")
SOURCE_INV = os.path.join(IN_DIR, "cmo_url_inventory.csv")
OUT_TEMPLATE = os.path.join(IN_DIR, "cmo_gt_analyst_TEMPLATE.xlsx")
OUT_INPUT_V2 = os.path.join(IN_DIR, "cmo_input_v2.xlsx")

YN = ("Yes / No / Not disclosed. No = the site says or clearly implies they "
      "don't; Not disclosed = the site is silent. If torn, Not disclosed.")

# (question text, guidance) — guidance is BOTH the extractor instruction in
# cmo_input_v2.xlsx and the analyst's answer-format row: same contract on
# both sides is what makes the comparison fair.
QUESTIONS = [
    ("Summary description of company's services",
     "1-3 sentences describing what the company offers, based on the website."),
    ("Is the company still operating independently?",
     "Answer Yes, No, or Not disclosed. Answer No if the company has been "
     "acquired, merged, or is now a division of another company."),
    ("If not operating independently: who acquired or absorbed the company?",
     "Name of the acquiring/parent company. Write n/a if the company is "
     "independent; write Not disclosed if an acquisition is mentioned but "
     "the acquirer is not named."),
    ("Where is the company headquarters located?",
     "City and country, e.g. 'Lund, Sweden'."),
    ("In which country/countries does manufacturing take place?",
     "One country per line. List every manufacturing country stated on the "
     "website."),
    ("Does the company have printed circuit board (PCB) manufacturing or "
     "assembly capability?", YN),
    ("Does the company have systems integration capability?", YN),
    ("Does the company have plastic moulding capability?", YN),
    ("Does the company have end-of-line (EOL) testing capability?", YN),
    ("Does the company have new product introduction (NPI) support?", YN),
    ("Does the company have tooling capability?", YN),
    ("Does the company have experience of medical device manufacturing?", YN),
    ("What is the company's yearly revenue?",
     "Value with currency and year where stated, e.g. 'EUR 45 million "
     "(2024)'. Write Not disclosed if not on the website."),
    ("How many employees does the company have?",
     "A number or range, with year if stated, e.g. '1,200 (2024)'."),
    ("What is their typical production volume?",
     "As stated on the website: units per year if given, otherwise the "
     "site's own wording (e.g. 'low-volume, high-mix')."),
    ("Do they produce low volumes (around 500-1000 products/devices per "
     "year)?", YN),
    ("Does manufacturing take place exclusively in China?", YN),
]

INSTRUCTIONS = [
    ("Purpose",
     "This sheet is the ground truth for evaluating an automated "
     "web-extraction tool. The tool reads ONLY each company's public "
     "website, so for a fair comparison please answer from the website too."),
    ("Which website",
     "Use the URL in the Website column — these have been verified and "
     "updated (some differ from the original list where companies moved, "
     "were acquired, or the old link was dead). Browse any pages and PDFs "
     "on that site."),
    ("Information not on the website",
     "Write 'Not disclosed' — even if you know the answer from elsewhere. "
     "If you want to record outside knowledge, put it in Notes / sources; "
     "it will not be scored against the tool."),
    ("Empty cells",
     "Leave a cell EMPTY only if you did not assess it. An empty cell means "
     "'not checked', never 'checked and found nothing' — that distinction "
     "is what the evaluation runs on."),
    ("Multiple answers in one cell",
     "Put one item per line inside the cell (Alt+Enter in Excel), e.g. one "
     "manufacturing country per line."),
    ("Yes/No questions",
     "Answer exactly Yes, No, or Not disclosed. 'No' needs positive evidence "
     "on the website — a statement or a clear implication from stated facts "
     "(e.g. manufacturing listed only in the US -> 'exclusively in China?' = "
     "No). 'Not disclosed' means the site doesn't say either way — expect "
     "this often; it is a useful answer, not a failure. If torn between No "
     "and Not disclosed, choose Not disclosed, and when you do answer No, "
     "paste the page that shows it in Notes / sources."),
    ("Answer format row",
     "The grey row under the headers shows the expected answer format for "
     "each question — please follow it; it keeps answers comparable."),
    ("Notes / sources (optional)",
     "For non-obvious answers (revenue, acquisitions), pasting the page URL "
     "you found it on makes disagreements easy to settle later."),
    ("Date checked",
     "The date you completed the row — websites change."),
    ("Excluded companies tab",
     "Companies we could not find a usable website for (duplicates, "
     "acquired-and-absorbed, no site found). No research needed — listed "
     "for transparency only."),
]

_HEADER_FILL = PatternFill("solid", start_color="2E4057")
_GUIDE_FILL = PatternFill("solid", start_color="EEEEEE")
_PREFILL_FILL = PatternFill("solid", start_color="E8EAF6")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = _WRAP_TOP


def build_template(urls: pd.DataFrame, inv: pd.DataFrame) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Ground Truth"
    ws.sheet_properties.tabColor = "4CAF50"
    headers = (["CMO", "Website"] + [q for q, _ in QUESTIONS]
               + ["Notes / sources", "Date checked"])
    ws.append(headers)
    # Guidance row: entity cell EMPTY on purpose — gt_convert skips leading
    # rows with no entity, so the filled sheet converts without deleting it.
    guide = ["", "ANSWER FORMAT ->"] + [g for _, g in QUESTIONS] + [
        "Optional: page URL where you found non-obvious answers", "When you finished the row"]
    ws.append(guide)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = _GUIDE_FILL
        cell.font = Font(italic=True, size=9, color="555555")
        cell.alignment = _WRAP_TOP
    for _, r in urls.iterrows():
        ws.append([r["entities"], r["url"]] + [""] * (len(QUESTIONS) + 2))
    for row in range(3, 3 + len(urls)):
        for c in (1, 2):
            ws.cell(row=row, column=c).fill = _PREFILL_FILL
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top")
        ws.cell(row=row, column=1).alignment = _WRAP_TOP
    _style_header(ws, len(headers))
    widths = [30, 36] + [28] * len(QUESTIONS) + [40, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 68
    ws.row_dimensions[2].height = 64
    ws.freeze_panes = "C3"

    ws_i = wb.create_sheet("Instructions")
    ws_i.sheet_properties.tabColor = "2196F3"
    ws_i.append(["Topic", "Instruction"])
    for topic, text in INSTRUCTIONS:
        ws_i.append([topic, text])
        ws_i.cell(row=ws_i.max_row, column=1).font = Font(bold=True)
        ws_i.cell(row=ws_i.max_row, column=2).alignment = _WRAP_TOP
    _style_header(ws_i, 2)
    ws_i.column_dimensions["A"].width = 30
    ws_i.column_dimensions["B"].width = 100

    ws_x = wb.create_sheet("Excluded companies")
    ws_x.sheet_properties.tabColor = "9E9E9E"
    ws_x.append(["Company", "Reason (manual URL research, July 2026)"])
    reasons = {
        "george_no_access": "No usable website found (site gone or inaccessible)",
        "george_unknown": "Could not identify the company / its website",
        "george_bad_landing": "Listed URL opens but lands on an unrelated page",
        "george_maybe": "Only an uncertain website candidate found — excluded",
    }
    excluded = inv[inv["cohort"].astype(str).str.startswith("george_")]
    for _, r in excluded.iterrows():
        note = str(r.get("george_note", "") or "").strip()
        extra = f" ({note})" if note and note.upper() not in (
            "NO_ACCESS", "UNKNOWN", "RANDOM LANDING PAGE") else ""
        ws_x.append([r["entity"], reasons.get(r["cohort"], r["cohort"]) + extra])
    ws_x.append([])
    ws_x.append(["(Duplicate rows in the original list were merged into one "
                 "row per company and are not shown here.)"])
    _style_header(ws_x, 2)
    ws_x.column_dimensions["A"].width = 34
    ws_x.column_dimensions["B"].width = 80
    for row in ws_x.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP_TOP

    wb.save(OUT_TEMPLATE)
    print(f"Analyst template written: {OUT_TEMPLATE} — {len(urls)} entities, "
          f"{len(QUESTIONS)} questions, {len(excluded)} excluded listed")


def build_input_v2(src: pd.ExcelFile) -> None:
    entities = pd.read_excel(src, "entities")
    urls = pd.read_excel(src, "urls")
    config = pd.read_excel(src, "config")
    questions = pd.DataFrame({
        "question": [q for q, _ in QUESTIONS],
        "instructions": [g for _, g in QUESTIONS],
    })
    with pd.ExcelWriter(OUT_INPUT_V2, engine="openpyxl") as w:
        entities.to_excel(w, sheet_name="entities", index=False)
        urls.to_excel(w, sheet_name="urls", index=False)
        questions.to_excel(w, sheet_name="questions", index=False)
        config.to_excel(w, sheet_name="config", index=False)
    print(f"Pipeline workbook written: {OUT_INPUT_V2} — {len(entities)} "
          f"entities, {len(QUESTIONS)} questions (v2 canonical, instructions on)")


def main() -> int:
    for path in (SOURCE_WB, SOURCE_INV):
        if not os.path.exists(path):
            sys.exit(f"Missing {path} — decode/build the CMO inputs first.")
    src = pd.ExcelFile(SOURCE_WB)
    urls = pd.read_excel(src, "urls")
    inv = pd.read_csv(SOURCE_INV)
    build_template(urls, inv)
    build_input_v2(src)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
