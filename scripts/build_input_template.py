"""Build a styled, self-explanatory input workbook a user can fill in and run.

    python scripts/build_input_template.py my_project.xlsx

Writes the four-sheet input format (entities / urls / questions / config)
plus a leading Instructions tab. This script is also what the web UI's
"Download blank template" button serves (webapp/server.py, /api/template),
so the instructions lead with the upload flow and mention the CLI second.

Presentation contract:
- every data cell wraps text and aligns top-left, so long questions and
  instructions stay readable;
- header rows are bold, filled, and frozen;
- example rows are highlighted and explicitly marked as replaceable;
- the config sheet lists EVERY workbook-overridable setting with its real
  default, its allowed values (as a dropdown where the set is closed), and a
  plain-language description. Defaults are imported from config.py so the
  template can never drift out of sync with the code.

Parsing contract (src/io_excel.py): sheet names and the column headers
'entity', 'url'/'depth'/'entities', 'question'/'instructions' and
'setting'/'value' are matched by name — do not rename them. Extra columns
(e.g. 'allowed values') are ignored by the reader, so they are safe to ship.
"""
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Import the real defaults so template and code cannot disagree.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config as cfg  # noqa: E402

# ── shared styles ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF6E5")  # soft amber: "replace me"
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
NOTE_FONT = Font(italic=True, color="808080")
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _style_data_sheet(ws, widths: dict[str, int], n_example_rows: int) -> None:
    """Apply the shared look: styled+frozen header, wrapped cells, highlighted
    example rows, column widths."""
    for cell in ws[1]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP
    ws.freeze_panes = "A2"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if cell.row <= 1 + n_example_rows:
                cell.fill = EXAMPLE_FILL


# ── Instructions sheet ───────────────────────────────────────────────────────

def _write_instructions(wb) -> None:
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    def put(text: str, font=None, height: float | None = None):
        ws.append([None, text])
        c = ws.cell(row=ws.max_row, column=2)
        c.alignment = WRAP
        if font:
            c.font = font
        if height:
            ws.row_dimensions[ws.max_row].height = height

    put("Entity Extraction Pipeline — input workbook", TITLE_FONT, 28)
    put("Answers your questions about a list of companies from their websites, "
        "with a verbatim quote backing every answer.", NOTE_FONT, 28)
    put("")
    put("How to run (web page)", SECTION_FONT)
    put("1.  Fill in the entities, urls and questions sheets (guidance below).", None)
    put("2.  Replace or delete the highlighted example rows — they are placeholders.", None)
    put("3.  Save this file, then go back to the web page and upload it.", None)
    put("4.  Press “Run pipeline” and watch the live log.", None)
    put("5.  When it finishes, download the results workbook from the same page.", None)
    put("")
    put("Command-line alternative:  python main.py --input <this file> --output results.xlsx",
        NOTE_FONT)
    put("")
    put("What goes in each sheet", SECTION_FONT)
    put("entities — one company per row, written exactly as you want it to appear "
        "in the output.", None, 28)
    put("urls — each company's website (full address, starting https://), a crawl "
        "depth, and the company it belongs to. Depth 1 = the page plus the most "
        "relevant links on it; 0 = just that page.", None, 42)
    put("questions — one question per row. The optional instructions column "
        "controls the answer format, and the output really does follow it — "
        "e.g. “Answer Yes, No, or Not disclosed.”", None, 42)
    put("config — optional run settings. Every row is pre-filled with the "
        "default; delete any row you don't want to touch. The “allowed values” "
        "column shows exactly what each setting accepts.", None, 42)
    put("")
    put("The three mistakes that cause empty output", SECTION_FONT)
    put("1.  The entities column on the urls sheet must match the entity names "
        "on the entities sheet EXACTLY (same spelling, same punctuation). If "
        "they don't match, the URL is not crawled for that company.", None, 42)
    put("2.  URLs must be full addresses: https://www.acme.com — not acme.com.", None, 28)
    put("3.  Vague questions get vague answers. “Does the company offer contract "
        "manufacturing?” beats “Manufacturing?”.", None, 28)


# ── config sheet rows: setting, default, allowed values, description ─────────

def _config_rows() -> list[tuple[str, object, str, str]]:
    return [
        ("CRAWL_MAX_PAGES", cfg.CRAWL_MAX_PAGES,
         "whole number, 1 or more",
         "Page budget per company. More pages = better coverage, longer runs."),
        ("DEFAULT_DEPTH", cfg.DEFAULT_DEPTH,
         "whole number, 0 or more",
         "Crawl depth used when the urls sheet leaves depth blank. "
         "0 = seed page only; 1 = seed plus its most relevant links."),
        ("CRAWL_SCOPE", cfg.CRAWL_SCOPE,
         "host | site",
         "host = stay on the exact website given; site = also crawl the "
         "company's own subdomains."),
        ("CRAWL_RENDER_FOR_DISCOVERY", str(cfg.CRAWL_RENDER_FOR_DISCOVERY).lower(),
         "true | false",
         "true helps sites whose menus only appear with JavaScript, at the "
         "cost of slower crawling."),
        ("CRAWL_BLOCK_INFRA_PATHS", str(cfg.CRAWL_BLOCK_INFRA_PATHS).lower(),
         "true | false",
         "true skips infrastructure links (API endpoints, template artefacts) "
         "at link discovery."),
        ("CRAWL_SCORER", cfg.CRAWL_SCORER,
         "baseline | experimental",
         "How links are ranked for crawling. Leave at baseline unless testing."),
        ("CRAWL_MIN_SCORE", cfg.CRAWL_MIN_SCORE,
         "decimal between 0 and 1",
         "Keyword-scorer threshold: links scoring below this are not followed."),
        ("CRAWL_MIN_SCORE_EMBED", cfg.CRAWL_MIN_SCORE_EMBED,
         "decimal between 0 and 1",
         "Embedding-scorer threshold: links scoring below this are not followed."),
        ("ACQUIRE_TOOL", cfg.ACQUIRE_TOOL,
         "playwright_pooled_hybrid | playwright_pooled | firecrawl | sgai | "
         "playwright | requests | local",
         "How pages are fetched. The default fetches statically and only opens "
         "a browser when needed; firecrawl uses the paid vendor service."),
        ("EXTRACT_TOOL", cfg.EXTRACT_TOOL,
         "azure | claude | sgai | llmapi",
         "Which LLM backend extracts the answers. Needs the matching API key "
         "in .env (see .env.example)."),
        ("CACHE_DIR", cfg.CACHE_DIR,
         "any folder path",
         "Where fetched pages are cached. Change only to keep runs isolated."),
    ]


_DROPDOWNS = {
    "CRAWL_SCOPE": '"host,site"',
    "CRAWL_RENDER_FOR_DISCOVERY": '"true,false"',
    "CRAWL_BLOCK_INFRA_PATHS": '"true,false"',
    "CRAWL_SCORER": '"baseline,experimental"',
    "ACQUIRE_TOOL": ('"playwright_pooled_hybrid,playwright_pooled,firecrawl,'
                     'sgai,playwright,requests,local"'),
    "EXTRACT_TOOL": '"azure,claude,sgai,llmapi"',
}


def main() -> int:
    out = sys.argv[1] if len(sys.argv) > 1 else "input_template.xlsx"
    if not out.endswith(".xlsx"):
        out += ".xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_instructions(wb)

    ws = wb.create_sheet("entities")
    ws.append(["entity"])
    ws.append(["Example Company Ltd"])
    ws.append(["Example Widgets Inc"])
    _style_data_sheet(ws, {"A": 44}, n_example_rows=2)

    ws = wb.create_sheet("urls")
    ws.append(["url", "depth", "entities"])
    ws.append(["https://www.example.com", 1, "Example Company Ltd"])
    ws.append(["https://www.example.org", 1, "Example Widgets Inc"])
    _style_data_sheet(ws, {"A": 44, "B": 10, "C": 34}, n_example_rows=2)

    ws = wb.create_sheet("questions")
    ws.append(["question", "instructions"])
    ws.append(["Where is the company headquartered?",
               "City and country, e.g. 'Cambridge, UK'."])
    ws.append(["Does the company have ISO 13485 certification?",
               "Answer Yes, No, or Not disclosed. Answer No only if the site "
               "states it explicitly."])
    ws.append(["Which products or services does the company offer?",
               "Comma-separated list of short phrases."])
    _style_data_sheet(ws, {"A": 52, "B": 52}, n_example_rows=3)

    ws = wb.create_sheet("config")
    ws.append(["setting", "value", "allowed values", "what it does"])
    for r, (key, value, allowed, meaning) in enumerate(_config_rows(), start=2):
        ws.append([key, value, allowed, meaning])
        formula = _DROPDOWNS.get(key)
        if formula:
            dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                                showErrorMessage=True,
                                error="Pick one of the allowed values.")
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=r, column=2))
    # Config rows are real defaults, not placeholders — no example highlight.
    _style_data_sheet(ws, {"A": 30, "B": 24, "C": 34, "D": 58}, n_example_rows=0)

    wb.save(out)
    print(f"Template written: {out}")
    print("Fill in the entities, urls and questions sheets, then either upload")
    print("it on the web page (python -m webapp) or run:")
    print(f"  python main.py --input {out} --output results.xlsx")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
