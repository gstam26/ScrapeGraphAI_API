"""One-command CMO scoring: pipeline workbook in, styled eval report out.

    python scripts/score_cmo.py outputs/<run>.xlsx
    python scripts/score_cmo.py outputs/<run>.xlsx --gt outputs/cmo_gt_final59.xlsx

Wraps generic_eval with the decided HEADLINE configuration — deliverable
grain (--sheet matrix), prose cells joined (--prose-cells), CE matcher — so
final numbers are always produced the same way, and stamps a Run Info sheet
(GT version, flags, matcher backend, date) into the report so no scoring is
ever ambiguous about how it was made.

Presentation: the report is styled for reading — frozen headers, wrapped
text, verdict colour-coding, autofilter — without touching a single scored
number (write_report_excel's data is written first, styling is applied to
the saved file afterwards).

The report also carries a "Matrix View" sheet: the analyst-template shape
(one row per CMO, one column per question, Website column from the run
input) with each cell showing the eval outcome — ✓ matched value, ✗ missed
GT value, + unmatched AI extra — colour-coded by cell outcome, with the
full GT/AI values and matcher scores in the cell comment. It is rendered
FROM the already-written Detail sheet, so it can never disagree with the
scored numbers; `--retrofit <report.xlsx>` adds it to an existing report
without re-scoring.
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

DEFAULT_GT = os.path.join("outputs", "cmo_gt_final59.xlsx")
DEFAULT_URLS = os.path.join("cmo-inputs", "cmo_input_v2.xlsx")

# Analyst-template column order (cmo_gt_analyst_TEMPLATE.xlsx "Ground Truth"
# sheet). The Matrix View follows this order so it reads like the artefact the
# analysts and the deliverable both use; questions not in this list (future
# tasks) are appended alphabetically after.
CMO_QUESTION_ORDER = [
    "Summary description of company's services",
    "Is the company still operating independently?",
    "If not operating independently: who acquired or absorbed the company?",
    "Where is the company headquarters located?",
    "In which country/countries does manufacturing take place?",
    "Does the company have printed circuit board (PCB) manufacturing or assembly capability?",
    "Does the company have systems integration capability?",
    "Does the company have plastic moulding capability?",
    "Does the company have end-of-line (EOL) testing capability?",
    "Does the company have new product introduction (NPI) support?",
    "Does the company have tooling capability?",
    "Does the company have experience of medical device manufacturing?",
    "What is the company's yearly revenue?",
    "How many employees does the company have?",
    "What is their typical production volume?",
    "Do they produce low volumes (around 500-1000 products/devices per year)?",
    "Does manufacturing take place exclusively in China?",
]

# Verdict fills — muted so text stays readable when printed.
_VERDICT_FILL = {
    "auto_match":      "C6EFCE",  # green
    "review":          "C6EFCE",
    "semantic_review": "DDEBF7",  # blue — CE-only credit, auditable
    "null_match":      "E2EFDA",  # pale green — correct abstention
    "auto_miss":       "FFC7CE",  # red — missed GT
    "no_ai_data":      "FFC7CE",
    "ai_only":         "FFE699",  # amber — unmatched claim (FP)
    "redundant":       "EDEDED",  # grey — restatement, precision-exempt
    "suppressed_null": "EDEDED",  # grey — abstention, precision-exempt
}

# Plain-English meaning of every verdict, written INTO the report (Run Info
# glossary) so a reader never needs the source code to interpret a row.
_VERDICT_GLOSSARY = [
    ("auto_match",
     "The AI value agrees with the GT value: strong lexical agreement, or an "
     "exact match on typed values (numbers, Yes/No). Counted as a true "
     "positive."),
    ("semantic_review",
     "Credited by the semantic matcher (cross-encoder) alone: the two texts "
     "differ in wording but were judged equivalent (e.g. a paraphrased "
     "description). Counted as a true positive; coloured blue so every "
     "matcher-only credit stays auditable."),
    ("null_match",
     "GT confirmed the information is absent ('None (not disclosed)') and "
     "the pipeline also reported no data — a correct abstention. Counted as "
     "a true positive."),
    ("auto_miss",
     "A GT value the pipeline failed to report, or reported something not "
     "judged equivalent. Counted as a false negative."),
    ("no_ai_data",
     "The pipeline produced nothing at all for this cell; each GT value in "
     "it counts as a false negative."),
    ("ai_only",
     "An AI claim with no matching GT value. Counted as a false positive — "
     "but on list questions the GT is non-exhaustive, so some of these may "
     "be real-but-unlisted (precision there is a lower bound)."),
    ("redundant",
     "A restatement of a claim already credited in the same cell (e.g. "
     "'Geneva' beside a matched 'Geneva, Switzerland'). Precision-exempt: "
     "neither a true nor a false positive."),
    ("suppressed_null",
     "A 'Not disclosed' claim sitting beside a substantive answer in the "
     "same cell — page-local absence, i.e. an abstention. An abstention "
     "asserts nothing, so it is never a false positive; any GT value the "
     "pipeline failed to find is still charged as a miss."),
    ("review",
     "Moderate lexical agreement band (embedding backend only; does not "
     "occur under the cross-encoder matcher). Counted as a true positive, "
     "flagged for inspection."),
]


def _write_glossary(wb) -> None:
    """Append the verdict glossary to Run Info (idempotent)."""
    if "Run Info" not in wb.sheetnames:
        return
    from openpyxl.styles import Alignment, Font, PatternFill
    info = wb["Run Info"]
    have = {str(r[0].value) for r in info.iter_rows(min_col=1, max_col=1)}
    if "Verdict glossary" in have:
        return
    wrap = Alignment(wrap_text=True, vertical="top")
    info.append([])
    info.append(["Verdict glossary",
                 "Meaning of every verdict in the Detail sheet (cell colours "
                 "match the Detail rows and the Matrix View):"])
    info.cell(row=info.max_row, column=1).font = Font(bold=True, size=12)
    info.cell(row=info.max_row, column=2).alignment = wrap
    for verdict, meaning in _VERDICT_GLOSSARY:
        info.append([verdict, meaning])
        c = info.cell(row=info.max_row, column=1)
        c.font = Font(bold=True)
        fill = _VERDICT_FILL.get(verdict)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        info.cell(row=info.max_row, column=2).alignment = wrap


def _git_stamp() -> str:
    """Short commit id of the scoring code, '+local changes' when dirty —
    so every report records exactly which evaluator produced it."""
    import subprocess
    try:
        head = subprocess.run(["git", "rev-parse", "--short", "HEAD"],
                              capture_output=True, text=True, timeout=10,
                              cwd=_REPO_ROOT).stdout.strip()
        dirty = subprocess.run(["git", "status", "--porcelain"],
                               capture_output=True, text=True, timeout=10,
                               cwd=_REPO_ROOT).stdout.strip()
        return (head or "unknown") + (" +local changes" if dirty else "")
    except Exception:  # noqa: BLE001 — provenance nicety, never fatal
        return "unknown"


def _reorder_summary(ws) -> None:
    """Rewrite Summary data rows into the analyst-template question order
    (unknown questions after, alphabetically; OVERALL last) — the same order
    as the Matrix View columns, so the two sheets read in parallel."""
    canon = {_qnorm(q): i for i, q in enumerate(CMO_QUESTION_ORDER)}
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]

    def _key(vals):
        q = str(vals[0])
        if q == "OVERALL":
            return (2, 0, "")
        i = canon.get(_qnorm(q))
        return (0, i, "") if i is not None else (1, 0, q)

    rows.sort(key=_key)
    for i, vals in enumerate(rows, start=2):
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)


def style_report(path: str, meta_rows: list[tuple[str, str]]) -> None:
    """Post-style the written report + prepend a Run Info sheet."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="44546A")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ── Run Info sheet, first position ────────────────────────────────────
    info = wb.create_sheet("Run Info", 0)
    info.append(["CMO evaluation report"])
    info["A1"].font = Font(bold=True, size=14)
    info.append([])
    for k, v in meta_rows:
        info.append([k, v])
        info.cell(row=info.max_row, column=1).font = Font(bold=True)
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 90
    for row in info.iter_rows(min_row=3):
        row[1].alignment = wrap

    # ── Summary: headers, widths, percent-ish rounding, freeze ────────────
    ws = wb["Summary"]
    _reorder_summary(ws)
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 62
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = wrap
        for c in row[5:9]:
            c.number_format = "0.000"
        if str(row[0].value) == "OVERALL":
            for c in row:
                c.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions

    # ── Detail: freeze entity+question, wrap the claim text, colour verdicts ──
    from openpyxl.comments import Comment
    ws = wb["Detail"]
    hdr = [str(c.value) for c in ws[1]]
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
    ws.cell(row=1, column=hdr.index("verdict") + 1).comment = Comment(
        "What each verdict means: see the Verdict glossary on the Run Info "
        "sheet.", "eval", height=80, width=320)
    ws.freeze_panes = "C2"
    widths = {"entity": 24, "question": 42, "gt_value": 46, "ai_value": 46,
              "is_list": 8, "value_score": 11, "quote_score": 11,
              "semantic": 11, "combined": 11, "verdict": 16}
    v_idx = hdr.index("verdict")
    for i, name in enumerate(hdr, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)
    for row in ws.iter_rows(min_row=2):
        for name, c in zip(hdr, row):
            if name in ("entity", "question", "gt_value", "ai_value"):
                c.alignment = wrap
        fill = _VERDICT_FILL.get(str(row[v_idx].value))
        if fill:
            row[v_idx].fill = PatternFill("solid", fgColor=fill)
    ws.auto_filter.ref = ws.dimensions

    _write_glossary(wb)
    wb.save(path)


def _qnorm(s: str) -> str:
    return " ".join(str(s).lower().split())


def _short(s: str, n: int = 110) -> str:
    s = str(s).strip()
    return s if len(s) <= n else s[: n - 1].rstrip() + "…"


def _read_url_map(urls_path: str) -> dict[str, str]:
    """entity(normalised) -> website, from the run-input urls sheet.
    Missing/odd input workbook degrades to blank Website cells, never fails."""
    try:
        import pandas as pd
        df = pd.read_excel(urls_path, sheet_name="urls")
        out: dict[str, str] = {}
        for _, r in df.iterrows():
            for ent in str(r.get("entities", "")).split(";"):
                out.setdefault(_qnorm(ent), str(r.get("url", "")).strip())
        return out
    except Exception as e:  # noqa: BLE001 — Website column is a nicety
        print(f"  [matrix view] no url map ({type(e).__name__}: {e}); "
              f"Website column left blank")
        return {}


# Verdicts that mean "this GT value was credited".
_TP_VERDICTS = frozenset({"auto_match", "review", "semantic_review", "null_match"})
# Precision-exempt rows: shown in the cell comment only, never in the grid.
_EXEMPT_VERDICTS = frozenset({"redundant", "suppressed_null"})

_CELL_FILL = {
    "right":   "C6EFCE",  # every GT value credited, no extras
    "abstain": "E2EFDA",  # correct abstention (all null_match)
    "partial": "FFE699",  # some credit, some miss/extra
    "wrong":   "FFC7CE",  # no credit at all
    "exempt":  "EDEDED",  # only restatements/abstentions — nothing scored
}


def _render_cell(rows: list[dict]) -> tuple[str, str, str, tuple[int, int, int]]:
    """One (entity, question) cell -> (visible text, outcome key, comment text).

    Visible text is one marker line per scored item — ✓ credited AI value
    (blue-credited pairs show AI ⇄ GT since the texts differ), ✗ missed GT
    value, + unmatched AI extra — so both sides of every disagreement are on
    the grid. The comment holds the full untruncated values, matcher scores
    and the precision-exempt rows."""
    vis: list[str] = []
    com: list[str] = []
    tp = fn = fp = nulls = 0
    for r in rows:
        v = r["verdict"]
        gt = str(r["gt_value"] or "").strip()
        ai = str(r["ai_value"] or "").strip()
        try:
            scores = f"lex={float(r['combined']):.2f} ce={float(r['semantic']):.2f}"
        except (TypeError, ValueError):
            scores = ""
        com.append(f"[{v}] GT: {gt or '—'} | AI: {ai or '—'}"
                   + (f"  ({scores})" if scores else ""))
        if v in _EXEMPT_VERDICTS:
            continue
        if v in _TP_VERDICTS:
            tp += 1
            if v == "null_match":
                nulls += 1
                vis.append("✓ Not disclosed")
            elif v == "semantic_review":
                vis.append(f"✓ {_short(ai, 60)} ⇄ GT: {_short(gt, 60)}")
            else:
                vis.append(f"✓ {_short(ai)}")
        elif v == "ai_only":
            fp += 1
            vis.append(f"+ AI: {_short(ai)}")
        else:  # auto_miss / no_ai_data
            fn += 1
            vis.append(f"✗ GT: {_short(gt)}")

    if tp and not fn and not fp:
        outcome = "abstain" if nulls == tp else "right"
    elif tp:
        outcome = "partial"
    elif fn or fp:
        outcome = "wrong"
    else:
        # Only precision-exempt rows (e.g. GT left blank by the analyst, AI
        # abstained) — nothing scored; say so rather than show an empty box.
        outcome = "exempt"
        vis.append("(not scored)")
    return "\n".join(vis), outcome, "\n".join(com), (tp, fn, fp)


def _cell_f1(tp: int, fn: int, fp: int) -> float:
    """Per-cell F1 with the evaluator's empty-denominator convention
    (no claims on a side => that side is perfect)."""
    r = tp / (tp + fn) if (tp + fn) else 1.0
    p = tp / (tp + fp) if (tp + fp) else 1.0
    return 2 * p * r / (p + r) if (p + r) else 0.0


_CELL_SUMMARY_COLS = ["cells_right", "cells_partial", "cells_wrong",
                      "cell_accuracy", "macro_F1"]


def _write_cell_summary(wb, qstats: dict[str, dict],
                        q_display: dict[str, str]) -> None:
    """Append the per-cell (macro) columns to the Summary sheet.

    cell_accuracy = fully-correct cells / scored cells (strict: every GT
    value credited, no extras; correct abstentions count). macro_F1 = mean
    of per-cell F1 — partial credit inside a cell, every cell equal weight.
    Complementary to the micro-pooled P/R/F1 to their left, which weigh
    cells by how many claims they contain. Idempotent: re-running overwrites
    the same columns."""
    if "Summary" not in wb.sheetnames:
        return
    from openpyxl.styles import Font, PatternFill
    sm = wb["Summary"]
    hdr = [str(c.value) for c in sm[1]]
    start = (hdr.index(_CELL_SUMMARY_COLS[0]) + 1
             if _CELL_SUMMARY_COLS[0] in hdr else sm.max_column + 1)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="44546A")
    for k, title in enumerate(_CELL_SUMMARY_COLS):
        c = sm.cell(row=1, column=start + k, value=title)
        c.font, c.fill = head_font, head_fill

    by_text = {q_display[qn]: s for qn, s in qstats.items()}
    tot = {"right": 0, "partial": 0, "wrong": 0, "f1s": []}
    for s in qstats.values():
        for k in ("right", "partial", "wrong"):
            tot[k] += s[k]
        tot["f1s"].extend(s["f1s"])

    for row in sm.iter_rows(min_row=2):
        q = str(row[0].value)
        s = tot if q == "OVERALL" else by_text.get(q)
        if s is None:
            continue
        n = s["right"] + s["partial"] + s["wrong"]
        vals = [s["right"], s["partial"], s["wrong"],
                round(s["right"] / n, 4) if n else None,
                round(sum(s["f1s"]) / n, 4) if n else None]
        for k, v in enumerate(vals):
            c = sm.cell(row=row[0].row, column=start + k, value=v)
            if k >= 3:
                c.number_format = "0.000"
            if q == "OVERALL":
                c.font = Font(bold=True)


def add_matrix_view(report_path: str, urls_path: str = DEFAULT_URLS) -> None:
    """Add/refresh the Matrix View sheet on a written eval report.

    Reads the report's own Detail sheet (never re-scores), so the view can
    never disagree with the Summary numbers."""
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(report_path)
    if "Matrix View" in wb.sheetnames:
        del wb["Matrix View"]
    det = wb["Detail"]
    hdr = [str(c.value) for c in det[1]]
    col = {name: hdr.index(name) for name in
           ("entity", "question", "gt_value", "ai_value",
            "semantic", "combined", "verdict")}

    # Group Detail rows into cells, preserving entity order of appearance.
    cells: dict[tuple[str, str], list[dict]] = {}
    entities: list[str] = []
    q_display: dict[str, str] = {}      # qnorm -> display text
    for raw in det.iter_rows(min_row=2, values_only=True):
        r = {name: raw[i] for name, i in col.items()}
        ent = str(r["entity"]).strip()
        qn = _qnorm(r["question"])
        if ent not in entities:
            entities.append(ent)
        q_display.setdefault(qn, str(r["question"]).strip())
        cells.setdefault((ent, qn), []).append(r)

    # Question order: analyst template first, unknown questions appended.
    canon = [_qnorm(q) for q in CMO_QUESTION_ORDER]
    q_order = [qn for qn in canon if qn in q_display]
    q_order += sorted(qn for qn in q_display if qn not in canon)

    url_map = _read_url_map(urls_path)

    pos = wb.sheetnames.index("Summary") + 1 if "Summary" in wb.sheetnames else 1
    ws = wb.create_sheet("Matrix View", pos)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="44546A")
    wrap = Alignment(wrap_text=True, vertical="top")
    fills = {k: PatternFill("solid", fgColor=v) for k, v in _CELL_FILL.items()}

    ws.append(["CMO", "Website"] + [q_display[qn] for qn in q_order] + ["Cells"])
    for c in ws[1]:
        c.font, c.fill, c.alignment = head_font, head_fill, wrap
    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    for i, qn in enumerate(q_order, start=3):
        ws.column_dimensions[get_column_letter(i)].width = (
            46 if "summary description" in qn else 30)
    ws.column_dimensions[get_column_letter(len(q_order) + 3)].width = 13

    # Per-question cell-level stats (macro view: every cell weighs the same,
    # unlike the Summary's micro-pooled claim counts).
    qstats: dict[str, dict] = {
        qn: {"right": 0, "partial": 0, "wrong": 0, "f1s": []} for qn in q_order
    }

    for ent in entities:
        row_i = ws.max_row + 1
        ws.cell(row=row_i, column=1, value=ent).alignment = wrap
        ws.cell(row=row_i, column=2,
                value=url_map.get(_qnorm(ent), "")).alignment = wrap
        counts = {"right": 0, "partial": 0, "wrong": 0}
        for j, qn in enumerate(q_order, start=3):
            rows = cells.get((ent, qn))
            if not rows:
                continue
            text, outcome, comment, (tp, fn, fp) = _render_cell(rows)
            c = ws.cell(row=row_i, column=j, value=text)
            c.alignment = wrap
            c.fill = fills[outcome]
            if comment:
                c.comment = Comment(comment, "eval", height=280, width=460)
            if outcome == "exempt":
                continue  # not scored — excluded from cell-level stats too
            key = ("right" if outcome in ("right", "abstain")
                   else "partial" if outcome == "partial" else "wrong")
            counts[key] += 1
            qstats[qn][key] += 1
            qstats[qn]["f1s"].append(_cell_f1(tp, fn, fp))
        ws.cell(row=row_i, column=len(q_order) + 3,
                value=f"{counts['right']}✓ {counts['partial']}◐ "
                      f"{counts['wrong']}✗").alignment = wrap

    _write_cell_summary(wb, qstats, q_display)
    _write_glossary(wb)

    # Legend on the Run Info sheet (when present), so the view explains itself.
    if "Run Info" in wb.sheetnames:
        info = wb["Run Info"]
        legend = ("✓ credited (⇄ shows AI vs GT when texts differ) · "
                  "✗ missed GT value · + unmatched AI extra. Cell colour: "
                  "green=fully correct, pale green=correct abstention, "
                  "amber=partial, red=wrong, grey=only restatements/"
                  "abstentions. Hover a cell for full values + matcher "
                  "scores. Cells column: per-CMO right/partial/wrong count.")
        cell_legend = ("Summary's cells_right/partial/wrong + cell_accuracy "
                       "+ macro_F1 are the CELL-LEVEL (macro) view: every "
                       "cell weighs the same regardless of how many claims "
                       "it holds. cell_accuracy is strict (cell fully "
                       "correct, abstentions count); macro_F1 gives partial "
                       "credit inside a cell. Complementary to the "
                       "micro-pooled P/R/F1, which weigh cells by claim "
                       "count.")
        have = {str(r[0].value) for r in info.iter_rows(min_col=1, max_col=1)}
        if "Matrix View legend" not in have:
            info.append(["Matrix View legend", legend])
            info.cell(row=info.max_row, column=1).font = Font(bold=True)
            info.cell(row=info.max_row, column=2).alignment = wrap
        if "Per-cell columns" not in have:
            info.append(["Per-cell columns", cell_legend])
            info.cell(row=info.max_row, column=1).font = Font(bold=True)
            info.cell(row=info.max_row, column=2).alignment = wrap

    wb.save(report_path)
    print(f"Matrix View sheet added: {report_path}")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Score a CMO pipeline workbook against the final GT "
                    "(headline config: matrix sheet, prose cells, CE matcher).")
    ap.add_argument("pipeline_output", nargs="?",
                    help="Pipeline output workbook (.xlsx)")
    ap.add_argument("--gt", default=DEFAULT_GT,
                    help=f"Flat GT workbook (default: {DEFAULT_GT})")
    ap.add_argument("--output", default=None,
                    help="Report path (default: outputs/eval_<run-name>_<date>.xlsx)")
    ap.add_argument("--urls", default=DEFAULT_URLS,
                    help="Run-input workbook whose 'urls' sheet fills the "
                         f"Matrix View Website column (default: {DEFAULT_URLS})")
    ap.add_argument("--retrofit", metavar="REPORT",
                    help="Add/refresh the Matrix View sheet on an EXISTING "
                         "eval report (built from its Detail sheet — no "
                         "re-scoring) and exit")
    args = ap.parse_args()

    if args.retrofit:
        if not os.path.exists(args.retrofit):
            sys.exit(f"Report not found: {args.retrofit}")
        add_matrix_view(args.retrofit, args.urls)
        return 0

    if args.pipeline_output is None:
        ap.error("pipeline_output is required unless --retrofit is used")
    if not os.path.exists(args.gt):
        sys.exit(f"GT not found: {args.gt} (pull, or pass --gt)")
    if not os.path.exists(args.pipeline_output):
        sys.exit(f"Pipeline workbook not found: {args.pipeline_output}")

    run_name = os.path.splitext(os.path.basename(args.pipeline_output))[0]
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out = args.output or os.path.join("outputs", f"eval_{run_name}_{stamp}.xlsx")

    from src.eval.generic_eval import (evaluate, print_report, read_gt,
                                       read_pipeline_matrix, read_run_entities,
                                       write_report_excel)

    gt = read_gt(args.gt)
    ai = read_pipeline_matrix(args.pipeline_output)
    run_ents = read_run_entities(args.pipeline_output)
    result = evaluate(gt, ai, semantic=True, semantic_backend="cross-encoder",
                      run_entities=run_ents, prose_cells=True)
    print_report(result)
    write_report_excel(result, out)

    o = result.overall
    style_report(out, [
        ("Pipeline run", os.path.basename(args.pipeline_output)),
        ("Ground truth", f"{os.path.basename(args.gt)}  "
                         f"({o['entities']} entities scored, {len(gt)} GT rows)"),
        ("Scored on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Code version", f"{_git_stamp()} (scripts/score_cmo.py)"),
        ("Configuration", "sheet=matrix (deliverable grain), prose-cells=on, "
                          "semantic-backend=cross-encoder (decisive)"),
        ("Headline", f"P={o['precision']:.3f}  R={o['recall']:.3f}  "
                     f"F1={o['F1']:.3f}  hallucination={o['hallucination_rate']:.3f}"),
        ("Reading guide", "List-question precision is a LOWER BOUND (GT is "
                          "non-exhaustive): unmatched items may be real but "
                          "unlisted. Grey Detail rows are precision-exempt "
                          "(restatements / abstentions). Blue rows are "
                          "matches credited by the semantic matcher alone."),
    ])
    add_matrix_view(out, args.urls)
    print(f"Styled report: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
