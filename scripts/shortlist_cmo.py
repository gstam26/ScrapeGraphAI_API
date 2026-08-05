"""Shortlist layer CLI — run, init-spec, and the two-run experiment.

    python scripts/shortlist_cmo.py init-spec                       # write default criteria workbook
    python scripts/shortlist_cmo.py run <workbook> [--spec S] [--out O] [--top K]
    python scripts/shortlist_cmo.py experiment <gt.xlsx> <ai.xlsx> [--spec S] [--out O]

Input auto-detection by sheet name: "Matrix" -> pipeline workbook (production
path), "GroundTruth" -> flat GT workbook, "Detail" -> eval report
(AI cells reconstructed; verified markers unavailable -> caveat recorded).

The experiment (George, 2026-08-05): the SAME spec file, byte-identical,
run twice — once on GT, once on the AI matrix. Output: top-k from each, a
full gate-verdict-level diff (every disagreement traced to the cell that
diverged), the sensitivity pass (threshold/weight perturbations + the
missing-policy flip), and the trivial-agreement guard (are both runs gating
on the same sparse disclosed minority?). Findings before interpretation.
"""
from __future__ import annotations

import argparse
import copy
import os
import sys
from datetime import datetime

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from src.shortlist import (Spec, diff_gates, gate_decisiveness, overlap_at_k,
                           read_eval_detail_cells, read_flat_gt_cells,
                           read_matrix_cells, read_spec, run_shortlist,
                           top_k, write_default_spec)

DEFAULT_SPEC = os.path.join("cmo-inputs", "shortlist_criteria.xlsx")
REF_YEAR = datetime.now().year


def read_cells_auto(path: str, sheet: str | None = None) -> tuple[dict, str]:
    """(cells, source_kind) — auto-detect the workbook format. An explicit
    --sheet always routes through the grid reader (works on Matrix, AI
    Summary, or any entity-rows × question-columns sheet)."""
    import pandas as pd
    if sheet is not None:
        return read_matrix_cells(path, sheet=sheet), f"grid sheet {sheet!r}"
    sheets = {s.lower() for s in pd.ExcelFile(path).sheet_names}
    if "matrix" in sheets or "ai summary" in sheets:
        cells = read_matrix_cells(path)
        return cells, "matrix" if "matrix" in sheets else "ai summary"
    if "groundtruth" in sheets:
        return read_flat_gt_cells(path), "flat_gt"
    if "detail" in sheets:
        return read_eval_detail_cells(path), "eval_detail (verified markers UNAVAILABLE)"
    raise SystemExit(f"{path}: no Matrix / AI Summary / GroundTruth / Detail "
                     f"sheet found — name one explicitly with --sheet")


# ---------------------------------------------------------------------------
# Consultant-facing explanations — composed from the SPEC, so the report can
# never describe criteria other than the ones actually used.
# ---------------------------------------------------------------------------
_FLAG_GLOSSARY = [
    ("QUALIFIED", "The value carried a qualifier ('over', '~', '+') — the stated number was used."),
    ("BOUND_OPEN", "The qualifier means the true value could exceed the stated number; the check passed on the stated value."),
    ("RANGE", "The value was a range; its bounds were used."),
    ("STALE", "The figure is older than the allowed age for this criterion."),
    ("PARENT_ATTRIBUTED", "The revenue figure belongs to the parent company, not this entity — it was NOT counted against the size limit."),
    ("NO_CURRENCY", "No currency was stated — USD assumed."),
    ("FORECAST", "Forward-looking figure (outlook / target)."),
    ("PARTIAL", "Part-year figure (e.g. first half)."),
    ("MULTI_VALUE", "The cell held several candidate answers; the lead one was used."),
    ("UNVERIFIED_ONLY", "Only unverified value(s) present — treated like missing data, kept with this flag."),
    ("UNVERIFIED_CELL", "The cell's values were not verified against source text."),
    ("CONFLICT", "Sources disagreed for this cell."),
    ("DEMOTED_RIVALS", "Other candidate answers exist — see the Provenance sheet of the pipeline workbook."),
    ("CAPPED", "The display cut additional items — see Provenance."),
    ("SUMMARY_FALLBACK", "The cell came from a fallback rendering of the summary layer."),
    ("VERIFIED_UNKNOWN", "This input carries no verification markers, so verified-only rules could not apply."),
]


def _plain_rule(c) -> str:
    """One criterion -> plain-English rule text."""
    if c.type == "hard_gate":
        if c.direction == "require_yes":
            return ("MUST-PASS - a company is excluded only on an explicit "
                    + ("verified " if c.unverified_policy == "verified_only" else "")
                    + "'No'. 'Not disclosed' or unreadable answers keep the "
                      "company in, with a note.")
        if c.direction == "require_no":
            return "MUST-PASS - excluded on an explicit 'Yes'; unknowns kept with a note."
        if c.direction == "require_match":
            pol = ("excluded" if c.missing_policy == "exclude"
                   else "kept in, with a note")
            return ("MUST-PASS - the company description must show they OFFER "
                    "manufacturing as a service (contract manufacturer / CMO / "
                    "EMS...), matched against the editable phrase list on the "
                    "spec's Keywords sheet. Descriptions with no matching "
                    f"phrase are {pol} - wording varies, so no-match is not "
                    "proof they only sell their own products.")
        if c.direction == "range":
            unit = " USD" if c.parser == "money" else ""
            pol = ("excluded" if c.missing_policy == "exclude"
                   else "kept in, with a note")
            own = (" Figures attributed to a parent company do not count."
                   if c.parser == "money" else "")
            return (f"MUST-PASS - the value must be between "
                    f"{c.threshold_lo:,.0f} and {c.threshold_hi:,.0f}{unit} "
                    f"(mid-sized band).{own} Companies with no disclosed "
                    f"figure are {pol}.")
        unit = " USD" if c.parser == "money" else ""
        lim = c.threshold_hi if c.direction == "max" else c.threshold_lo
        side = "above" if c.direction == "max" else "below"
        own = (" Figures attributed to a parent company do not count."
               if c.parser == "money" else "")
        pol = ("excluded" if c.missing_policy == "exclude"
               else "kept in, with a note")
        return (f"MUST-PASS - excluded if the value is {side} {lim:,.0f}{unit}."
                f"{own} Companies with no disclosed figure are {pol}.")
    return (f"RANKING (weight {c.weight:g}) - 'Yes' ranks above 'No'; "
            f"companies with no answer are not penalised on this criterion "
            f"but their coverage drops.")


def _entity_reason(r, spec_by_id: dict) -> str:
    """Plain-English 'why' for one entity row."""
    label = lambda cid: cid.replace("_", " ")
    if r.excluded_by:
        v = next(g for g in r.gates if g.criterion_id == r.excluded_by)
        c = spec_by_id[r.excluded_by]
        if c.parser == "binary":
            return (f"Excluded - answered '{v.parsed}' on {label(c.id)} "
                    f"(the site states this explicitly).")
        unit = " USD" if c.parser == "money" else ""
        if c.direction == "range":
            side = "below" if "below" in v.label else "above"
            band = f"{c.threshold_lo:,.0f}-{c.threshold_hi:,.0f}{unit}"
            return (f"Excluded - {label(c.id)}: {v.parsed} is {side} the "
                    f"{band} band.")
        lim = c.threshold_hi if c.direction == "max" else c.threshold_lo
        return (f"Excluded - {label(c.id)}: {v.parsed} is over the "
                f"{lim:,.0f}{unit} limit.")
    notes = []
    for v in r.gates:
        if "not disclosed" in v.label:
            notes.append(f"{label(v.criterion_id)} not disclosed")
        elif "no data" in v.label:
            notes.append(f"no data for {label(v.criterion_id)}")
        elif "unparseable" in v.label:
            notes.append(f"{label(v.criterion_id)} answer unreadable")
        elif "parent-attributed" in v.label:
            notes.append("revenue figure is the parent company's")
        elif "unverified only" in v.label:
            notes.append(f"{label(v.criterion_id)} unverified")
        elif "no service-language match" in v.label:
            notes.append("description has no recognised service phrase "
                         "(check it manually)")
    basis = "; ".join(
        f"{label(s.criterion_id)}: {s.label}" for s in r.scores
        if s.score is not None and spec_by_id[s.criterion_id].weight > 0
    ) or "no ranking data"
    why = f"Passed all must-pass checks. Ranked on {basis}."
    if notes:
        why += " Note: " + ", ".join(notes) + "."
    return why


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------
_HEAD = dict(bold=True, color="FFFFFF")
_STATUS_FILL = {"SHORTLISTED": "C6EFCE", "RANKED": "E2EFDA",
                "EXCLUDED": "FFC7CE"}
_DIFF_FILL = {"SAME": None, "VALUE_DISAGREES": "FFC7CE",
              "PARSE_DIVERGENCE": "FFE699", "MISSING_DIVERGENCE": "DDEBF7"}


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill
    for c in ws[1]:
        c.font = Font(**_HEAD)
        c.fill = PatternFill("solid", fgColor="44546A")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _howto_sheet(wb, spec) -> None:
    """Consultant 'How to read' sheet — generated FROM the spec in use."""
    from openpyxl.styles import Alignment, Font
    ws = wb.create_sheet("How To Read")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _section(title):
        ws.append([])
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    ws.append(["How this shortlist was made"])
    ws["A1"].font = Font(bold=True, size=14)
    _section("The criteria (in the order applied)")
    for c in spec.criteria:
        if c.type == "scored" and not c.weight:
            continue   # switched off — do not describe rules not in force
        ws.append([c.id.replace("_", " "), c.question, _plain_rule(c)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.alignment = wrap
    _section("Ranking logic")
    ws.append(["", "", "Companies failing any MUST-PASS check are excluded "
                       "(the 'why' column names the check and the value). "
                       "Survivors are ranked by the weighted ranking criteria; "
                       "'coverage' shows how much ranking data the company "
                       "disclosed. Ties break by coverage, then name. "
                       "IMPORTANT: 'Not disclosed' is never treated as 'No' — "
                       "unknowns stay in with a note, so a company can rank "
                       "highly while key facts are undisclosed (check its "
                       "notes)."])
    ws.cell(row=ws.max_row, column=3).alignment = wrap
    _section("Notes/flags you may see")
    for flag, meaning in _FLAG_GLOSSARY:
        ws.append(["", flag, meaning])
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=3).alignment = wrap
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 100


def _shortlist_sheet(wb, name: str, result) -> None:
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(name)
    gates = [c for c in result.spec.criteria if c.type == "hard_gate"]
    scored = [c for c in result.spec.criteria if c.type == "scored"]
    spec_by_id = {c.id: c for c in result.spec.criteria}
    hdr = ["rank", "entity", "status", "why", "total", "coverage", "excluded_by"]
    for g in gates:
        hdr += [f"gate:{g.id}", f"{g.id}:value", f"{g.id}:source_cell"]
    for s in scored:
        hdr += [f"{s.id}"]
    hdr += ["flags"]
    ws.append(hdr)
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in result.entities:
        status = ("SHORTLISTED" if r.rank is not None and r.rank <= result.k
                  else "RANKED" if r.rank is not None else "EXCLUDED")
        flags = sorted({f for v in r.gates for f in v.flags})
        row = [r.rank, r.entity, status, _entity_reason(r, spec_by_id),
               r.total if r.total is not None else "",
               r.coverage if r.rank is not None else "", r.excluded_by]
        for g in gates:
            v = next((x for x in r.gates if x.criterion_id == g.id), None)
            row += [v.label if v else "", v.parsed if v else "",
                    (v.raw[:120] if v else "")]
        for s in scored:
            sv = next((x for x in r.scores if x.criterion_id == s.id), None)
            row += [sv.label if sv else ""]
        row += ["; ".join(flags)]
        ws.append(row)
        fill = _STATUS_FILL.get(status)
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor=fill)
        for c in ws[ws.max_row]:
            c.alignment = wrap
    _style_header(ws)
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["D"].width = 60
    for i in range(5, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def write_shortlist_workbook(result, out: str, meta: list[tuple[str, str]]) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    info = wb.create_sheet("Run Info")
    info.append(["Shortlist run"])
    info["A1"].font = Font(bold=True, size=14)
    info.append([])
    for k, v in meta:
        info.append([k, v])
        info.cell(row=info.max_row, column=1).font = Font(bold=True)
        info.cell(row=info.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 100
    _howto_sheet(wb, result.spec)
    _shortlist_sheet(wb, "Shortlist", result)
    wb.save(out)
    print(f"Shortlist written: {out}")


# ---------------------------------------------------------------------------
# The experiment
# ---------------------------------------------------------------------------
def _perturbations(spec: Spec) -> list[tuple[str, Spec]]:
    """One-at-a-time threshold (±10 %, ±25 %) and weight (±25 %)
    perturbations, plus the missing-policy flip on both size gates."""
    out: list[tuple[str, Spec]] = []
    for c in spec.criteria:
        if c.type == "hard_gate" and c.threshold_hi is not None:
            for pct in (-0.25, -0.10, 0.10, 0.25):
                s = copy.deepcopy(spec)
                cc = next(x for x in s.criteria if x.id == c.id)
                cc.threshold_hi = c.threshold_hi * (1 + pct)
                out.append((f"{c.id} threshold {pct:+.0%}", s))
        if c.type == "scored" and c.weight:
            for pct in (-0.25, 0.25):
                s = copy.deepcopy(spec)
                cc = next(x for x in s.criteria if x.id == c.id)
                cc.weight = c.weight * (1 + pct)
                out.append((f"{c.id} weight {pct:+.0%}", s))
    flip = copy.deepcopy(spec)
    for cc in flip.criteria:
        if cc.type == "hard_gate" and cc.missing_policy == "flag":
            cc.missing_policy = "exclude"
    out.append(("missing-policy flip: all flag-gates -> exclude", flip))
    return out


def run_experiment(gt_path: str, ai_path: str, spec_path: str, out: str,
                   k: int, ai_sheet: str | None = None) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    spec = read_spec(spec_path)
    gt_cells, gt_kind = read_cells_auto(gt_path)
    ai_cells, ai_kind = read_cells_auto(ai_path, sheet=ai_sheet)
    gt_res = run_shortlist(gt_cells, spec, k=k, ref_year=REF_YEAR)
    ai_res = run_shortlist(ai_cells, spec, k=k, ref_year=REF_YEAR)

    diff = diff_gates(gt_res, ai_res)
    n_diverge = sum(1 for d in diff if d["class"] != "SAME")
    base_overlap = overlap_at_k(gt_res, ai_res, k)

    # Sensitivity: same perturbation applied to BOTH sides.
    sens_rows = []
    for label, pspec in _perturbations(spec):
        g2 = run_shortlist(gt_cells, pspec, k=k, ref_year=REF_YEAR)
        a2 = run_shortlist(ai_cells, pspec, k=k, ref_year=REF_YEAR)
        sens_rows.append({
            "scenario": label,
            "gt_vs_base": overlap_at_k(gt_res, g2, k),
            "ai_vs_base": overlap_at_k(ai_res, a2, k),
            "gt_vs_ai": overlap_at_k(g2, a2, k),
            "gt_top": ", ".join(top_k(g2, k)),
            "ai_top": ", ".join(top_k(a2, k)),
        })

    dec_gt, dec_ai = gate_decisiveness(gt_res), gate_decisiveness(ai_res)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wrap = Alignment(wrap_text=True, vertical="top")

    info = wb.create_sheet("Run Info")
    info.append(["Shortlist experiment — same spec, GT vs AI matrix"])
    info["A1"].font = Font(bold=True, size=14)
    info.append([])
    for kk, v in [
        ("GT input", f"{os.path.basename(gt_path)}  [{gt_kind}]"),
        ("AI input", f"{os.path.basename(ai_path)}  [{ai_kind}]"),
        ("Criteria spec", f"{os.path.basename(spec_path)} (identical file both runs)"),
        ("Run on", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Reference year (staleness)", str(REF_YEAR)),
        ("k", str(k)),
        ("Top-k overlap (base)", f"{base_overlap:.2f}"),
        ("Gate verdicts diverging", f"{n_diverge} of {len(diff)}"),
        ("Approved gate semantics",
         "independence: exclude on explicit verified No only, ND flags "
         "through (polarity recorded in spec, may flip). Giant gates: own "
         "size only — parent-attributed revenue passes with flag; "
         "missing_policy=flag (exclude would shortlist website "
         "transparency, not capability). Criteria are PLACEHOLDERS "
         "(mechanism test)."),
        ("Caveat",
         "AI cells reconstructed from the eval report's Detail sheet on "
         "this machine: verified/unverified and conflict markers are "
         "unavailable, so unverified_policy cannot bite on the AI side. "
         "Re-run against the raw pipeline workbook on the laptop for "
         "marker-aware verdicts." if "eval_detail" in ai_kind else "—"),
    ]:
        info.append([kk, v])
        info.cell(row=info.max_row, column=1).font = Font(bold=True)
        info.cell(row=info.max_row, column=2).alignment = wrap
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 110

    # Top-k side by side
    ws = wb.create_sheet("Top K")
    ws.append(["rank", "GT shortlist", "AI shortlist", "in both?"])
    gt_top, ai_top = top_k(gt_res, k), top_k(ai_res, k)
    for i in range(max(len(gt_top), len(ai_top))):
        g = gt_top[i] if i < len(gt_top) else ""
        a = ai_top[i] if i < len(ai_top) else ""
        both = "YES" if g and g in ai_top else ("yes (AI)" if a and a in gt_top else "")
        ws.append([i + 1, g, a, both])
    _style_header(ws)
    ws.column_dimensions["B"].width = ws.column_dimensions["C"].width = 30

    _howto_sheet(wb, spec)
    _shortlist_sheet(wb, "Shortlist GT", gt_res)
    _shortlist_sheet(wb, "Shortlist AI", ai_res)

    # Gate diff — every disagreement traced to its cell
    ws = wb.create_sheet("Gate Diff")
    hdr = ["entity", "gate", "class", "gt_verdict", "gt_value", "ai_verdict",
           "ai_value", "gt_raw_cell", "ai_raw_cell"]
    ws.append(hdr)
    for d in sorted(diff, key=lambda x: (x["class"] == "SAME", x["gate"], x["entity"])):
        ws.append([d["entity"], d["gate"], d["class"], d["gt_verdict"],
                   d["gt_value"], d["ai_verdict"], d["ai_value"],
                   d["gt_raw"][:150], d["ai_raw"][:150]])
        base_cls = d["class"].split(" (")[0]
        fill = _DIFF_FILL.get(base_cls)
        if fill and d["class"] != "SAME":
            ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor=fill)
        for c in ws[ws.max_row]:
            c.alignment = wrap
    _style_header(ws)
    ws.column_dimensions["A"].width = 24
    for col in ("D", "E", "F", "G"):
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["H"].width = ws.column_dimensions["I"].width = 40

    # Sensitivity
    ws = wb.create_sheet("Sensitivity")
    ws.append(["scenario", "GT overlap vs base", "AI overlap vs base",
               "GT vs AI overlap", "GT top-k", "AI top-k"])
    for s in sens_rows:
        ws.append([s["scenario"], s["gt_vs_base"], s["ai_vs_base"],
                   s["gt_vs_ai"], s["gt_top"], s["ai_top"]])
        for c in ws[ws.max_row]:
            c.alignment = wrap
    _style_header(ws)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["E"].width = ws.column_dimensions["F"].width = 55

    # Decisiveness — the trivial-agreement guard
    ws = wb.create_sheet("Gate Decisiveness")
    ws.append(["gate", "side", "decided_on_data", "of_which_FAIL",
               "flagged_through", "decisive_%"])
    for gid in sorted(set(dec_gt) | set(dec_ai)):
        for side, dec in (("GT", dec_gt), ("AI", dec_ai)):
            d = dec.get(gid, {})
            n = d.get("decided_on_data", 0) + d.get("flagged_through", 0)
            ws.append([gid, side, d.get("decided_on_data", 0), d.get("failed", 0),
                       d.get("flagged_through", 0),
                       round(d.get("decided_on_data", 0) / n, 3) if n else ""])
    _style_header(ws)

    wb.save(out)
    print(f"\nExperiment report: {out}")
    print(f"  GT top-{k}: {gt_top}")
    print(f"  AI top-{k}: {ai_top}")
    print(f"  overlap@{k}: {base_overlap:.2f} | diverging gate verdicts: "
          f"{n_diverge}/{len(diff)}")


# ---------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="Shortlist layer CLI")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("init-spec", help="write the default criteria workbook")
    p.add_argument("--out", default=DEFAULT_SPEC)
    p.add_argument("--force", action="store_true")

    p = sub.add_parser("run", help="shortlist one workbook")
    p.add_argument("workbook")
    p.add_argument("--spec", default=DEFAULT_SPEC)
    p.add_argument("--out", default=None)
    p.add_argument("--top", type=int, default=5)
    p.add_argument("--sheet", default=None,
                   help="explicit sheet to read (e.g. 'AI Summary'); default "
                        "auto-detects Matrix, then AI Summary")

    p = sub.add_parser("experiment",
                       help="same spec twice: GT vs AI, diff + sensitivity")
    p.add_argument("gt_workbook")
    p.add_argument("ai_workbook")
    p.add_argument("--spec", default=DEFAULT_SPEC)
    p.add_argument("--out", default=None)
    p.add_argument("--top", type=int, default=5)
    p.add_argument("--ai-sheet", default=None,
                   help="explicit sheet to read from the AI workbook")

    args = ap.parse_args()

    if args.cmd == "init-spec":
        if os.path.exists(args.out) and not args.force:
            sys.exit(f"{args.out} exists (use --force to overwrite)")
        os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
        write_default_spec(args.out)
        print(f"Spec written: {args.out}")
        return 0

    if not os.path.exists(args.spec):
        sys.exit(f"Spec not found: {args.spec} — run init-spec first")

    if args.cmd == "run":
        cells, kind = read_cells_auto(args.workbook, sheet=args.sheet)
        spec = read_spec(args.spec)
        res = run_shortlist(cells, spec, k=args.top, ref_year=REF_YEAR)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out = args.out or os.path.join(
            "outputs", f"shortlist_{os.path.splitext(os.path.basename(args.workbook))[0]}_{stamp}.xlsx")
        write_shortlist_workbook(res, out, [
            ("Input", f"{os.path.basename(args.workbook)}  [{kind}]"),
            ("Spec", os.path.basename(args.spec)),
            ("Run on", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Top-k", ", ".join(top_k(res))),
        ])
        return 0

    if args.cmd == "experiment":
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out = args.out or os.path.join("outputs", f"shortlist_experiment_{stamp}.xlsx")
        run_experiment(args.gt_workbook, args.ai_workbook, args.spec, out,
                       args.top, ai_sheet=args.ai_sheet)
        return 0
    return 2


if __name__ == "__main__":
    sys.exit(main())
