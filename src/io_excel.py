import os
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import MATRIX_MAX_DISPLAY_ITEMS
from models import ColumnSpec, PipelineInput, PipelineResult, UrlSpec

# Excel's hard per-cell character limit; text beyond it is silently truncated
# by openpyxl/Excel. We clamp below it with an explicit marker instead.
_EXCEL_CELL_MAX = 32767
_TRUNCATION_MARKER = "\n[truncated — full list in Provenance]"


def _clamp_cell_text(text: str) -> str:
    """Keep cell text under Excel's 32,767-char limit, marked, never silent."""
    if len(text) <= _EXCEL_CELL_MAX:
        return text
    keep = _EXCEL_CELL_MAX - len(_TRUNCATION_MARKER)
    clipped = text[:keep]
    # Cut on a line boundary so we never show half a claim.
    if "\n" in clipped:
        clipped = clipped.rsplit("\n", 1)[0]
    return clipped + _TRUNCATION_MARKER

# Colour palette
_HEADER_FILL = "2E4057"
_HEADER_FONT = "FFFFFF"
_ALT_ROW = "F5F5F5"
_RED_FILL = "FFCDD2"
_RED_FONT = "C62828"
_ORANGE_FILL = "FFE0B2"
_LORANGE_FILL = "FFF3E0"  # light orange for mixed verified/unverified

_TAB_COLORS = {
    "Summary": "2E4057",
    "Matrix": "4CAF50",
    "Digest": "8BC34A",
    "Provenance": "009688",
    "Grouped Themes": "8BC34A",
    "Acquire Log": "FF9800",
    "Crawl Candidates": "FF9800",
    "Filter Log": "2196F3",
    "Extract Log": "F44336",
    "Verify Log": "9C27B0",
}

_SUPPORTED_CONFIG_KEYS = {
    "ACQUIRE_TOOL",
    "EXTRACT_TOOL",
    "CRAWL_MIN_SCORE",
    "CRAWL_MIN_SCORE_EMBED",
    "CRAWL_SCORER",
    "CRAWL_MAX_PAGES",
    "DEFAULT_DEPTH",
}


# Public helpers used by main.py

def _clean_str(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _find_sheet(xls: pd.ExcelFile, target: str) -> str | None:
    target_lower = target.lower()
    for sheet_name in xls.sheet_names:
        if sheet_name.strip().lower() == target_lower:
            return sheet_name
    return None


def _find_column(df: pd.DataFrame, target: str) -> str | None:
    target_lower = target.lower()
    for col in df.columns:
        if str(col).strip().lower() == target_lower:
            return col
    return None


def _find_url_column(df: pd.DataFrame) -> str:
    for col in df.columns:
        name = str(col).strip().lower()
        if "url" in name or "link" in name:
            return col
    if len(df.columns) == 0:
        raise ValueError("URL sheet must contain a url column")
    return df.columns[0]


def _parse_depth(value: Any) -> int:
    if value is None or pd.isna(value) or str(value).strip() == "":
        return 0
    try:
        depth = int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Depth must be an integer, got {value!r}") from exc
    if depth not in {0, 1, 2}:
        raise ValueError(f"Depth must be 0, 1, or 2, got {depth!r}")
    return depth


def _parse_entity_list(value: Any) -> list[str]:
    raw = _clean_str(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(",") if part.strip()]


def _read_entities_sheet(xls: pd.ExcelFile, sheet_name: str) -> list[str]:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    entity_col = _find_column(df, "entity") or (df.columns[0] if len(df.columns) else None)
    if entity_col is None:
        return []

    entities: list[str] = []
    seen: set[str] = set()
    for value in df[entity_col].tolist():
        entity = _clean_str(value)
        if entity and entity not in seen:
            entities.append(entity)
            seen.add(entity)
    return entities


def _read_urls_sheet(xls: pd.ExcelFile, sheet_name: str) -> list[UrlSpec]:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    url_col = _find_url_column(df)
    depth_col = _find_column(df, "depth")
    entities_col = _find_column(df, "entities")

    url_specs: list[UrlSpec] = []
    for _, row in df.iterrows():
        url = _clean_str(row.get(url_col))
        if not url:
            continue

        depth = _parse_depth(row.get(depth_col)) if depth_col is not None else 0
        entities = _parse_entity_list(row.get(entities_col)) if entities_col is not None else []
        url_specs.append(UrlSpec(url=url, depth=depth, entities=entities))

    return url_specs


def _read_questions_sheet(xls: pd.ExcelFile, sheet_name: str) -> list[ColumnSpec]:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    question_col = _find_column(df, "question") or (df.columns[0] if len(df.columns) else None)
    if question_col is None:
        return []

    instructions_col = _find_column(df, "instructions")
    columns: list[ColumnSpec] = []
    for _, row in df.iterrows():
        question = _clean_str(row.get(question_col))
        if not question:
            continue
        instruction = _clean_str(row.get(instructions_col)) if instructions_col is not None else ""
        columns.append(ColumnSpec(name=question, instruction=instruction or None))
    return columns


def _coerce_config_value(key: str, value: Any) -> Any:
    if value is None or pd.isna(value):
        return None

    if key in {"ACQUIRE_TOOL", "EXTRACT_TOOL", "CRAWL_SCORER"}:
        return str(value).strip()

    if key in {"CRAWL_MAX_PAGES", "DEFAULT_DEPTH"}:
        return int(float(value))

    if key in {"CRAWL_MIN_SCORE", "CRAWL_MIN_SCORE_EMBED"}:
        return float(value)

    return value


def _read_config_sheet(xls: pd.ExcelFile, sheet_name: str) -> dict[str, Any]:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    setting_col = _find_column(df, "setting") or (df.columns[0] if len(df.columns) else None)
    value_col = _find_column(df, "value") or (df.columns[1] if len(df.columns) > 1 else None)
    if setting_col is None or value_col is None:
        return {}

    overrides: dict[str, Any] = {}
    for _, row in df.iterrows():
        key = _clean_str(row.get(setting_col)).upper()
        if not key:
            continue
        if key not in _SUPPORTED_CONFIG_KEYS:
            raise ValueError(
                f"Unsupported config setting {key!r}. Supported settings: "
                + ", ".join(sorted(_SUPPORTED_CONFIG_KEYS))
            )
        value = _coerce_config_value(key, row.get(value_col))
        if value is not None:
            overrides[key] = value
    return overrides


def _resolve_url_entities(urls: list[UrlSpec], entities: list[str], has_entities_sheet: bool) -> None:
    if has_entities_sheet:
        entity_set = set(entities)
        for spec in urls:
            if not spec.entities:
                spec.entities = list(entities)
                continue
            unknown = [entity for entity in spec.entities if entity not in entity_set]
            if unknown:
                raise ValueError(
                    f"URL {spec.url!r} references unknown entities: {', '.join(unknown)}"
                )
        return

    # Backward compatibility: without an entities sheet, each URL is its own entity.
    for spec in urls:
        spec.entities = [spec.url]


def read_input(filepath: str) -> PipelineInput:
    """Read the four-sheet input workbook.

    Expected sheets:
    - entities: entity
    - urls: url, depth, entities
    - questions: question, instructions
    - config: setting, value

    If the workbook has no entities sheet, each URL becomes its own entity.
    """
    xls = pd.ExcelFile(filepath)
    try:
        entities_sheet = _find_sheet(xls, "entities")
        urls_sheet = _find_sheet(xls, "urls")
        questions_sheet = _find_sheet(xls, "questions")
        config_sheet = _find_sheet(xls, "config")

        has_entities_sheet = entities_sheet is not None

        if urls_sheet is None:
            if has_entities_sheet:
                raise ValueError("Input workbook must contain a urls sheet")
            urls_sheet = xls.sheet_names[0]

        entities = _read_entities_sheet(xls, entities_sheet) if entities_sheet else []
        urls = _read_urls_sheet(xls, urls_sheet)

        if not entities:
            entities = []
            seen_urls: set[str] = set()
            for spec in urls:
                if spec.url not in seen_urls:
                    entities.append(spec.url)
                    seen_urls.add(spec.url)

        _resolve_url_entities(urls, entities, has_entities_sheet)

        columns = _read_questions_sheet(xls, questions_sheet) if questions_sheet else []
        config_overrides = _read_config_sheet(xls, config_sheet) if config_sheet else {}
    finally:
        xls.close()

    return PipelineInput(
        entities=entities,
        urls=urls,
        columns=columns,
        config_overrides=config_overrides,
    )


def read_urls_from_excel(filepath: str) -> list[tuple[str, int]]:
    """Backward-compatible URL reader."""
    xls = pd.ExcelFile(filepath)
    try:
        urls_sheet = _find_sheet(xls, "urls") or xls.sheet_names[0]
        return [(spec.url, spec.depth) for spec in _read_urls_sheet(xls, urls_sheet)]
    finally:
        xls.close()


def parse_columns(raw_columns: list[str]) -> list[ColumnSpec]:
    """Parse old terminal-style question specs."""
    columns = []
    for raw in raw_columns:
        if ":" in raw:
            name, instruction = raw.split(":", 1)
            columns.append(ColumnSpec(name=name.strip(), instruction=instruction.strip()))
        else:
            columns.append(ColumnSpec(name=raw.strip()))
    return columns


# DataFrame builders

def _make_df(rows: list[dict], keys: list[str], col_names: list[str]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=col_names)
    records = [{col: row.get(key, "") for col, key in zip(col_names, keys)} for row in rows]
    return pd.DataFrame(records, columns=col_names)


def _make_summary_df(diag: dict | None, result: PipelineResult) -> pd.DataFrame:
    col_names = [
        "Entity", "Pages Fetched", "Pages Crawled", "Total Claims Found",
        "Claims Verified", "Claims Unverified", "Cells With No Data",
        "Total Fetch Time", "Total Extract Time", "Acquire Tool Used", "Extract Tool Used",
    ]
    keys = [
        "entity", "pages_fetched", "pages_crawled", "total_claims_found",
        "claims_verified", "claims_unverified", "cells_with_no_data",
        "total_fetch_time", "total_extract_time", "acquire_tool_used", "extract_tool_used",
    ]
    if diag and diag.get("summary"):
        return _make_df(diag["summary"], keys, col_names)

    rows = []
    for row in result.rows:
        src = row.all_cells  # always raw — never read aggregated cells here
        all_ev = [e for c in src for e in c.evidence]
        rows.append({
            "entity": row.entity,
            "pages_fetched": len({c.source_url for c in src}),
            "pages_crawled": "",
            "total_claims_found": len(all_ev),
            "claims_verified": sum(1 for e in all_ev if e.verified),
            "claims_unverified": sum(1 for e in all_ev if not e.verified),
            "cells_with_no_data": sum(1 for c in src if not c.evidence),
            "total_fetch_time": "",
            "total_extract_time": "",
            "acquire_tool_used": "",
            "extract_tool_used": "",
        })
    return pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)


def _match_type_str(verified: bool, score) -> str:
    if score is None or score == "":
        return "none"
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "none"
    if s >= 100:
        return "exact"
    if verified:
        return "fuzzy"
    return "none"


def _make_matrix_df(
    result: PipelineResult,
    columns: list[ColumnSpec],
) -> tuple[pd.DataFrame, dict]:
    """Returns (DataFrame, cell_fills) where cell_fills maps (row, col) to hex color."""
    col_order = ["Entity"] + [c.name for c in columns]
    matrix_rows = []
    fills: dict[tuple[int, int], str] = {}

    for row_data_idx, row in enumerate(result.rows):
        excel_row = row_data_idx + 2  # 1-indexed + header
        src = row.cells if row.cells else row.all_cells  # prefer aggregated/deduplicated
        matrix_row: dict = {"Entity": row.entity}

        for col_excel_idx, col in enumerate(columns, start=2):
            # Aggregated src has at most one cell per (entity, column).
            agg_cell = next((c for c in src if c.column == col.name), None)

            if agg_cell is None or not agg_cell.evidence:
                text = "No data found"
                fills[(excel_row, col_excel_idx)] = _RED_FILL
                matrix_row[col.name] = text
                continue

            # Look up verified status from ranked evidence, but render only the
            # fuzzy-deduped values from agg_cell.value so _DEDUP_RATIO takes effect.
            ev_verified: dict[str, bool] = {}
            for ev in agg_cell.evidence:
                val_str = str(ev.value).strip() if ev.value is not None else ""
                if val_str and val_str not in ev_verified:
                    ev_verified[val_str] = ev.verified

            raw_val = agg_cell.value
            if isinstance(raw_val, list):
                display_vals = [str(v).strip() for v in raw_val if v not in (None, "", [])]
            elif raw_val not in (None, "", []):
                display_vals = [str(raw_val).strip()]
            else:
                display_vals = []
            verified_vals = [v for v in display_vals if ev_verified.get(v, agg_cell.verified)]
            unverified_vals = [v for v in display_vals if not ev_verified.get(v, agg_cell.verified)]

            # Cap rendered items (verified kept preferentially); everything
            # remains in Provenance and the overflow is stated in the cell.
            hidden = 0
            total_items = len(verified_vals) + len(unverified_vals)
            if total_items > MATRIX_MAX_DISPLAY_ITEMS:
                hidden = total_items - MATRIX_MAX_DISPLAY_ITEMS
                keep_v = min(len(verified_vals), MATRIX_MAX_DISPLAY_ITEMS)
                verified_vals = verified_vals[:keep_v]
                unverified_vals = unverified_vals[:MATRIX_MAX_DISPLAY_ITEMS - keep_v]

            if not verified_vals and not unverified_vals:
                text = "No data found"
                fills[(excel_row, col_excel_idx)] = _RED_FILL
            elif agg_cell.has_conflict:
                lines = ["- " + v for v in verified_vals]
                if verified_vals and unverified_vals:
                    lines += ["-- Unverified --"]
                lines += ["- " + v for v in unverified_vals]
                text = "(sources conflict)\n" + "\n".join(lines)
                fills[(excel_row, col_excel_idx)] = _ORANGE_FILL
            elif not verified_vals:
                lines = ["- " + v for v in unverified_vals] + ["(unverified)"]
                text = "\n".join(lines)
                fills[(excel_row, col_excel_idx)] = _ORANGE_FILL
            elif unverified_vals:
                lines = ["- " + v for v in verified_vals]
                lines += ["-- Unverified --"] + ["- " + v for v in unverified_vals]
                text = "\n".join(lines)
                fills[(excel_row, col_excel_idx)] = _LORANGE_FILL
            else:
                text = "\n".join("- " + v for v in verified_vals)

            if hidden:
                text += f"\n[+{hidden} more items — see Provenance]"
            matrix_row[col.name] = _clamp_cell_text(text)

        matrix_rows.append(matrix_row)

    df = pd.DataFrame(matrix_rows, columns=col_order) if matrix_rows else pd.DataFrame(columns=col_order)
    return df, fills


def _norm_claim(text) -> str:
    # Mirrors aggregate.py/group.py _normalise_value so claim-ID lookups from
    # the Grouped Themes / Digest builders land on the same Provenance rows.
    return " ".join(str(text).strip().lower().split())


def _make_provenance_df(
    result: PipelineResult,
    columns: list[ColumnSpec],
    diag: dict | None,
) -> tuple[pd.DataFrame, dict]:
    """Returns (df, claim_index).

    claim_index maps (entity, question, normalised claim) -> (claim_id,
    provenance_excel_row) — the anchor the Grouped Themes and Digest sheets
    hyperlink back to. The anchor is the first VERIFIED occurrence of each
    claim, falling back to the first occurrence when no verified one exists
    (standing decision 2026-07-06: grouping/digest cite verified rows only;
    since group.py feeds them verified claims only, the fallback is never
    consultant-visible). Claim IDs are sequential in Provenance order, which
    is deterministic (spec merge order).
    """
    col_names = [
        "Claim ID", "Entity", "Source URL", "Question", "Claim", "Verbatim Quote",
        "Page Title", "Extraction Method", "Confidence Score", "Verified",
        "Verification Score", "Match Type", "Semantic Score", "Source Page Depth",
    ]

    url_to_depth: dict[str, int] = {}
    if diag:
        for r in diag.get("acquire_log", []):
            url_to_depth[r.get("page_url", "")] = r.get("depth", 0)

    rows = []
    claim_index: dict[tuple[str, str, str], tuple[str, int]] = {}
    verified_anchor_keys: set[tuple[str, str, str]] = set()
    for entity_row in result.rows:
        src = entity_row.all_cells  # always granular — never read aggregated cells here
        for cell in src:
            for ev in cell.evidence:
                if ev.value is None:
                    continue
                score = ev.verification_score
                verified = ev.verified
                source_url = ev.source_url or cell.source_url
                claim_id = f"C{len(rows) + 1:04d}"
                entity = cell.entity or entity_row.entity
                key = (entity, cell.column, _norm_claim(ev.value))
                if key not in claim_index or (verified and key not in verified_anchor_keys):
                    claim_index[key] = (claim_id, len(rows) + 2)  # +2: 1-based + header
                    if verified:
                        verified_anchor_keys.add(key)
                rows.append({
                    "Claim ID": claim_id,
                    "Entity": entity,
                    "Source URL": source_url,
                    "Question": cell.column,
                    "Claim": str(ev.value),
                    "Verbatim Quote": ev.quote or "",
                    "Page Title": ev.page_title,
                    "Extraction Method": ev.extraction_method,
                    "Confidence Score": round(ev.confidence_score, 3) if isinstance(ev.confidence_score, float) else "",
                    "Verified": verified,
                    "Verification Score": round(score, 1) if isinstance(score, float) else "",
                    "Match Type": _match_type_str(verified, score),
                    "Semantic Score": round(ev.semantic_score, 3) if isinstance(ev.semantic_score, float) else "",
                    "Source Page Depth": url_to_depth.get(source_url, 0),
                })

    df = pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)
    return df, claim_index


def _group_claim_refs(group: dict, claim_index: dict) -> tuple[list[tuple[str, str]], tuple[str, int] | None]:
    """Resolve a theme's member values against the Provenance claim index.

    Returns ([(value, claim_id_or_empty)], anchor) where anchor is the
    (claim_id, provenance_excel_row) of the theme-label claim itself, falling
    back to the first member that resolved — the row the theme hyperlinks to.
    """
    entity = group.get("entity", "")
    question = group.get("question", "")

    def _lookup(value) -> tuple[str, int] | None:
        return claim_index.get((entity, question, _norm_claim(value)))

    pairs: list[tuple[str, str]] = []
    anchor = _lookup(group.get("theme", ""))
    for v in group.get("values", []):
        if v in (None, "", []):
            continue
        hit = _lookup(v)
        pairs.append((str(v).strip(), hit[0] if hit else ""))
        if anchor is None and hit is not None:
            anchor = hit
    return pairs, anchor


def _make_grouped_themes_df(
    claim_groups: list[dict],
    claim_index: dict,
) -> tuple[pd.DataFrame, dict[int, int]]:
    """Grouped Themes sheet: deterministic claim clusters per aggregated cell.

    Traceability (Advisory requirement): every bullet carries its Provenance
    claim ID (`[C0042]`), a "Claim IDs" column lists the theme's references,
    and the returned theme_links map {sheet_excel_row: provenance_excel_row}
    is used post-write to hyperlink each Theme cell to its anchor claim.

    Mirrors the Matrix writer's display conventions without touching it:
    bullets capped at MATRIX_MAX_DISPLAY_ITEMS with the same overflow marker,
    final text clamped below Excel's hard cell limit. Every member claim
    remains fully listed in Provenance.
    """
    col_names = ["Entity", "Question", "Theme", "Items", "Values", "Claim IDs", "Distinct Sources"]
    rows = []
    theme_links: dict[int, int] = {}
    for group in claim_groups:
        pairs, anchor = _group_claim_refs(group, claim_index)
        # Claim IDs column must stay complete even when the bullet display is
        # capped — it is the traceability escape hatch for oversized cells
        # (e.g. HORIBA's 328-item Recent news theme), so truncating it in
        # lockstep with the display bullets silently dropped every ID past
        # MATRIX_MAX_DISPLAY_ITEMS, leaving the "+N more — see Provenance"
        # items with no ID to search Provenance by. Compute ids from the
        # FULL pairs list, before slicing for display.
        all_ids = [cid for _, cid in pairs if cid]
        hidden = 0
        display_pairs = pairs
        if len(pairs) > MATRIX_MAX_DISPLAY_ITEMS:
            hidden = len(pairs) - MATRIX_MAX_DISPLAY_ITEMS
            display_pairs = pairs[:MATRIX_MAX_DISPLAY_ITEMS]
        bullets = [
            f"- {value} [{cid}]" if cid else f"- {value}"
            for value, cid in display_pairs
        ]
        text = "\n".join(bullets)
        if hidden:
            text += f"\n[+{hidden} more items — see Provenance]"

        ids_text = ", ".join(all_ids)

        excel_row = len(rows) + 2  # 1-based + header
        if anchor is not None:
            theme_links[excel_row] = anchor[1]
        rows.append({
            "Entity": group.get("entity", ""),
            "Question": group.get("question", ""),
            "Theme": group.get("theme", ""),
            "Items": group.get("n_items", ""),
            "Values": _clamp_cell_text(text),
            "Claim IDs": _clamp_cell_text(ids_text),
            "Distinct Sources": group.get("sources", ""),
        })
    df = pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)
    return df, theme_links


def _make_digest_df(
    claim_groups: list[dict],
    claim_index: dict,
) -> tuple[pd.DataFrame, dict[int, int]]:
    """Digest sheet: one deterministic template line per grouped cell.

    NO LLM — the text is assembled mechanically from the theme structure, so
    it is faithful and traceable by construction: theme labels are verbatim
    member claims and each carries its Provenance claim ID. Returns
    (df, digest_links) where digest_links maps {digest_excel_row:
    grouped_themes_excel_row} for post-write hyperlinking.
    """
    from src.group import ALL_ITEMS_THEME

    col_names = ["Entity", "Question", "Items", "Themes", "Digest"]
    # Preserve claim_groups order; remember each cell's first Grouped Themes row.
    cells: dict[tuple[str, str], dict] = {}
    for i, group in enumerate(claim_groups):
        key = (group.get("entity", ""), group.get("question", ""))
        entry = cells.setdefault(key, {"groups": [], "first_row": i + 2})
        entry["groups"].append(group)

    rows = []
    digest_links: dict[int, int] = {}
    for (entity, question), entry in cells.items():
        groups = entry["groups"]
        total = sum(g.get("n_items", 0) for g in groups)
        real = [g for g in groups if g.get("theme") != ALL_ITEMS_THEME]
        if not real:
            digest = f"{total} items (below grouping threshold — see Grouped Themes)."
        else:
            tops = []
            for g in real[:3]:  # groups arrive size-desc from group_rows
                label = str(g.get("theme", "")).strip()
                hit = claim_index.get((entity, question, _norm_claim(label)))
                ref = f" [{hit[0]}]" if hit else ""
                tops.append(f"“{label}” ({g.get('n_items', '?')} items){ref}")
            digest = f"{total} items across {len(groups)} themes. Top: " + "; ".join(tops) + "."
        excel_row = len(rows) + 2
        digest_links[excel_row] = entry["first_row"]
        rows.append({
            "Entity": entity,
            "Question": question,
            "Items": total,
            "Themes": len(groups),
            "Digest": _clamp_cell_text(digest),
        })
    df = pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)
    return df, digest_links


# Workbook formatting

def _style_sheet(ws, tab_color: str, matrix_fills: dict | None = None) -> None:
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"

    h_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    h_font = Font(name="Arial", bold=True, color=_HEADER_FONT, size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
    ws.row_dimensions[1].height = 28

    verified_cols: set[int] = set()
    for cell in ws[1]:
        if cell.value and "verified" in str(cell.value).lower():
            verified_cols.add(cell.column)

    b_font = Font(name="Arial", size=10)
    b_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alt = PatternFill("solid", fgColor=_ALT_ROW)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = b_font
            cell.alignment = b_align
            if row_idx % 2 == 0:
                cell.fill = alt

    if matrix_fills:
        for (r, c), hex_color in matrix_fills.items():
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=hex_color)
            if hex_color == _RED_FILL:
                cell.font = Font(name="Arial", size=10, color=_RED_FONT)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            vl = val.lower()

            if "no data found" in vl:
                cell.fill = PatternFill("solid", fgColor=_RED_FILL)
                cell.font = Font(name="Arial", size=10, color=_RED_FONT)
            elif cell.column in verified_cols and vl == "false":
                # Excel stores booleans as Python bool -> str(False) == "False",
                # so an exact "FALSE" comparison never matched; compare on the
                # lowercased value. This is the analyst-review flag for
                # unverified Provenance / Verify Log rows.
                cell.fill = PatternFill("solid", fgColor=_ORANGE_FILL)
            elif any(kw in vl for kw in ("timed out", "timeout", "status: error")):
                cell.fill = PatternFill("solid", fgColor=_RED_FILL)

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                longest = max((len(line) for line in str(cell.value).split("\n")), default=0)
                max_len = max(max_len, longest)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    if ws.title == "Matrix":
        for row_idx in range(2, ws.max_row + 1):
            max_lines = 1
            for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
                if col[0].value:
                    max_lines = max(max_lines, str(col[0].value).count("\n") + 1)
            ws.row_dimensions[row_idx].height = min(max(max_lines * 14, 16), 200)


# Main entry point

def write_output_excel(
    result: PipelineResult,
    columns: list[ColumnSpec],
    output_path: str,
    diag: dict | None = None,
) -> None:
    try:
        from config import DIAGNOSTICS
    except Exception:
        DIAGNOSTICS = False

    write_diag = DIAGNOSTICS and diag is not None

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    summary_df = _make_summary_df(diag, result)
    matrix_df, matrix_fills = _make_matrix_df(result, columns)
    provenance_df, claim_index = _make_provenance_df(result, columns, diag)

    acq_col_keys = [
        "entities", "seed_url", "page_url", "parent_url", "depth", "crawl_score",
        "above_threshold", "fetch_tool", "crawl_scorer", "page_length", "fetch_time_ms",
        "from_cache", "status", "skip_reason",
    ]
    acq_col_names = [
        "Entities", "Seed URL", "Page URL", "Parent URL", "Depth", "Crawl Score",
        "Above Threshold", "Fetch Tool", "Crawl Scorer", "Page Length (chars)", "Fetch Time (ms)",
        "From Cache", "Status", "Skip Reason",
    ]

    cand_col_keys = [
        "seed_url", "entities", "parent_url", "candidate_url", "anchor_text",
        "url_path", "crawl_score", "crawl_scorer", "threshold", "followed", "skip_reason",
    ]
    cand_col_names = [
        "Seed URL", "Entities", "Parent URL", "Candidate URL", "Anchor Text",
        "URL Path", "Crawl Score", "Crawl Scorer", "Threshold", "Followed", "Skip Reason",
    ]

    flt_col_keys = [
        "url", "column", "embedding_score", "keyword_gate", "included", "reason",
    ]
    flt_col_names = [
        "URL", "Column", "Embedding Score", "Keyword Gate", "Included", "Reason",
    ]

    ext_col_keys = [
        "entity", "source_url", "question", "extract_tool", "items_extracted",
        "extraction_time_ms", "timed_out", "retry_count", "page_length_input",
        "raw_answer_preview",
    ]
    ext_col_names = [
        "Entity", "Source URL", "Question", "Extract Tool", "Items Extracted",
        "Extraction Time (ms)", "Timed Out", "Retry Count", "Page Length Input (chars)",
        "Raw Answer Preview",
    ]

    ver_col_keys = [
        "entity", "source_url", "question", "claim_preview", "quote_preview",
        "verified", "match_type", "verification_score", "semantic_score", "verifier_tool",
    ]
    ver_col_names = [
        "Entity", "Source URL", "Question", "Claim Preview", "Quote Preview",
        "Verified", "Match Type", "Verification Score", "Semantic Score", "Verifier Tool",
    ]

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Summary", summary_df),
        ("Matrix", matrix_df),
        ("Provenance", provenance_df),
    ]

    # Digest + Grouped Themes are deliverable-facing (consultant view of big
    # cells), written whenever grouping produced rows — NOT gated on
    # DIAGNOSTICS. Traceability chain: Digest -> Grouped Themes -> Provenance
    # (claim IDs + internal hyperlinks) -> source URL.
    claim_groups = (diag or {}).get("claim_groups") or []
    theme_links: dict[int, int] = {}
    digest_links: dict[int, int] = {}
    if claim_groups:
        themes_df, theme_links = _make_grouped_themes_df(claim_groups, claim_index)
        digest_df, digest_links = _make_digest_df(claim_groups, claim_index)
        sheets.insert(2, ("Digest", digest_df))  # after Matrix, before Provenance
        sheets.append(("Grouped Themes", themes_df))

    if write_diag:
        sheets += [
            ("Acquire Log", _make_df(diag.get("acquire_log", []), acq_col_keys, acq_col_names)),
            ("Crawl Candidates", _make_df(diag.get("crawl_candidates", []), cand_col_keys, cand_col_names)),
            ("Filter Log", _make_df(diag.get("filter_log", []), flt_col_keys, flt_col_names)),
            ("Extract Log", _make_df(diag.get("extract_log", []), ext_col_keys, ext_col_names)),
            ("Verify Log", _make_df(diag.get("verify_log", []), ver_col_keys, ver_col_names)),
        ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        fills = matrix_fills if ws.title == "Matrix" else None
        _style_sheet(ws, _TAB_COLORS.get(ws.title, _HEADER_FILL), matrix_fills=fills)

    # Traceability hyperlinks (applied after styling so the link font wins).
    link_font = Font(name="Arial", size=9, color="0563C1", underline="single")
    if theme_links and "Grouped Themes" in wb.sheetnames:
        ws = wb["Grouped Themes"]
        for row, prov_row in theme_links.items():
            c = ws.cell(row=row, column=3)  # Theme
            c.hyperlink = f"#Provenance!A{prov_row}"
            c.font = link_font
    if digest_links and "Digest" in wb.sheetnames and "Grouped Themes" in wb.sheetnames:
        ws = wb["Digest"]
        for row, gt_row in digest_links.items():
            c = ws.cell(row=row, column=2)  # Question
            c.hyperlink = f"#'Grouped Themes'!A{gt_row}"
            c.font = link_font
    if "Provenance" in wb.sheetnames:
        # Source URL column (C) as real links — the last hop of the chain.
        # Excel caps workbook hyperlinks around 65k; stay well below.
        ws = wb["Provenance"]
        for row in range(2, min(ws.max_row, 20000) + 1):
            c = ws.cell(row=row, column=3)
            url = str(c.value or "")
            if url.startswith("http"):
                c.hyperlink = url
                c.font = link_font

    wb.save(output_path)
