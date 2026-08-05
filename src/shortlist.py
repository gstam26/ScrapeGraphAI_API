"""
Shortlist layer — criteria-driven ranked shortlist from a deliverable Matrix.

Design: brain/proposals/shortlist-layer.md (approved 2026-08-05). Mechanism
test, not a client deliverable: criteria are PLACEHOLDERS to play with.

George's approved gate semantics (2026-08-05), recorded here and in the spec:
  * independence: EXCLUDE on explicit verified "No" (acquired companies out);
    "Not disclosed" / no-data / unparseable FLAG THROUGH. Polarity may flip
    later — it is spec data, not code.
  * giant gate (employees and/or revenue): measures the ENTITY'S OWN size
    only. Parent-attributed revenue (a non-year parenthetical naming another
    company) NEVER trips it — PASS with PARENT_ATTRIBUTED flag; the
    independence gate handles those cases. Attribution undeterminable →
    PARSE_FAIL with reason.
  * missing_policy = flag for both launch gates: exclude would shortlist
    website transparency, not capability.

Everything the layer decides is deterministic: parsers never guess (a cell
that resists parsing is PARSE_FAIL with a reason, taking missing_policy),
and the three unknown states — NOT DISCLOSED, NO DATA, PARSE_FAIL — stay
visibly distinct in every output.

Criteria are data: a spec workbook (sheet "Criteria" + aux "FX", "Guards")
read at run time; add/remove/reweight without touching code.
`write_default_spec()` emits the launch spec.

Inputs are read-only. Three readers, one cell model:
  * read_matrix_cells    — a pipeline output workbook's Matrix sheet (the
                           production path; parses the writer's cell grammar
                           incl. conflict/unverified/demotion markers)
  * read_flat_gt_cells   — a flat GT workbook (GroundTruth sheet), for the
                           shortlist(criteria, GT) legs of the evaluation
  * read_eval_detail_cells — reconstructs AI cells from an eval report's
                           Detail sheet (Dell fallback when the raw pipeline
                           workbook is on another machine; verified markers
                           are unavailable → every cell carries
                           VERIFIED_UNKNOWN and unverified_policy cannot bite)
"""
from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from rapidfuzz import fuzz

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class Criterion:
    id: str
    question: str
    type: str                 # hard_gate | scored
    parser: str               # binary | count | money
    direction: str            # require_yes | require_no | max | min | prefer_yes | prefer_no
    threshold_lo: Optional[float]
    threshold_hi: Optional[float]
    weight: float
    missing_policy: str       # exclude | pass | flag
    unverified_policy: str    # use_flagged | verified_only
    max_age_years: Optional[int]
    notes: str = ""


@dataclass
class Spec:
    criteria: list[Criterion]
    fx: dict[str, float]              # currency code -> USD multiplier
    guards: dict[str, list[str]]      # guard class -> keyword list
    keywords: dict[str, list[str]] = field(default_factory=dict)  # criterion id -> match phrases (parser=keyword)
    path: str = ""


@dataclass
class CellItem:
    raw: str
    verified: bool = True


@dataclass
class Cell:
    items: list[CellItem] = field(default_factory=list)
    flags: set = field(default_factory=set)   # CONFLICT, UNVERIFIED_CELL, CAPPED, DEMOTED_RIVALS, VERIFIED_UNKNOWN
    state: str = "data"                       # data | ND | NO_DATA


@dataclass
class Parsed:
    ok: bool
    value: Optional[float] = None      # canonical point (count; USD for money; 1/0 for binary)
    lo: Optional[float] = None         # guaranteed lower bound (qualifiers/ranges)
    hi: Optional[float] = None         # guaranteed upper bound
    vintage: Optional[int] = None
    flags: set = field(default_factory=set)
    reason: str = ""                   # PARSE_FAIL reason when not ok


@dataclass
class GateVerdict:
    criterion_id: str
    outcome: str          # PASS | FAIL
    label: str            # PASS | FAIL | PASS (not disclosed) | PASS (no data) | PASS (unparseable: ...) | PASS (parent-attributed) | ...
    parsed: str           # human-readable parsed value ("" when none)
    raw: str              # the source cell text (traceability)
    flags: list = field(default_factory=list)


@dataclass
class ScoredValue:
    criterion_id: str
    score: Optional[float]   # None = no data (excluded from renorm)
    label: str
    raw: str


@dataclass
class EntityResult:
    entity: str
    gates: list[GateVerdict]
    excluded_by: str = ""            # first failing gate's criterion id ("" = survived)
    scores: list[ScoredValue] = field(default_factory=list)
    total: Optional[float] = None    # renormalised weighted score (survivors only)
    coverage: float = 0.0            # fraction of scored-criterion weight with data
    rank: Optional[int] = None


@dataclass
class ShortlistResult:
    entities: list[EntityResult]     # ranked survivors first (by rank), then excluded
    k: int
    spec: Spec


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
class CellMap(dict):
    """dict[(entity_norm, question_norm) -> Cell] that also carries the
    original display casing of each entity (join keys stay normalised so the
    GT and AI sides align; output shows the real names)."""
    def __init__(self):
        super().__init__()
        self.display: dict[str, str] = {}


def _norm(text: str) -> str:
    return " ".join(str(text).strip().lower().split())


def _is_nd(text: str) -> bool:
    t = _norm(text)
    return t == "none (not disclosed)" or "not disclosed" in t


_YEAR_RE = re.compile(r"(19|20)\d{2}")

# AI Summary cell grammar (io_excel._make_ai_summary_df): prose or verdicts
# carry [C####] citations, deterministic multi-answers are "; "-joined, and
# degraded cells end with a "[fallback: ...]" marker line.
_CITATION_RE = re.compile(r"\s*\[C\d[\d,\sC]*\]")
_MORE_RE = re.compile(r"\(more in provenance\)", re.IGNORECASE)


def _strip_citations(text: str) -> str:
    return _MORE_RE.sub("", _CITATION_RE.sub("", text)).strip()


# ---------------------------------------------------------------------------
# Spec I/O
# ---------------------------------------------------------------------------
_CRITERIA_COLS = ["id", "question", "type", "parser", "direction",
                  "threshold_lo", "threshold_hi", "weight", "missing_policy",
                  "unverified_policy", "max_age_years", "notes"]

# Launch spec — PLACEHOLDER criteria for the mechanism test (George,
# 2026-08-05). Two hard gates per the approved decisions; scored criteria
# exist to exercise ranking/renormalisation, their weights are toys.
_DEFAULT_CRITERIA = [
    # George 2026-08-05(3): client criteria REINTERPRETED. The first must is
    # not corporate independence but "do they OFFER manufacturing as a
    # service" (CMO/EMS/CDMO vs selling own products) — judged on the
    # Summary-description prose via the service-language keyword list
    # (sheet "Keywords"; deterministic, editable). Second must: MID-SIZED by
    # revenue (band, not just a giant cap). Independence gate superseded —
    # re-add a hard_gate row to restore it.
    ("cmo_services", "Summary description of company's services",
     "hard_gate", "keyword", "require_match", None, None, 0, "flag", "use_flagged", None,
     "MUST: description shows they offer manufacturing services (keyword list = data). No match -> kept, flagged (keywords can miss phrasings; LLM normaliser is the gated follow-up)."),
    ("rev_midsize", "What is the company's yearly revenue?",
     "hard_gate", "money", "range", 10e6, 1e9, 0, "flag", "use_flagged", 10,
     "PLACEHOLDER band: mid-sized = USD 10M-1B own revenue. Parent-attributed figures never decide."),
    ("medical", "Does the company have experience of medical device manufacturing?",
     "scored", "binary", "prefer_yes", None, None, 1, "flag", "use_flagged", None,
     "the scored criterion: client-intent + reliable (eval F1 0.84)"),
]

# Service-language markers for the cmo_services gate — DATA (sheet
# "Keywords"), extend freely. Short entries (<=4 chars) are matched as whole
# words, longer ones as substrings of the normalised description.
_DEFAULT_KEYWORDS = {
    "cmo_services": [
        "contract manufactur", "contract electronics", "cmo", "cdmo", "ems",
        "electronic manufacturing services", "manufacturing services",
        "manufacturing service", "toll manufactur", "turnkey manufactur",
        "build to print", "build-to-print", "assembly services",
        "manufacturing partner", "on behalf of", "oems", "odm",
    ],
}

# Static coarse FX (order-of-magnitude gates; reviewed manually — Q7).
_DEFAULT_FX = {"USD": 1.0, "GBP": 1.27, "EUR": 1.08, "CHF": 1.10, "SEK": 0.095}

_DEFAULT_GUARDS = {
    # count: value is a subset/team, not total headcount -> PARSE_FAIL
    "subset": ["engineers", "designers", "experts", "specialists", "staff of",
               "team of", "r&d", "scientists"],
    # count: not a current figure -> PARSE_FAIL
    "not_current": ["planned", "within next", "target", "projected", "by 20"],
    # money: forward-looking / partial-period figures parse but carry a flag
    "forecast": ["outlook", "forecast", "projected", "guidance", "target"],
    "partial_period": ["first half", "second half", "half of", "h1", "h2",
                       "q1", "q2", "q3", "q4", "quarter"],
}


def write_default_spec(path: str) -> None:
    import pandas as pd
    crit = pd.DataFrame(_DEFAULT_CRITERIA, columns=_CRITERIA_COLS)
    fx = pd.DataFrame(sorted(_DEFAULT_FX.items()), columns=["currency", "usd_rate"])
    guards = pd.DataFrame(
        [(cls, kw) for cls, kws in _DEFAULT_GUARDS.items() for kw in kws],
        columns=["guard_class", "keyword"])
    kws = pd.DataFrame(
        [(cid, kw) for cid, lst in _DEFAULT_KEYWORDS.items() for kw in lst],
        columns=["criterion_id", "phrase"])
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        crit.to_excel(w, sheet_name="Criteria", index=False)
        fx.to_excel(w, sheet_name="FX", index=False)
        guards.to_excel(w, sheet_name="Guards", index=False)
        kws.to_excel(w, sheet_name="Keywords", index=False)


def read_spec(path: str) -> Spec:
    import pandas as pd

    def _opt_float(v):
        if v is None or (isinstance(v, float) and v != v) or str(v).strip() == "":
            return None
        return float(v)

    df = pd.read_excel(path, sheet_name="Criteria")
    criteria = []
    for _, r in df.iterrows():
        criteria.append(Criterion(
            id=str(r["id"]).strip(),
            question=str(r["question"]).strip(),
            type=str(r["type"]).strip(),
            parser=str(r["parser"]).strip(),
            direction=str(r["direction"]).strip(),
            threshold_lo=_opt_float(r.get("threshold_lo")),
            threshold_hi=_opt_float(r.get("threshold_hi")),
            weight=float(r.get("weight") or 0),
            missing_policy=str(r["missing_policy"]).strip(),
            unverified_policy=str(r["unverified_policy"]).strip(),
            max_age_years=(int(v) if (v := _opt_float(r.get("max_age_years"))) is not None else None),
            notes=str(r.get("notes") or ""),
        ))
    fx = {str(r["currency"]).strip().upper(): float(r["usd_rate"])
          for _, r in pd.read_excel(path, sheet_name="FX").iterrows()}
    guards: dict[str, list[str]] = {}
    for _, r in pd.read_excel(path, sheet_name="Guards").iterrows():
        guards.setdefault(str(r["guard_class"]).strip(), []).append(
            str(r["keyword"]).strip().lower())
    keywords: dict[str, list[str]] = {}
    try:
        for _, r in pd.read_excel(path, sheet_name="Keywords").iterrows():
            keywords.setdefault(str(r["criterion_id"]).strip(), []).append(
                str(r["phrase"]).strip().lower())
    except ValueError:
        pass  # pre-keyword spec workbooks stay readable
    return Spec(criteria=criteria, fx=fx, guards=guards, keywords=keywords,
                path=path)


# ---------------------------------------------------------------------------
# Cell readers
# ---------------------------------------------------------------------------
_MARKER_LINES = {"(sources conflict)", "(unverified)", "-- unverified --"}


def read_matrix_cells(filepath: str,
                      sheet: Optional[str] = None) -> dict[tuple[str, str], Cell]:
    """Cells from any entity-rows x question-columns sheet of a pipeline
    output workbook. Default sheet: Matrix, falling back to AI Summary.
    Handles both cell grammars: the Matrix writer's bullets/markers
    (io_excel._make_matrix_df) and the AI Summary's [C####] citations,
    "; "-joined deterministic verdicts and [fallback: ...] lines."""
    import pandas as pd
    with pd.ExcelFile(filepath) as xls:
        if sheet is not None:
            sheet = next((s for s in xls.sheet_names
                          if _norm(s) == _norm(sheet)), None)
            if sheet is None:
                raise ValueError(f"{filepath!r}: sheet not found. "
                                 f"Available: {xls.sheet_names}")
        else:
            sheet = next((s for s in xls.sheet_names if s.lower() == "matrix"),
                         None) or next((s for s in xls.sheet_names
                                        if s.lower() == "ai summary"), None)
        if sheet is None:
            raise ValueError(f"{filepath!r} has no Matrix / AI Summary sheet: "
                             f"{xls.sheet_names}")
        df = pd.read_excel(xls, sheet_name=sheet)
    ent_col = next((c for c in df.columns if _norm(str(c)) == "entity"), df.columns[0])

    cells = CellMap()
    for _, r in df.iterrows():
        entity = str(r.get(ent_col) or "").strip()
        if not entity:
            continue
        cells.display.setdefault(_norm(entity), entity)
        for q in (c for c in df.columns if c != ent_col):
            text = r.get(q)
            text = "" if text is None or (isinstance(text, float) and text != text) else str(text).strip()
            if not text:
                continue
            cell = Cell()
            if _norm(text) == "no data found":
                cell.state = "NO_DATA"
            else:
                verified = not any(_norm(l) == "(unverified)" for l in text.split("\n"))
                if not verified:
                    cell.flags.add("UNVERIFIED_CELL")
                section_verified = verified
                for line in text.split("\n"):
                    s = line.strip()
                    sn = _norm(s)
                    if not s:
                        continue
                    if sn == "-- unverified --":
                        section_verified = False
                        continue
                    if sn == "(sources conflict)":
                        cell.flags.add("CONFLICT")
                        continue
                    if s.startswith("[+") and "candidate" in sn:
                        cell.flags.add("DEMOTED_RIVALS")
                        continue
                    if s.startswith("[+") or sn.startswith("[truncated"):
                        cell.flags.add("CAPPED")
                        continue
                    if sn.startswith("[fallback:"):
                        cell.flags.add("SUMMARY_FALLBACK")
                        continue
                    if sn in _MARKER_LINES:
                        continue
                    value = s[2:].strip() if s.startswith("- ") else s
                    # AI Summary grammar: strip [C####] citations, then split
                    # "; "-joined deterministic verdicts into items. Prose
                    # fragments produced by the split simply PARSE_FAIL at
                    # the criterion parser and flag through — never guessed.
                    value = _strip_citations(value)
                    for frag in (f.strip() for f in value.split("; ")):
                        if frag:
                            cell.items.append(CellItem(raw=frag, verified=section_verified))
                if cell.items and all(_is_nd(i.raw) for i in cell.items):
                    cell.state = "ND"
                elif not cell.items:
                    cell.state = "NO_DATA"
            cells[(_norm(entity), _norm(str(q)))] = cell
    return cells


def read_flat_gt_cells(filepath: str) -> dict[tuple[str, str], Cell]:
    """Cells from a flat GT workbook (GroundTruth sheet: entity|question|value)."""
    import pandas as pd
    df = pd.read_excel(filepath, sheet_name="GroundTruth")
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    cells = CellMap()
    for _, r in df.iterrows():
        entity, question = str(r.get("entity") or "").strip(), str(r.get("question") or "").strip()
        value = str(r.get("value") or "").strip()
        if not entity or not question or not value or value.lower() == "nan":
            continue
        cells.display.setdefault(_norm(entity), entity)
        key = (_norm(entity), _norm(question))
        cell = cells.setdefault(key, Cell())
        if _is_nd(value):
            if not cell.items:
                cell.state = "ND"
        else:
            cell.items.append(CellItem(raw=value, verified=True))
            cell.state = "data"
    return cells


def read_eval_detail_cells(filepath: str) -> dict[tuple[str, str], Cell]:
    """Reconstruct the AI side's cells from an eval report's Detail sheet —
    the Dell fallback when the raw pipeline workbook is on another machine.
    LIMITATION (recorded on every cell): verified/unverified and
    conflict/demotion markers are not in Detail, so every cell carries
    VERIFIED_UNKNOWN and unverified_policy cannot bite."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if "Detail" not in wb.sheetnames:
        raise ValueError(f"{filepath!r} has no Detail sheet")
    ws = wb["Detail"]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c) for c in next(rows)]
    i_ent, i_q = hdr.index("entity"), hdr.index("question")
    i_ai, i_v = hdr.index("ai_value"), hdr.index("verdict")

    cells = CellMap()
    seen: dict[tuple[str, str], set] = {}
    for r in rows:
        cells.display.setdefault(_norm(str(r[i_ent])), str(r[i_ent]).strip())
        key = (_norm(str(r[i_ent])), _norm(str(r[i_q])))
        cell = cells.setdefault(key, Cell(state="NO_DATA", flags={"VERIFIED_UNKNOWN"}))
        ai = str(r[i_ai] or "").strip()
        if not ai:
            continue
        if _is_nd(ai):
            if cell.state == "NO_DATA":
                cell.state = "ND"
            continue
        dedup = seen.setdefault(key, set())
        if ai not in dedup:
            dedup.add(ai)
            cell.items.append(CellItem(raw=ai, verified=True))
            cell.state = "data"
    wb.close()
    return cells


# ---------------------------------------------------------------------------
# Parsers — deterministic, never guess
# ---------------------------------------------------------------------------
_BINARY = {"yes": 1.0, "true": 1.0, "no": 0.0, "false": 0.0}


def parse_binary(text: str) -> Parsed:
    v = _BINARY.get(_norm(text).rstrip("."))
    if v is None:
        return Parsed(ok=False, reason=f"not a bare Yes/No: {text[:60]!r}")
    return Parsed(ok=True, value=v, lo=v, hi=v)


_QUAL_LOWER = re.compile(r"^(over|above|more than|at least|exceeding|\+?\s*)")
_QUAL_LOWER_WORDS = ("over", "above", "more than", "at least", "exceeding")
_QUAL_APPROX_WORDS = ("around", "about", "approx", "approximately", "circa", "c.", "~")


def _split_annotations(text: str) -> tuple[str, list[str]]:
    """Return (body without parentheticals, list of parenthetical contents)."""
    notes = re.findall(r"\(([^)]*)\)", text)
    body = re.sub(r"\([^)]*\)", " ", text)
    return " ".join(body.split()), [n.strip() for n in notes]


def _classify_annotation(note: str, guards: dict) -> tuple[str, Optional[int]]:
    """One parenthetical -> ('year'|'forecast'|'partial'|'not_current'|'company', vintage)."""
    n = _norm(note)
    year = None
    m = _YEAR_RE.search(n)
    if m:
        year = int(m.group(0))
    if any(kw in n for kw in guards.get("not_current", [])):
        return "not_current", year
    if any(kw in n for kw in guards.get("forecast", [])):
        return "forecast", year
    if any(kw in n for kw in guards.get("partial_period", [])):
        return "partial", year
    stripped = re.sub(r"(fy|fiscal year|as of|year)", "", n)
    if year is not None and not re.search(r"[a-z]", re.sub(_YEAR_RE, "", stripped)):
        return "year", year
    return "company", year


def _qualifiers(body: str) -> tuple[str, str]:
    """Return (body without qualifier, kind: ''|'lower_bound'|'approx')."""
    b = _norm(body)
    kind = ""
    for w in _QUAL_LOWER_WORDS:
        if b.startswith(w + " "):
            b, kind = b[len(w):].strip(), "lower_bound"
            break
    for w in _QUAL_APPROX_WORDS:
        if b.startswith(w + " ") or b.startswith(w):
            if kind == "":
                b, kind = b[len(w):].strip(), "approx"
            break
    if b.endswith("+"):
        b, kind = b[:-1].strip(), "lower_bound"
    elif re.search(r"\d\+", b):
        # digit-attached plus mid-value: "2+ billion", "1.6 billion +"
        b, kind = re.sub(r"(?<=\d)\+", "", b), "lower_bound"
    return b, kind


def parse_count(text: str, guards: dict, ref_year: int,
                max_age: Optional[int]) -> Parsed:
    t = _norm(text)
    for cls in ("subset", "not_current"):
        hit = next((kw for kw in guards.get(cls, []) if kw in t), None)
        if hit:
            return Parsed(ok=False, reason=f"{cls} guard ({hit!r}): {text[:60]!r}")
    body, notes = _split_annotations(text)
    flags: set = set()
    vintage = None
    for n in notes:
        kind, year = _classify_annotation(n, guards)
        if kind == "not_current":
            return Parsed(ok=False, reason=f"not_current annotation: ({n})")
        if kind == "year":
            vintage = year
        elif kind in ("forecast", "partial"):
            flags.add(kind.upper())
            vintage = vintage or year
        else:
            flags.add("ANNOTATION")
    body, qual = _qualifiers(body)
    if qual:
        flags.add("QUALIFIED")
    m = re.fullmatch(r"(\d[\d,\s]*)\s*(?:-|–|to)\s*(\d[\d,\s]*)", body)
    if m:
        lo = float(m.group(1).replace(",", "").replace(" ", ""))
        hi = float(m.group(2).replace(",", "").replace(" ", ""))
        flags.add("RANGE")
        return Parsed(ok=True, value=(lo + hi) / 2, lo=lo, hi=hi,
                      vintage=vintage, flags=flags)
    b = body.replace(",", "").replace(" ", "")
    if not re.fullmatch(r"\d+(\.\d+)?", b or ""):
        return Parsed(ok=False, reason=f"not a headcount number: {text[:60]!r}")
    v = float(b)
    p = Parsed(ok=True, value=v, lo=v, hi=None if qual == "lower_bound" else v,
               vintage=vintage, flags=flags)
    if max_age is not None and vintage is not None and ref_year - vintage > max_age:
        p.flags.add("STALE")
    return p


_CUR_TOKENS = {"$": "USD", "usd": "USD", "us$": "USD", "£": "GBP", "gbp": "GBP",
               "€": "EUR", "eur": "EUR", "chf": "CHF", "sek": "SEK"}
_SCALES = {"billion": 1e9, "bn": 1e9, "b": 1e9,
           "million": 1e6, "m": 1e6, "mio": 1e6}
_MONEY_RE = re.compile(
    r"(?P<cur1>\$|£|€|usd|gbp|eur|chf|sek|us\$)?\s*"
    r"(?P<amt>\d[\d,]*(?:\.\d+)?)\s*"
    r"(?P<scale>billion|million|bn|mio|b|m|msek)?\s*"
    r"(?P<cur2>usd|gbp|eur|chf|sek)?", re.IGNORECASE)


def _one_money(fragment: str, fx: dict) -> Optional[tuple[float, set]]:
    """Parse one amount like '$29.8B' / '45M' / '1,426.50 MSEK' -> (usd, flags)."""
    m = _MONEY_RE.search(fragment)
    if not m or not m.group("amt"):
        return None
    amt = float(m.group("amt").replace(",", ""))
    scale_tok = (m.group("scale") or "").lower()
    flags: set = set()
    if scale_tok == "msek":
        cur, scale = "SEK", 1e6
    else:
        scale = _SCALES.get(scale_tok, 1.0)
        cur_tok = (m.group("cur1") or m.group("cur2") or "").lower()
        cur = _CUR_TOKENS.get(cur_tok, "")
        if not cur:
            cur = "USD"
            flags.add("NO_CURRENCY")
    rate = fx.get(cur)
    if rate is None:
        return None
    return amt * scale * rate, flags


def parse_money(text: str, guards: dict, fx: dict, ref_year: int,
                max_age: Optional[int]) -> Parsed:
    # Human-typed thousands with a space after the comma ("4, 153 million
    # SEK" = 4,153) — rejoin before the amount regex, which otherwise stops
    # at the comma and reads 4 (found scoring the final59 GT, Rosti).
    text = re.sub(r"(?<=\d),\s+(?=\d{3})", ",", text)
    body, notes = _split_annotations(text)
    flags: set = set()
    vintage = None
    for n in notes:
        kind, year = _classify_annotation(n, guards)
        if kind == "company":
            # Q1 (approved): a non-year parenthetical naming another company
            # = parent-attributed revenue. Parses, flagged, never trips the
            # giant gate.
            flags.add("PARENT_ATTRIBUTED")
        elif kind == "year":
            vintage = year
        else:
            flags.add(kind.upper())
            vintage = vintage or year
    t = _norm(body)
    for cls in ("forecast", "partial_period"):
        if any(kw in t for kw in guards.get(cls, [])):
            flags.add("FORECAST" if cls == "forecast" else "PARTIAL")
    body_q, qual = _qualifiers(body)
    if qual:
        flags.add("QUALIFIED")
    # Range "X to Y" / "X - Y"
    parts = re.split(r"\s+to\s+|\s-\s", body_q)
    if len(parts) == 2:
        a, b = _one_money(parts[0], fx), _one_money(parts[1], fx)
        if a and b:
            flags |= a[1] | b[1] | {"RANGE"}
            lo, hi = sorted((a[0], b[0]))
            return Parsed(ok=True, value=(lo + hi) / 2, lo=lo, hi=hi,
                          vintage=vintage, flags=flags)
    one = _one_money(body_q, fx)
    if one is None:
        return Parsed(ok=False, reason=f"no parseable amount: {text[:60]!r}",
                      flags=flags)
    usd, f2 = one
    flags |= f2
    p = Parsed(ok=True, value=usd, lo=usd,
               hi=None if qual == "lower_bound" else usd,
               vintage=vintage, flags=flags)
    if max_age is not None and vintage is not None and ref_year - vintage > max_age:
        p.flags.add("STALE")
    return p


def keyword_match(text: str, phrases: list[str]) -> Optional[str]:
    """First matching service-language phrase, or None. Short phrases
    (<= 4 chars: 'cmo', 'ems', 'odm') match as whole words only — 'cmo'
    must not fire inside 'cmos'; longer phrases match as substrings of the
    normalised text (hyphens folded to spaces)."""
    t = _norm(text).replace("-", " ")
    for kw in phrases:
        k = kw.replace("-", " ")
        if len(k) <= 4:
            if re.search("(?<![a-z0-9])" + re.escape(k) + "(?![a-z0-9])", t):
                return kw
        elif k in t:
            return kw
    return None


def parse_value(crit: Criterion, text: str, spec: Spec, ref_year: int) -> Parsed:
    if crit.parser == "binary":
        return parse_binary(text)
    if crit.parser == "count":
        return parse_count(text, spec.guards, ref_year, crit.max_age_years)
    if crit.parser == "money":
        return parse_money(text, spec.guards, spec.fx, ref_year, crit.max_age_years)
    return Parsed(ok=False, reason=f"unknown parser {crit.parser!r}")


# ---------------------------------------------------------------------------
# Gate + score engine
# ---------------------------------------------------------------------------
def _pick_item(cell: Cell, crit: Criterion) -> tuple[Optional[CellItem], set]:
    """The item a criterion judges: the lead (first) item, verified-first
    when the policy demands it. Extra items -> MULTI_VALUE flag."""
    flags: set = set()
    items = cell.items
    if crit.unverified_policy == "verified_only" and "VERIFIED_UNKNOWN" not in cell.flags:
        verified = [i for i in items if i.verified]
        if items and not verified:
            return None, {"UNVERIFIED_ONLY"}
        items = verified
    if not items:
        return None, flags
    if len(items) > 1:
        flags.add("MULTI_VALUE")
    return items[0], flags


def _missing_verdict(crit: Criterion, label: str, raw: str, flags: set) -> GateVerdict:
    if crit.missing_policy == "exclude":
        return GateVerdict(crit.id, "FAIL", f"EXCLUDED ({label})", "", raw, sorted(flags))
    return GateVerdict(crit.id, "PASS", f"PASS ({label})", "", raw, sorted(flags))


def eval_gate(crit: Criterion, cell: Optional[Cell], spec: Spec,
              ref_year: int) -> GateVerdict:
    raw = "\n".join(i.raw for i in cell.items) if cell else ""
    cell_flags = set(cell.flags) if cell else set()
    if cell is None or cell.state == "NO_DATA":
        return _missing_verdict(crit, "no data", raw, cell_flags)
    if cell.state == "ND":
        return _missing_verdict(crit, "not disclosed", "Not disclosed", cell_flags)

    # keyword criteria judge the WHOLE cell prose (a description is one
    # answer, not competing candidates), not the lead item.
    if crit.parser == "keyword":
        text = " ".join(i.raw for i in cell.items)
        hit = keyword_match(text, spec.keywords.get(crit.id, []))
        if hit is not None:
            return GateVerdict(crit.id, "PASS", "PASS",
                               f"matches {hit!r}", raw, sorted(cell_flags))
        # No phrase found != a refusal: descriptions phrase things freely.
        # Treated like missing data, visibly labelled.
        return _missing_verdict(crit, "no service-language match", raw,
                                cell_flags)

    item, pick_flags = _pick_item(cell, crit)
    flags = cell_flags | pick_flags
    if item is None:
        # Q3 (approved): only an explicit VERIFIED value can exclude; an
        # unverified-only cell flags through like missing data.
        return _missing_verdict(crit, "unverified only", raw, flags)

    p = parse_value(crit, item.raw, spec, ref_year)
    flags |= p.flags
    if not p.ok:
        return _missing_verdict(crit, f"unparseable: {p.reason}", raw, flags)

    if crit.parser == "money" and "PARENT_ATTRIBUTED" in p.flags:
        # Q1 (approved): giant gate measures the entity's OWN size only.
        return GateVerdict(crit.id, "PASS", "PASS (parent-attributed)",
                           f"{p.value:,.0f} USD (parent's)", raw, sorted(flags))

    parsed_str = (f"{p.value:,.0f}" if crit.parser == "count"
                  else f"{p.value:,.0f} USD" if crit.parser == "money"
                  else ("Yes" if p.value else "No"))

    if crit.direction in ("require_yes", "require_no"):
        want = 1.0 if crit.direction == "require_yes" else 0.0
        outcome = "PASS" if p.value == want else "FAIL"
        return GateVerdict(crit.id, outcome, outcome, parsed_str, raw, sorted(flags))

    if crit.direction == "max":
        t = crit.threshold_hi
        if p.lo is not None and p.lo > t:
            return GateVerdict(crit.id, "FAIL", "FAIL", parsed_str, raw, sorted(flags))
        if p.hi is not None and p.hi <= t:
            label = "PASS" if not (flags & {"QUALIFIED", "RANGE", "STALE"}) else "PASS (flagged)"
            return GateVerdict(crit.id, "PASS", label, parsed_str, raw, sorted(flags))
        # Qualifier leaves the bound open on the deciding side ("over 1,500"
        # vs max 10,000): the stated number passes, the guarantee doesn't.
        flags.add("BOUND_OPEN")
        return GateVerdict(crit.id, "PASS", "PASS (bound open)", parsed_str, raw, sorted(flags))

    if crit.direction == "min":
        t = crit.threshold_lo
        if p.hi is not None and p.hi < t:
            return GateVerdict(crit.id, "FAIL", "FAIL", parsed_str, raw, sorted(flags))
        if p.lo is not None and p.lo >= t:
            return GateVerdict(crit.id, "PASS", "PASS", parsed_str, raw, sorted(flags))
        flags.add("BOUND_OPEN")
        return GateVerdict(crit.id, "PASS", "PASS (bound open)", parsed_str, raw, sorted(flags))

    if crit.direction == "range":
        lo_t, hi_t = crit.threshold_lo, crit.threshold_hi
        if p.hi is not None and p.hi < lo_t:
            return GateVerdict(crit.id, "FAIL", "FAIL (below band)",
                               parsed_str, raw, sorted(flags))
        if p.lo is not None and p.lo > hi_t:
            return GateVerdict(crit.id, "FAIL", "FAIL (above band)",
                               parsed_str, raw, sorted(flags))
        if p.lo is not None and p.hi is not None and lo_t <= p.lo and p.hi <= hi_t:
            label = "PASS" if not (flags & {"QUALIFIED", "RANGE", "STALE"}) else "PASS (flagged)"
            return GateVerdict(crit.id, "PASS", label, parsed_str, raw, sorted(flags))
        flags.add("BOUND_OPEN")
        return GateVerdict(crit.id, "PASS", "PASS (bound open)", parsed_str,
                           raw, sorted(flags))

    return _missing_verdict(crit, f"unknown direction {crit.direction!r}", raw, flags)


def eval_scored(crit: Criterion, cell: Optional[Cell], spec: Spec,
                ref_year: int) -> ScoredValue:
    raw = "\n".join(i.raw for i in cell.items) if cell else ""
    if cell is None or cell.state in ("NO_DATA", "ND"):
        label = "no data" if (cell is None or cell.state == "NO_DATA") else "not disclosed"
        return ScoredValue(crit.id, None, label, raw)
    item, _ = _pick_item(cell, crit)
    if item is None:
        return ScoredValue(crit.id, None, "unverified only", raw)
    p = parse_value(crit, item.raw, spec, ref_year)
    if not p.ok:
        return ScoredValue(crit.id, None, f"unparseable: {p.reason}", raw)
    if crit.direction == "prefer_yes":
        return ScoredValue(crit.id, p.value, "Yes" if p.value else "No", raw)
    if crit.direction == "prefer_no":
        return ScoredValue(crit.id, 1.0 - p.value, "No" if not p.value else "Yes", raw)
    # numeric prefer_low / prefer_high: piecewise linear between thresholds
    lo, hi = crit.threshold_lo or 0.0, crit.threshold_hi or 1.0
    x = max(0.0, min(1.0, (p.value - lo) / (hi - lo))) if hi != lo else 0.0
    s = 1.0 - x if crit.direction == "prefer_low" else x
    return ScoredValue(crit.id, s, f"{p.value:,.0f}", raw)


def _match_questions(spec: Spec, cells: dict) -> dict[str, str]:
    """criterion id -> question_norm actually present in the cells (exact
    then fuzzy >= 0.70, same rule as the evaluator)."""
    qnorms = {q for _, q in cells.keys()}
    out = {}
    for c in spec.criteria:
        want = _norm(c.question)
        if want in qnorms:
            out[c.id] = want
            continue
        best, best_s = None, 0.0
        for qn in qnorms:
            s = fuzz.token_sort_ratio(want, qn) / 100.0
            if s > best_s:
                best, best_s = qn, s
        out[c.id] = best if best_s >= 0.70 else want
    return out


def run_shortlist(cells: dict[tuple[str, str], Cell], spec: Spec,
                  k: int = 5, ref_year: int = 2026) -> ShortlistResult:
    entities = sorted({e for e, _ in cells.keys()})
    qmap = _match_questions(spec, cells)
    gates = [c for c in spec.criteria if c.type == "hard_gate"]
    scored = [c for c in spec.criteria if c.type == "scored"]

    display = getattr(cells, "display", {})
    results: list[EntityResult] = []
    for ent in entities:
        er = EntityResult(entity=display.get(ent, ent), gates=[])
        for c in gates:
            v = eval_gate(c, cells.get((ent, qmap[c.id])), spec, ref_year)
            er.gates.append(v)
            if v.outcome == "FAIL" and not er.excluded_by:
                er.excluded_by = c.id
        if not er.excluded_by:
            er.scores = [eval_scored(c, cells.get((ent, qmap[c.id])), spec, ref_year)
                         for c in scored]
            wsum = sum(c.weight for c in scored)
            have = [(c, s) for c, s in zip(scored, er.scores) if s.score is not None]
            w_have = sum(c.weight for c, _ in have)
            er.coverage = round(w_have / wsum, 4) if wsum else 0.0
            er.total = (round(sum(c.weight * s.score for c, s in have) / w_have, 4)
                        if w_have else 0.0)
        results.append(er)

    survivors = [r for r in results if not r.excluded_by]
    survivors.sort(key=lambda r: (-(r.total or 0), -r.coverage, r.entity.casefold()))
    for i, r in enumerate(survivors, start=1):
        r.rank = i
    ranked = survivors + [r for r in results if r.excluded_by]
    return ShortlistResult(entities=ranked, k=k, spec=spec)


# ---------------------------------------------------------------------------
# Comparison helpers (the two-run experiment)
# ---------------------------------------------------------------------------
def top_k(result: ShortlistResult, k: Optional[int] = None) -> list[str]:
    k = k or result.k
    return [r.entity for r in result.entities if r.rank is not None][:k]


def overlap_at_k(a: ShortlistResult, b: ShortlistResult,
                 k: Optional[int] = None) -> float:
    """|top-k(a) ∩ top-k(b)| / k (denominator shrinks if a run has fewer
    than k survivors, so a short list is not penalised for its length)."""
    k = k or a.k
    sa, sb = set(top_k(a, k)), set(top_k(b, k))
    denom = min(k, max(len(sa), len(sb))) or 1
    return round(len(sa & sb) / denom, 4)


def diff_gates(gt: ShortlistResult, ai: ShortlistResult) -> list[dict]:
    """Per (entity, gate): how the two runs' verdicts diverge, classified so
    every disagreement traces to the specific cell that diverged."""
    def _by_ent(res):
        return {r.entity: r for r in res.entities}
    g, a = _by_ent(gt), _by_ent(ai)
    rows = []
    for ent in sorted(set(g) | set(a)):
        gv = {v.criterion_id: v for v in g[ent].gates} if ent in g else {}
        av = {v.criterion_id: v for v in a[ent].gates} if ent in a else {}
        for cid in sorted(set(gv) | set(av)):
            vg, va = gv.get(cid), av.get(cid)
            if vg is None or va is None:
                cls = "ENTITY_MISSING_ONE_SIDE"
            elif vg.outcome != va.outcome:
                if "unparseable" in vg.label or "unparseable" in va.label:
                    cls = "PARSE_DIVERGENCE"
                elif ("no data" in vg.label or "not disclosed" in vg.label
                      or "no data" in va.label or "not disclosed" in va.label):
                    cls = "MISSING_DIVERGENCE"
                else:
                    cls = "VALUE_DISAGREES"
            elif vg.label != va.label or vg.parsed != va.parsed:
                if "unparseable" in vg.label or "unparseable" in va.label:
                    cls = "PARSE_DIVERGENCE (same outcome)"
                elif vg.parsed != va.parsed and vg.parsed and va.parsed:
                    cls = "VALUE_DIFFERS (same outcome)"
                else:
                    cls = "MISSING_DIVERGENCE (same outcome)"
            else:
                cls = "SAME"
            rows.append({
                "entity": ent, "gate": cid, "class": cls,
                "gt_verdict": vg.label if vg else "", "gt_value": vg.parsed if vg else "",
                "gt_raw": vg.raw if vg else "",
                "ai_verdict": va.label if va else "", "ai_value": va.parsed if va else "",
                "ai_raw": va.raw if va else "",
            })
    return rows


def gate_decisiveness(result: ShortlistResult) -> dict[str, dict]:
    """Per gate: how often it actually judged DATA vs flagged through missing
    — the trivial-agreement guard. A gate that mostly flags through produces
    agreement between runs that says nothing about pipeline quality."""
    out: dict[str, dict] = {}
    for r in result.entities:
        for v in r.gates:
            d = out.setdefault(v.criterion_id,
                               {"decided_on_data": 0, "flagged_through": 0, "failed": 0})
            if v.outcome == "FAIL":
                d["failed"] += 1
                d["decided_on_data"] += 1
            elif v.label.startswith("PASS (") and any(
                    s in v.label for s in ("no data", "not disclosed", "unparseable",
                                           "unverified only",
                                           "no service-language match")):
                d["flagged_through"] += 1
            else:
                d["decided_on_data"] += 1
    return out
