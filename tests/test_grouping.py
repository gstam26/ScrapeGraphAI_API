"""
Tests for the deterministic claim-grouping layer (src/group.py) and its
integration points — fully offline: embed_batch is monkeypatched with
hand-built vectors (3 orthogonal direction families + small deterministic
noise) so the expected clusters are known exactly and no Ollama is needed.
"""
import openpyxl
import pytest

import pipeline as pipeline_mod
import src.group as group_mod
from config import GROUP_MIN_ITEMS, MATRIX_MAX_DISPLAY_ITEMS
from models import (
    ColumnSpec,
    ExtractedCell,
    ExtractedRow,
    PipelineInput,
    PipelineResult,
    SourceQuote,
    UrlSpec,
)
from src.group import ALL_ITEMS_THEME, group_rows
from src.io_excel import write_output_excel


# ── Fake embeddings: 3 orthogonal families, small deterministic noise ────────

def _fake_embed_batch(texts: list[str]) -> list[list[float]]:
    vecs = []
    for t in texts:
        low = t.lower()
        if "regulatory" in low:
            base = [1.0, 0.0, 0.0, 0.0]
        elif "launch" in low:
            base = [0.0, 1.0, 0.0, 0.0]
        elif "partnership" in low:
            base = [0.0, 0.0, 1.0, 0.0]
        else:
            base = [0.0, 0.0, 0.0, 1.0]
        # Deterministic small noise: same input -> same vector, within-family
        # cosine stays ~0.99, cross-family ~0.0 (well clear of GROUP_SIMILARITY).
        noise = (len(t) % 7) * 0.02
        vecs.append([base[0], base[1], base[2], base[3] + noise if base[3] == 0.0 else base[3]])
    return vecs


_REG = [
    "Zeta regulatory clearance for assay Z",
    "Alpha regulatory clearance in Europe",
    "Mid regulatory approval for kit A",
]
_LAUNCH = [
    "Launch of product line A",
    "New product launch in Japan",
    "Big launch of analyzer C",
]
_PARTNER = [
    "Partnership with Acme Corp",
    "Global partnership deal signed",
]

# Mixed display order, deliberately different from sorted() order.
_DISPLAY = [_LAUNCH[1], _REG[0], _LAUNCH[0], _PARTNER[0], _REG[1], _LAUNCH[2], _REG[2], _PARTNER[1]]


def _big_cell() -> ExtractedCell:
    evidence = (
        [SourceQuote(value=v, source_url="https://a.example.com/news") for v in _REG[:2]]
        + [SourceQuote(value=_REG[2], source_url="https://b.example.com/press")]
        + [SourceQuote(value=v, source_url="https://a.example.com/news") for v in _LAUNCH]
        + [SourceQuote(value=_PARTNER[0], source_url="https://c.example.com/pr")]
        # _PARTNER[1] intentionally has no matching evidence -> omitted from count
    )
    return ExtractedCell(
        entity="HORIBA",
        source_url="https://a.example.com/news",
        column="Recent news",
        value=list(_DISPLAY),
        evidence=evidence,
    )


def _rows_with_big_cell() -> list[ExtractedRow]:
    cell = _big_cell()
    return [ExtractedRow(entity="HORIBA", cells=[cell], all_cells=[cell])]


# ── Clustering behaviour ──────────────────────────────────────────────────────

def test_cluster_memberships_match_vector_families(monkeypatch):
    monkeypatch.setattr(group_mod, "embed_batch", _fake_embed_batch)
    groups = group_rows(_rows_with_big_cell())

    assert len(groups) == 3
    member_sets = [frozenset(g["values"]) for g in groups]
    assert frozenset(_REG) in member_sets
    assert frozenset(_LAUNCH) in member_sets
    assert frozenset(_PARTNER) in member_sets

    # Sorted size desc, then theme label asc.
    keys = [(-g["n_items"], g["theme"]) for g in groups]
    assert keys == sorted(keys)
    assert [g["n_items"] for g in groups] == [3, 3, 2]

    for g in groups:
        assert g["entity"] == "HORIBA"
        assert g["question"] == "Recent news"
        # Members are reported in original cell display order.
        display_pos = [_DISPLAY.index(v) for v in g["values"]]
        assert display_pos == sorted(display_pos)
    print("OK test_cluster_memberships_match_vector_families passed")


def test_medoid_theme_is_a_member_string(monkeypatch):
    monkeypatch.setattr(group_mod, "embed_batch", _fake_embed_batch)
    groups = group_rows(_rows_with_big_cell())
    for g in groups:
        assert g["theme"] in g["values"], "theme must be a real member claim, never synthesized"
    # 2-member cluster: label is the first member in sorted order.
    partner = next(g for g in groups if set(g["values"]) == set(_PARTNER))
    assert partner["theme"] == min(_PARTNER)
    print("OK test_medoid_theme_is_a_member_string passed")


def test_distinct_source_counts(monkeypatch):
    monkeypatch.setattr(group_mod, "embed_batch", _fake_embed_batch)
    groups = group_rows(_rows_with_big_cell())
    by_members = {frozenset(g["values"]): g for g in groups}
    assert by_members[frozenset(_REG)]["sources"] == 2      # a.example + b.example
    assert by_members[frozenset(_LAUNCH)]["sources"] == 1   # a.example only
    # One partner value matched (c.example); the unmatched one is omitted.
    assert by_members[frozenset(_PARTNER)]["sources"] == 1
    print("OK test_distinct_source_counts passed")


def test_determinism_two_calls_identical(monkeypatch):
    monkeypatch.setattr(group_mod, "embed_batch", _fake_embed_batch)
    first = group_rows(_rows_with_big_cell())
    second = group_rows(_rows_with_big_cell())
    assert first == second, "same rows + same embeddings must give identical output"
    print("OK test_determinism_two_calls_identical passed")


def test_small_cell_single_all_items_group_no_embedding(monkeypatch):
    def _must_not_embed(texts):
        raise AssertionError(f"embed_batch must not be called for small cells: {texts}")

    monkeypatch.setattr(group_mod, "embed_batch", _must_not_embed)
    values = ["alpha", "beta", "gamma"]
    assert len(values) < GROUP_MIN_ITEMS
    cell = ExtractedCell(
        entity="Small Co",
        source_url="https://s.example.com",
        column="Company type",
        value=list(values),
        evidence=[SourceQuote(value=v, source_url="https://s.example.com") for v in values],
    )
    groups = group_rows([ExtractedRow(entity="Small Co", cells=[cell], all_cells=[cell])])

    assert len(groups) == 1
    g = groups[0]
    assert g["theme"] == ALL_ITEMS_THEME
    assert g["values"] == values  # original display order preserved
    assert g["n_items"] == 3
    assert g["sources"] == 1
    print("OK test_small_cell_single_all_items_group_no_embedding passed")


def test_scalar_value_normalised_to_one_item_list(monkeypatch):
    monkeypatch.setattr(group_mod, "embed_batch", _fake_embed_batch)
    cell = ExtractedCell(
        entity="Scalar Co",
        source_url="https://x.example.com",
        column="Company type",
        value="Corporate",
    )
    groups = group_rows([ExtractedRow(entity="Scalar Co", cells=[cell], all_cells=[cell])])
    assert len(groups) == 1
    assert groups[0]["values"] == ["Corporate"] and groups[0]["theme"] == ALL_ITEMS_THEME
    print("OK test_scalar_value_normalised_to_one_item_list passed")


def test_empty_and_null_sentinel_cells_skipped(monkeypatch):
    monkeypatch.setattr(group_mod, "embed_batch", _fake_embed_batch)
    cells = [
        ExtractedCell(entity="E", source_url="u", column="Q1", value=None),
        ExtractedCell(entity="E", source_url="u", column="Q2", value=[]),
        ExtractedCell(entity="E", source_url="u", column="Q3", value=""),
        ExtractedCell(entity="E", source_url="u", column="Q4",
                      value=["None (not disclosed on the site)"]),
        ExtractedCell(entity="E", source_url="u", column="Q5", value=[None, "", []]),
    ]
    groups = group_rows([ExtractedRow(entity="E", cells=cells, all_cells=cells)])
    assert groups == []
    print("OK test_empty_and_null_sentinel_cells_skipped passed")


def test_embedding_failure_raises_clean_runtime_error(monkeypatch):
    def _broken(texts):
        raise OSError("connection refused")

    monkeypatch.setattr(group_mod, "embed_batch", _broken)
    with pytest.raises(RuntimeError, match="claim-grouping embedding failed"):
        group_rows(_rows_with_big_cell())
    print("OK test_embedding_failure_raises_clean_runtime_error passed")


# ── Pipeline degradation: grouping can never fail the run ────────────────────

def test_run_pipeline_survives_group_rows_raising(monkeypatch, capsys):
    """group_rows raising (e.g. Ollama unreachable off-network) must not fail
    the run — the sheet is simply absent. Uses the trivial-_process_url_spec
    pattern from tests/test_pipeline_resilience.py."""
    def fake_process(spec, request, cfg, all_entities):
        return {
            "entities": spec.entities or all_entities,
            "diag": {"acquire_log": [], "crawl_candidates": [], "filter_log": [],
                     "extract_log": [], "verify_log": []},
            "pages": [], "cells": [], "extract_time_ms": 0, "error": None,
        }

    def broken_group_rows(rows):
        raise RuntimeError("claim-grouping embedding failed: Ollama unreachable")

    monkeypatch.setattr(pipeline_mod, "_process_url_spec", fake_process)
    monkeypatch.setattr(pipeline_mod, "group_rows", broken_group_rows)

    request = PipelineInput(
        entities=["Good Co"],
        urls=[UrlSpec(url="https://good.example.com", depth=0, entities=["Good Co"])],
        columns=[ColumnSpec(name="Q1")],
    )
    result, diag = pipeline_mod.run_pipeline(request)

    assert {row.entity for row in result.rows} == {"Good Co"}
    assert "claim_groups" not in diag
    assert "! Grouping skipped:" in capsys.readouterr().out
    print("OK test_run_pipeline_survives_group_rows_raising passed")


def test_run_pipeline_stores_claim_groups_on_success(monkeypatch):
    def fake_process(spec, request, cfg, all_entities):
        return {
            "entities": spec.entities or all_entities,
            "diag": {"acquire_log": [], "crawl_candidates": [], "filter_log": [],
                     "extract_log": [], "verify_log": []},
            "pages": [], "cells": [], "extract_time_ms": 0, "error": None,
        }

    fake_groups = [{"entity": "Good Co", "question": "Q1", "theme": "t",
                    "n_items": 1, "values": ["t"], "sources": 1}]
    monkeypatch.setattr(pipeline_mod, "_process_url_spec", fake_process)
    monkeypatch.setattr(pipeline_mod, "group_rows", lambda rows: fake_groups)

    request = PipelineInput(
        entities=["Good Co"],
        urls=[UrlSpec(url="https://good.example.com", depth=0, entities=["Good Co"])],
        columns=[ColumnSpec(name="Q1")],
    )
    _, diag = pipeline_mod.run_pipeline(request)
    assert diag["claim_groups"] == fake_groups
    print("OK test_run_pipeline_stores_claim_groups_on_success passed")


# ── Sheet writer ──────────────────────────────────────────────────────────────

def _minimal_result() -> PipelineResult:
    cell = ExtractedCell(
        entity="HORIBA",
        source_url="https://a.example.com/news",
        column="Recent news",
        value=["item"],
        evidence=[SourceQuote(value="item", quote="item", verified=True)],
        verified=True,
    )
    return PipelineResult(rows=[ExtractedRow(entity="HORIBA", cells=[cell], all_cells=[cell])])


def test_grouped_themes_sheet_written_with_cap_marker(tmp_path):
    n_values = MATRIX_MAX_DISPLAY_ITEMS + 10
    big_values = [f"news item {i} about product {i}" for i in range(n_values)]
    diag = {
        "claim_groups": [
            {"entity": "HORIBA", "question": "Recent news", "theme": big_values[0],
             "n_items": n_values, "values": big_values, "sources": 3},
            {"entity": "HORIBA", "question": "Recent news", "theme": "small theme",
             "n_items": 2, "values": ["small theme", "other"], "sources": 1},
        ],
    }
    out = str(tmp_path / "grouped.xlsx")
    write_output_excel(_minimal_result(), [ColumnSpec(name="Recent news")], out, diag=diag)

    wb = openpyxl.load_workbook(out)
    assert "Grouped Themes" in wb.sheetnames
    # Deliverable-facing: sits right after Provenance.
    assert wb.sheetnames.index("Grouped Themes") == wb.sheetnames.index("Provenance") + 1

    ws = wb["Grouped Themes"]
    header = [c.value for c in ws[1]]
    assert header == ["Entity", "Question", "Theme", "Items", "Values", "Distinct Sources"]

    values_cell = ws.cell(row=2, column=5).value
    bullets = [line for line in values_cell.split("\n") if line.startswith("- ")]
    assert len(bullets) == MATRIX_MAX_DISPLAY_ITEMS
    assert f"[+10 more items — see Provenance]" in values_cell
    assert ws.cell(row=2, column=4).value == n_values
    assert ws.cell(row=3, column=3).value == "small theme"
    wb.close()
    print("OK test_grouped_themes_sheet_written_with_cap_marker passed")


def test_no_sheet_when_no_claim_groups(tmp_path):
    out = str(tmp_path / "no_groups.xlsx")
    write_output_excel(_minimal_result(), [ColumnSpec(name="Recent news")], out,
                       diag={"claim_groups": []})
    wb = openpyxl.load_workbook(out)
    assert "Grouped Themes" not in wb.sheetnames
    wb.close()
    print("OK test_no_sheet_when_no_claim_groups passed")


# ── Mean-centering (anisotropy correction) ────────────────────────────────────

def test_centering_separates_families_sharing_dominant_component():
    """The real-data failure geometry (2026-07-03 calibration): every claim in
    a cell shares a large company/domain component, so RAW cosines all sit in
    a narrow high band and one giant cluster forms at any usable threshold.
    center_vector_map must remove the shared component so the same threshold
    separates the families."""
    from src.group import center_vector_map, cluster_values

    # Two families, both dominated by the same shared direction [10,0,...]:
    # raw cosine between ANY pair is ~0.995 — indistinguishable.
    fam_a = {f"a{i}": [10.0, 1.0 + 0.01 * i, 0.0, 0.0] for i in range(4)}
    fam_b = {f"b{i}": [10.0, -1.0 - 0.01 * i, 0.0, 0.0] for i in range(4)}
    vectors = {**fam_a, **fam_b}
    values = sorted(vectors)

    raw_clusters = cluster_values(values, vectors, threshold=0.9)
    assert len(raw_clusters) == 1, "raw vectors must reproduce the one-blob failure"

    centered = center_vector_map(vectors)
    cent_clusters = cluster_values(values, centered, threshold=0.30)
    assert len(cent_clusters) == 2, f"centered space must separate the families, got {len(cent_clusters)}"
    memberships = {frozenset(c) for c in cent_clusters}
    assert frozenset(fam_a) in memberships and frozenset(fam_b) in memberships
    print("OK test_centering_separates_families_sharing_dominant_component passed")


def test_centering_deterministic_and_zero_safe():
    """Centering is a pure function; a vector equal to the mean keeps its raw
    vector (no zero-norm cosine)."""
    from src.group import center_vector_map

    vectors = {"x": [1.0, 1.0], "y": [1.0, 1.0]}  # both ARE the mean
    centered = center_vector_map(vectors)
    assert centered["x"] == [1.0, 1.0] and centered["y"] == [1.0, 1.0]

    v2 = {"a": [2.0, 0.0], "b": [0.0, 2.0]}
    assert center_vector_map(v2) == center_vector_map(v2), "must be deterministic"
    assert center_vector_map(v2)["a"] == [1.0, -1.0]
    print("OK test_centering_deterministic_and_zero_safe passed")
