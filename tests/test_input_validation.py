"""User-facing input validation: malformed workbooks must produce actionable,
row-numbered errors — never tracebacks or silent skips. Built for the
production handoff (users author their own input workbooks)."""
import openpyxl
import pytest

from src.io_excel import read_input


def _workbook(tmp_path, entities=None, urls=None, questions=None, config=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("entities"); ws.append(["entity"])
    for e in (entities or []): ws.append([e])
    ws = wb.create_sheet("urls"); ws.append(["url", "depth", "entities"])
    for u in (urls or []): ws.append(u)
    ws = wb.create_sheet("questions"); ws.append(["question", "instructions"])
    for q in (questions or []): ws.append(q)
    ws = wb.create_sheet("config"); ws.append(["setting", "value"])
    for c in (config or []): ws.append(c)
    path = tmp_path / "input.xlsx"
    wb.save(path)
    return str(path)


def test_missing_file_and_non_excel_are_clear(tmp_path):
    with pytest.raises(ValueError, match="not found"):
        read_input(str(tmp_path / "nope.xlsx"))
    bad = tmp_path / "notexcel.xlsx"
    bad.write_text("this is not a workbook")
    with pytest.raises(ValueError, match="Excel workbook"):
        read_input(str(bad))


def test_schemeless_url_fixed_with_note(tmp_path, capsys):
    p = _workbook(tmp_path, entities=["Acme"],
                  urls=[["www.acme.com", 1, "Acme"]],
                  questions=[["HQ?", ""]])
    result = read_input(p)
    assert result.urls[0].url == "https://www.acme.com"
    assert "row 2" in capsys.readouterr().out


def test_garbage_url_is_row_numbered_error(tmp_path):
    p = _workbook(tmp_path, entities=["Acme"],
                  urls=[["https://www.acme.com", 1, "Acme"],
                        ["not a url", 1, "Acme"]],
                  questions=[["HQ?", ""]])
    with pytest.raises(ValueError, match="row 3"):
        read_input(p)


def test_duplicate_question_is_row_numbered_error(tmp_path):
    p = _workbook(tmp_path, entities=["Acme"],
                  urls=[["https://www.acme.com", 1, "Acme"]],
                  questions=[["HQ?", ""], ["Revenue?", ""], ["HQ?", "again"]])
    with pytest.raises(ValueError, match="row 4.*duplicate|duplicate.*row 4"):
        read_input(p)


def test_unknown_entity_reference_still_errors(tmp_path):
    p = _workbook(tmp_path, entities=["Acme"],
                  urls=[["https://www.acme.com", 1, "Acme; Ghost Corp"]],
                  questions=[["HQ?", ""]])
    with pytest.raises(ValueError, match="Ghost Corp"):
        read_input(p)


def test_crawl_flags_accepted_in_config_sheet(tmp_path):
    p = _workbook(tmp_path, entities=["Acme"],
                  urls=[["https://www.acme.com", 1, "Acme"]],
                  questions=[["HQ?", ""]],
                  config=[["CRAWL_SCOPE", "site"],
                          ["CRAWL_RENDER_FOR_DISCOVERY", "true"],
                          ["CRAWL_MAX_PAGES", 40]])
    result = read_input(p)
    assert result.config_overrides["CRAWL_SCOPE"] == "site"
    # end-to-end: the override must reach the Config object parsed
    from pipeline import _build_config
    cfg = _build_config(result.config_overrides)
    assert cfg.crawl_scope == "site"
    assert cfg.crawl_render_for_discovery is True
    assert cfg.crawl_max_pages == 40


def test_unsupported_config_key_lists_supported(tmp_path):
    p = _workbook(tmp_path, entities=["Acme"],
                  urls=[["https://www.acme.com", 1, "Acme"]],
                  questions=[["HQ?", ""]],
                  config=[["CRAWL_SPEED", "fast"]])
    with pytest.raises(ValueError, match="CRAWL_SPEED.*Supported"):
        read_input(p)
