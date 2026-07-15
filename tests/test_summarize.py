"""
Tests for the LLM summary layer (src/summarize.py) and its workbook
integration — fully offline: azure_chat is monkeypatched with canned
responses so the Tier-1 mechanical gate, prompt construction, fallback
behaviour and sheet writing are all exercised without Azure or a key.
Design under test: brain/proposals/llm-summary-layer.md.
"""
import openpyxl
import pytest

import src.summarize as summarize_mod
from config import SUMMARY_MAX_CLAIMS_PER_THEME
from models import ColumnSpec, ExtractedCell, ExtractedRow, PipelineResult, SourceQuote
from src.group import ALL_ITEMS_THEME
from src.io_excel import build_claim_index, write_output_excel
from src.summarize import mechanical_gate, summarize_groups


# ── Fixture data: one entity/question cell, 8 verified claims -> C0001-C0008 ──
# Values are deliberately prose-length (> SUMMARY_TAG_MAX_CHARS): the
# 2026-07-15 deterministic answer route diverts all-short-value cells away
# from the LLM, and these fixtures exercise the LLM path.

_VALUES = [
    "Regulatory clearance for assay Z was granted by the FDA following an expedited review of the submission",
    "Regulatory approval in Europe was confirmed by the EMA for the full diagnostic product family this spring",
    "Regulatory approval for kit A was announced alongside an expanded reimbursement agreement in three markets",
    "Launch of product line A took place at the annual industry congress with immediate availability in the US",
    "New product launch in Japan followed local registration and a distribution partnership with a Tokyo firm",
    "Partnership with Acme Corp was signed to co-develop companion diagnostics over an initial five-year term",
    "Global partnership deal signed with a major pharmaceutical group covering biomarker discovery worldwide",
    "Opened new facility in Kyoto to expand regional manufacturing capacity and shorten delivery lead times",
]


def _rows() -> list[ExtractedRow]:
    cell = ExtractedCell(
        entity="Acme",
        source_url="https://a.example.com/news",
        column="Recent news",
        value=list(_VALUES),
        evidence=[
            SourceQuote(value=v, quote=v, source_url="https://a.example.com/news", verified=True)
            for v in _VALUES
        ],
    )
    return [ExtractedRow(entity="Acme", cells=[cell], all_cells=[cell])]


def _claim_groups() -> list[dict]:
    # Hand-built grouping output, size-desc like group_rows emits it.
    return [
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[0],
         "n_items": 3, "values": _VALUES[0:3], "sources": 1},
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[3],
         "n_items": 2, "values": _VALUES[3:5], "sources": 1},
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[5],
         "n_items": 2, "values": _VALUES[5:7], "sources": 1},
    ]


def _patch_azure(monkeypatch, text=None, error=None, fingerprint="fp_test"):
    """Replace make_client + azure_chat; returns the list of captured prompts."""
    prompts: list[str] = []

    def fake_chat(client, prompt, **kwargs):
        prompts.append(prompt)
        if error is not None:
            return {"text": None, "system_fingerprint": None, "error": error, "duration_ms": 1}
        return {"text": text, "system_fingerprint": fingerprint, "error": None, "duration_ms": 1}

    monkeypatch.setattr(summarize_mod, "make_client", lambda: object())
    monkeypatch.setattr(summarize_mod, "azure_chat", fake_chat)
    return prompts


# ── Claim IDs: pipeline-time index matches the Provenance writer ─────────────

def test_build_claim_index_matches_provenance_order():
    index = build_claim_index(_rows())
    for i, v in enumerate(_VALUES, start=1):
        cid, prov_row = index[("Acme", "Recent news", v.lower())]
        assert cid == f"C{i:04d}"
        assert prov_row == i + 1  # header offset
    print("OK test_build_claim_index_matches_provenance_order passed")


# ── Mechanical gate (Tier 1) ──────────────────────────────────────────────────

_TOP_SETS = [("theme A", {"C0001", "C0002"}), ("theme B", {"C0004"})]
_INPUT_IDS = {"C0001", "C0002", "C0003", "C0004", "C0005"}


def test_gate_passes_fully_cited_covering_summary():
    text = "Regulatory work dominated [C0001], [C0002]. A launch followed [C0004]."
    reasons, cited, uncited = mechanical_gate(text, _INPUT_IDS, _TOP_SETS)
    assert reasons == []
    assert cited == {"C0001", "C0002", "C0004"}
    assert uncited == []
    print("OK test_gate_passes_fully_cited_covering_summary passed")


def test_gate_fails_invented_citation():
    text = "Regulatory work dominated [C0001], [C9999]. A launch followed [C0004]."
    reasons, _, _ = mechanical_gate(text, _INPUT_IDS, _TOP_SETS)
    assert any("invented" in r and "C9999" in r for r in reasons)
    print("OK test_gate_fails_invented_citation passed")


def test_gate_fails_uncited_sentence():
    text = "Regulatory work dominated [C0001]. The company had a busy year. Launch [C0004]."
    reasons, _, uncited = mechanical_gate(text, _INPUT_IDS, _TOP_SETS)
    assert any("uncited" in r for r in reasons)
    assert uncited == ["The company had a busy year."]
    print("OK test_gate_fails_uncited_sentence passed")


def test_gate_fails_uncovered_top_theme():
    # Cites only theme B — theme A (top by size) unrepresented.
    text = "A major launch happened [C0004]."
    reasons, _, _ = mechanical_gate(text, _INPUT_IDS, _TOP_SETS)
    assert any("top theme not cited" in r and "theme A" in r for r in reasons)
    print("OK test_gate_fails_uncovered_top_theme passed")


def test_gate_fails_empty_summary():
    reasons, _, _ = mechanical_gate("", _INPUT_IDS, _TOP_SETS)
    assert any("empty" in r for r in reasons)
    print("OK test_gate_fails_empty_summary passed")


def test_multi_id_citation_parsing():
    # The 2026-07-07 laptop eval: the model batches IDs in one bracket, which
    # the old single-ID regex missed -> whole sentences read as uncited.
    from src.summarize import cited_ids, has_citation

    assert cited_ids("clearances [C0183, C0184, C0185]") == ["C0183", "C0184", "C0185"]
    assert cited_ids("chained [C0001][C0002]") == ["C0001", "C0002"]
    assert cited_ids("single [C0042].") == ["C0042"]
    assert cited_ids("semicolons [C0001; C0002]") == ["C0001", "C0002"]
    assert cited_ids("no citation here") == []
    assert has_citation("clearances [C0183, C0184]") is True
    assert has_citation("no citation") is False

    # A sentence with a multi-ID bracket must NOT count as uncited.
    text = "Danaher advanced bioprocessing [C0183, C0184, C0185, C0189]."
    reasons, cited, uncited = mechanical_gate(
        text, {"C0183", "C0184", "C0185", "C0189"}, [("t", {"C0183"})])
    assert uncited == []
    assert cited == {"C0183", "C0184", "C0185", "C0189"}
    assert reasons == []
    print("OK test_multi_id_citation_parsing passed")


def test_sentence_split_merges_abbreviation_fragments():
    # The 2026-07-07 laptop eval: "Ltd."/"U.S." split prose into citation-less
    # fragments that failed the gate and mis-fed the judge.
    from src.summarize import _split_sentences

    text = ("Aalto Scientific Ltd. develops calibrators [C0001], [C0002]. "
            "It ships to the U.S. and Europe [C0004].")
    assert _split_sentences(text) == [
        "Aalto Scientific Ltd. develops calibrators [C0001], [C0002].",
        "It ships to the U.S. and Europe [C0004].",
    ]
    reasons, _, _ = mechanical_gate(text, _INPUT_IDS, _TOP_SETS)
    assert not any("uncited" in r for r in reasons)
    print("OK test_sentence_split_merges_abbreviation_fragments passed")


def test_split_multiline_s4_output_by_line():
    # s4 compact format: one line per theme. Multi-line text splits on
    # newlines (bullet markers stripped defensively — the prompt forbids
    # them); items separated by commas/periods within a line must NOT be
    # sub-split.
    from src.summarize import _join_units, _split_sentences

    text = ("R&D sites: U.S., Germany, France (more in Provenance) [C0001, C0002]\n"
            "• Focus areas: cytology, molecular diagnostics [C0004]")
    units = _split_sentences(text)
    assert units == [
        "R&D sites: U.S., Germany, France (more in Provenance) [C0001, C0002]",
        "Focus areas: cytology, molecular diagnostics [C0004]",
    ]
    # Gate operates per line: both cited, top themes covered.
    reasons, cited, uncited = mechanical_gate(text, _INPUT_IDS, _TOP_SETS)
    assert reasons == [] and uncited == []
    assert cited == {"C0001", "C0002", "C0004"}
    # Round-trip preserves shape for the corruption legs.
    assert _join_units(units, text) == "\n".join(units)
    assert _join_units(["a [C1].", "b [C2]."], "a [C1]. b [C2].") == "a [C1]. b [C2]."
    print("OK test_split_multiline_s4_output_by_line passed")


def test_tag_only_cell_routed_deterministically(monkeypatch):
    # A cell whose citable input is a single short claim renders without any
    # LLM call — make_client must not even be constructed (it would raise
    # here: no key in the test env, and _patch_azure is deliberately absent).
    tag_cell = ExtractedCell(
        entity="Acme", source_url="https://a.example.com", column="Company type",
        value=["own-product"],
        evidence=[SourceQuote(value="own-product", quote="own-product",
                              source_url="https://a.example.com", verified=True)],
    )
    rows = [ExtractedRow(entity="Acme", cells=[tag_cell], all_cells=[tag_cell])]
    groups = [{"entity": "Acme", "question": "Company type", "theme": ALL_ITEMS_THEME,
               "n_items": 1, "values": ["own-product"], "sources": 1}]

    monkeypatch.setattr(summarize_mod, "make_client",
                        lambda: (_ for _ in ()).throw(AssertionError("LLM client built for tag-only cell")))
    out = summarize_groups(groups, rows)

    assert len(out) == 1
    s = out[0]
    assert s["summary"] == "own-product [C0001]"
    assert s["gate"] == "pass"
    assert s["model"] == "deterministic-answer"
    assert s["cited_ids"] == ["C0001"] and s["input_claim_ids"] == ["C0001"]
    assert s["prompt_version"] == summarize_mod.PROMPT_VERSION
    # The judge/eval legs read Raw Response — empty made tag cells
    # unjudgeable (13 "no sentences" on the 2026-07-14 CMO run).
    assert s["raw_response"] == "own-product [C0001]"
    print("OK test_tag_only_cell_routed_deterministically passed")


# ── Deterministic answer route (2026-07-15, George's analyst format) ─────────

def _short_value_fixture(values: list[str], question: str = "EOL testing?"):
    cell = ExtractedCell(
        entity="Acme", source_url="https://a.example.com", column=question,
        value=list(values),
        evidence=[SourceQuote(value=v, quote=v, source_url="https://a.example.com",
                              verified=True) for v in values],
    )
    rows = [ExtractedRow(entity="Acme", cells=[cell], all_cells=[cell])]
    groups = [{"entity": "Acme", "question": question, "theme": ALL_ITEMS_THEME,
               "n_items": len(values), "values": list(values), "sources": 1}]
    return groups, rows


def test_binary_consensus_collapses_to_one_verdict(monkeypatch):
    # The George-pasted EOL cell: {MIL-STD 810 testing, Yes, True} previously
    # went to the LLM ("Benchmark Electronics confirms having end-of-line
    # testing capability") — now a verbatim verdict + evidence line, no call.
    groups, rows = _short_value_fixture(["MIL-STD 810 testing", "Yes", "True"])
    monkeypatch.setattr(summarize_mod, "make_client",
                        lambda: (_ for _ in ()).throw(AssertionError("LLM built for short-value cell")))
    out = summarize_groups(groups, rows)

    assert len(out) == 1
    s = out[0]
    assert s["summary"] == "Yes [C0002, C0003]; MIL-STD 810 testing [C0001]"
    assert s["model"] == "deterministic-answer" and s["gate"] == "pass"
    assert s["raw_response"] == s["summary"]
    assert s["cited_ids"] == ["C0001", "C0002", "C0003"]
    print("OK test_binary_consensus_collapses_to_one_verdict passed")


def test_binary_conflict_never_merges_verdicts(monkeypatch):
    groups, rows = _short_value_fixture(["Yes", "No"])
    monkeypatch.setattr(summarize_mod, "make_client",
                        lambda: (_ for _ in ()).throw(AssertionError("LLM built for short-value cell")))
    out = summarize_groups(groups, rows)
    assert out[0]["summary"] == "Conflicting: Yes [C0001] / No [C0002]"
    print("OK test_binary_conflict_never_merges_verdicts passed")


def test_multi_value_cell_routes_to_llm_merge(monkeypatch):
    # George's s6c review: "Tempe, Arizona [C..]; Tempe, AZ [C..]" repeats the
    # same fact under variant spellings — semantic dedup is the LLM's job.
    values = ["Tempe, Arizona", "Tempe, AZ", "U.S."]
    groups, rows = _short_value_fixture(values, question="HQ Location(s)?")
    prompts = _patch_azure(monkeypatch, text="Tempe, AZ [C0001, C0002]; U.S. [C0003]")
    out = summarize_groups(groups, rows)

    assert len(prompts) == 1
    p = prompts[0]
    assert "[C0001] Tempe, Arizona" in p and "[C0002] Tempe, AZ" in p
    assert "Merge values that say the same thing" in p
    s = out[0]
    assert s["gate"] == "pass"
    assert s["model"] != "deterministic-answer"
    assert s["summary"] == "Tempe, AZ [C0001, C0002]; U.S. [C0003]"
    # Readable degradation ready if the gate had failed: the verbatim render.
    assert s["fallback_text"] == "Tempe, Arizona [C0001]; Tempe, AZ [C0002]; U.S. [C0003]"
    print("OK test_multi_value_cell_routes_to_llm_merge passed")


def test_merge_route_prompt_uncapped_fallback_capped(monkeypatch):
    from config import SUMMARY_MAX_ITEMS_PER_LINE

    values = [f"Site {chr(65 + i)}, Country {i}" for i in range(SUMMARY_MAX_ITEMS_PER_LINE + 2)]
    groups, rows = _short_value_fixture(values, question="Manufacturing locations")
    prompts = _patch_azure(monkeypatch, text="uncited")  # gate will fail
    out = summarize_groups(groups, rows)

    # The model sees every value (merging may collapse them below the cap)...
    assert all(f"[C{i + 1:04d}] {v}" in prompts[0] for i, v in enumerate(values))
    s = out[0]
    assert s["gate"].startswith("failed citation gate")
    # ...while the deterministic fallback stays capped and marked.
    assert s["fallback_text"].endswith("(more in Provenance)")
    assert s["fallback_text"].count(";") == SUMMARY_MAX_ITEMS_PER_LINE - 1
    assert len(s["input_claim_ids"]) == len(values)
    print("OK test_merge_route_prompt_uncapped_fallback_capped passed")


def test_mixed_cell_verdict_prepended_and_booleans_kept_from_prompt(monkeypatch):
    # Prose route: bare booleans render as a deterministic verdict line; the
    # LLM never sees them, so a '"True" (2 items)' top theme can no longer
    # fail the coverage gate (3 of 6 gate failures on the s6c CMO run).
    long_claim = ("The company confirmed end-of-line testing capability across "
                  "all three manufacturing campuses following the 2024 audit cycle")
    assert len(long_claim) > 80
    groups, rows = _short_value_fixture(["Yes", "True", long_claim])
    prompts = _patch_azure(monkeypatch, text="EOL capability: confirmed across campuses [C0003]")
    out = summarize_groups(groups, rows)

    assert len(prompts) == 1
    assert "[C0003]" in prompts[0]
    assert "C0001" not in prompts[0] and "C0002" not in prompts[0]  # booleans excluded
    s = out[0]
    assert s["gate"] == "pass"
    assert s["summary"] == ("Yes [C0001, C0002]\n"
                            "EOL capability: confirmed across campuses [C0003]")
    assert s["raw_response"] == s["summary"]  # judge/eval see the verdict line too
    assert s["cited_ids"] == ["C0001", "C0002", "C0003"]
    assert s["input_claim_ids"] == ["C0001", "C0002", "C0003"]
    print("OK test_mixed_cell_verdict_prepended_and_booleans_kept_from_prompt passed")


def test_gate_exempts_trailing_provenance_marker():
    # s6c: the model placed "(more in Provenance)" after the final period of
    # single-line output; the fragment must not count as an uncited sentence.
    text = "Tecan acts as an OEM partner [C0001]. Helps with regulation [C0002]. (more in Provenance)"
    reasons, _, uncited = mechanical_gate(text, {"C0001", "C0002"}, [])
    assert uncited == [] and reasons == []
    # A genuinely uncited sentence still fails.
    reasons, _, uncited = mechanical_gate(
        "Cited [C0001]. Totally uncited claim here.", {"C0001"}, [])
    assert uncited == ["Totally uncited claim here."]
    print("OK test_gate_exempts_trailing_provenance_marker passed")


def test_theme_fallback_renders_medoid_claims_not_bookkeeping():
    from src.summarize import _theme_fallback

    values = list(_VALUES)
    cell = ExtractedCell(
        entity="Acme", source_url="u", column="Recent news", value=values,
        evidence=[SourceQuote(value=v, quote=v, verified=True) for v in values],
    )
    rows = [ExtractedRow(entity="Acme", cells=[cell], all_cells=[cell])]
    index = build_claim_index(rows)
    groups = [
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[0],
         "n_items": 3, "values": _VALUES[0:3], "sources": 1},
        # A pure-boolean theme must be skipped, not rendered as "True [C...]".
        {"entity": "Acme", "question": "Recent news", "theme": "True",
         "n_items": 2, "values": [], "sources": 1},
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[3],
         "n_items": 2, "values": _VALUES[3:5], "sources": 1},
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[5],
         "n_items": 2, "values": _VALUES[5:7], "sources": 1},
        {"entity": "Acme", "question": "Recent news", "theme": _VALUES[7],
         "n_items": 1, "values": [_VALUES[7]], "sources": 1},
    ]
    fb = _theme_fallback("Acme", "Recent news", groups, index)
    lines = fb.splitlines()
    assert len(lines) == 3  # SUMMARY_MAX_LINES_PER_CELL
    assert lines[0] == f"{_VALUES[0]} [C0001, C0002, C0003]"
    assert lines[1] == f"{_VALUES[3]} [C0004, C0005]"
    assert lines[2].endswith("(more in Provenance)")  # theme 5 omitted
    assert "True" not in fb
    print("OK test_theme_fallback_renders_medoid_claims_not_bookkeeping passed")


def test_deterministic_answer_unit_rules():
    from src.summarize import deterministic_answer

    # Case/punctuation-insensitive verdicts; tight vocabulary.
    assert deterministic_answer([("C0001", "TRUE"), ("C0002", "yes.")]) == "Yes [C0001, C0002]"
    assert deterministic_answer([("C0001", "False")]) == "No [C0001]"
    # Qualified answers are NOT verdicts — verbatim like any short claim.
    assert deterministic_answer([("C0001", "Yes, via subcontractors")]) == \
        "Yes, via subcontractors [C0001]"
    # Any long value -> None (LLM path).
    assert deterministic_answer([("C0001", "x" * 81)]) is None
    assert deterministic_answer([]) is None
    print("OK test_deterministic_answer_unit_rules passed")


# ── summarize_groups end-to-end (mocked Azure) ────────────────────────────────

# Cites >=1 member of each of the 3 themes; every sentence cited.
_GOOD_SUMMARY = (
    "Acme's news is dominated by regulatory clearances [C0001], [C0002], [C0003]. "
    "It also launched new products [C0004], [C0005]. "
    "Partnerships rounded out the year [C0006], [C0007]."
)


def test_summarize_groups_pass(monkeypatch):
    prompts = _patch_azure(monkeypatch, text=_GOOD_SUMMARY)
    out = summarize_groups(_claim_groups(), _rows())

    assert len(out) == 1
    s = out[0]
    assert s["gate"] == "pass"
    assert s["entity"] == "Acme" and s["question"] == "Recent news"
    assert s["summary"] == _GOOD_SUMMARY
    assert s["cited_ids"] == [f"C{i:04d}" for i in range(1, 8)]
    assert s["input_claim_ids"] == [f"C{i:04d}" for i in range(1, 8)]
    assert s["uncited_sentences"] == []
    assert s["system_fingerprint"] == "fp_test"
    assert s["prompt_version"] == summarize_mod.PROMPT_VERSION
    assert s["prompt"] == prompts[0]

    # Prompt carries the closed ID-tagged set and the theme structure.
    assert '"Recent news"' in prompts[0]
    assert f'Theme "{_VALUES[0]}" (3 claims):' in prompts[0]
    assert f"[C0001] {_VALUES[0]}" in prompts[0]
    print("OK test_summarize_groups_pass passed")


def test_summarize_groups_gate_failure_recorded(monkeypatch):
    _patch_azure(monkeypatch, text="Acme had a great year. Launches happened [C0004].")
    out = summarize_groups(_claim_groups(), _rows())
    assert out[0]["gate"].startswith("failed citation gate:")
    assert out[0]["summary"]  # raw text kept for the Summary Log
    print("OK test_summarize_groups_gate_failure_recorded passed")


def test_summarize_groups_call_error_fail_soft(monkeypatch):
    _patch_azure(monkeypatch, error="connection refused")
    out = summarize_groups(_claim_groups(), _rows())
    assert out[0]["gate"] == "call failed: connection refused"
    assert out[0]["summary"] == ""
    assert out[0]["error"] == "connection refused"
    print("OK test_summarize_groups_call_error_fail_soft passed")


def test_summarize_groups_missing_key_raises(monkeypatch):
    monkeypatch.setattr(summarize_mod, "AZURE_API_KEY", None)
    with pytest.raises(RuntimeError, match="AZURE_API_KEY"):
        summarize_groups(_claim_groups(), _rows())
    print("OK test_summarize_groups_missing_key_raises passed")


def test_prompt_member_cap_marks_overflow_and_hides_ids(monkeypatch):
    n = SUMMARY_MAX_CLAIMS_PER_THEME + 3
    values = [
        f"Distinct verified claim number {i} describing a separate regulatory "
        "clearance for one of the company's diagnostic assay product lines"
        for i in range(n)
    ]
    cell = ExtractedCell(
        entity="Big", source_url="u", column="Q", value=list(values),
        evidence=[SourceQuote(value=v, quote=v, verified=True) for v in values],
    )
    rows = [ExtractedRow(entity="Big", cells=[cell], all_cells=[cell])]
    groups = [{"entity": "Big", "question": "Q", "theme": values[0],
               "n_items": n, "values": values, "sources": 1}]

    prompts = _patch_azure(monkeypatch, text=f"Many claims [C0001]. More [C0002].")
    out = summarize_groups(groups, rows)

    assert "(+3 more claims in this theme, not shown)" in prompts[0]
    # Hidden members' IDs never appear in the prompt (the gate's closed set);
    # the record's input_claim_ids lists the whole cell since s7.
    assert f"C{n:04d}" not in prompts[0]
    assert out[0]["input_claim_ids"] == [f"C{i:04d}" for i in range(1, n + 1)]
    print("OK test_prompt_member_cap_marks_overflow_and_hides_ids passed")


def test_all_items_cell_prompt_has_no_theme_label(monkeypatch):
    values = [
        "alpha corp operates as an original equipment manufacturer serving hospital laboratory customers",
        "beta gmbh operates as a contract developer of reagents and consumables for point-of-care testing",
    ]
    cell = ExtractedCell(
        entity="S", source_url="u", column="Company type", value=list(values),
        evidence=[SourceQuote(value=v, quote=v, verified=True) for v in values],
    )
    rows = [ExtractedRow(entity="S", cells=[cell], all_cells=[cell])]
    groups = [{"entity": "S", "question": "Company type", "theme": ALL_ITEMS_THEME,
               "n_items": 2, "values": values, "sources": 1}]

    prompts = _patch_azure(monkeypatch, text="Both are companies [C0001], [C0002].")
    out = summarize_groups(groups, rows)

    assert "Theme " not in prompts[0]
    assert "not grouped into themes" in prompts[0]
    assert out[0]["gate"] == "pass"  # no top-theme coverage requirement
    print("OK test_all_items_cell_prompt_has_no_theme_label passed")


def test_unresolvable_members_omitted_and_uncitable_cell_skipped(monkeypatch):
    # Values with no matching evidence resolve to no claim ID -> omitted; a
    # cell with nothing citable produces no summary row at all.
    cell = ExtractedCell(
        entity="X", source_url="u", column="Q", value=["known claim"],
        evidence=[SourceQuote(value="known claim", quote="q", verified=True)],
    )
    rows = [ExtractedRow(entity="X", cells=[cell], all_cells=[cell])]
    groups = [{"entity": "X", "question": "Q", "theme": ALL_ITEMS_THEME,
               "n_items": 1, "values": ["completely unknown value"], "sources": 0}]

    prompts = _patch_azure(monkeypatch, text="irrelevant")
    out = summarize_groups(groups, rows)
    assert out == []
    assert prompts == []  # no Azure call made
    print("OK test_unresolvable_members_omitted_and_uncitable_cell_skipped passed")


# ── Workbook integration ─────────────────────────────────────────────────────

# Second question deliberately has no summary -> exercises "No data found".
_COLUMNS = [ColumnSpec(name="Recent news"), ColumnSpec(name="R&D location")]


def _write_workbook(tmp_path, cell_summaries):
    diag = {"claim_groups": _claim_groups()}
    if cell_summaries is not None:
        diag["cell_summaries"] = cell_summaries
    path = str(tmp_path / "out.xlsx")
    write_output_excel(PipelineResult(rows=_rows()), _COLUMNS, path, diag=diag)
    return openpyxl.load_workbook(path)


def _summary_record(**overrides):
    base = {
        "entity": "Acme", "question": "Recent news", "summary": _GOOD_SUMMARY,
        "cited_ids": [f"C{i:04d}" for i in range(1, 8)],
        "uncited_sentences": [], "input_claim_ids": [f"C{i:04d}" for i in range(1, 8)],
        "gate": "pass", "model": "gpt-4.1-mini", "prompt_version": "s1",
        "generated_at": "2026-07-07T12:00:00+00:00", "system_fingerprint": "fp_test",
        "prompt": "PROMPT", "raw_response": _GOOD_SUMMARY, "duration_ms": 900,
        "error": None,
    }
    base.update(overrides)
    return base


def test_ai_summary_sheet_is_matrix_shaped_after_digest(tmp_path):
    wb = _write_workbook(tmp_path, [_summary_record()])
    names = wb.sheetnames
    assert "AI Summary" in names
    assert names.index("AI Summary") == names.index("Digest") + 1
    assert "Summary Log" in names  # DIAGNOSTICS=True in config

    ws = wb["AI Summary"]
    header = [c.value for c in ws[1]]
    # Matrix form: Entity column (carrying the disclaimer) + one column per question.
    assert header[0].startswith("Entity")
    assert "AI-synthesized prose" in header[0]
    assert header[1] == "Recent news" and header[2] == "R&D location"
    assert ws.max_row == 2  # one row per entity, not per (entity, question)
    assert ws.cell(row=2, column=1).value == "Acme"
    assert ws.cell(row=2, column=2).value == _GOOD_SUMMARY
    # No summary for this question -> Matrix conventions.
    assert ws.cell(row=2, column=3).value == "No data found"
    print("OK test_ai_summary_sheet_is_matrix_shaped_after_digest passed")


def test_gate_failed_cell_shows_marked_digest_fallback(tmp_path):
    failed = _summary_record(gate="failed citation gate: 1 uncited sentence(s)")
    wb = _write_workbook(tmp_path, [failed])

    digest_ws = wb["Digest"]
    digest_line = digest_ws.cell(row=2, column=5).value  # Digest column
    cell = wb["AI Summary"].cell(row=2, column=2).value
    assert cell.startswith(digest_line)
    assert "[fallback: deterministic digest — citation gate failed; see Summary Log]" in cell
    print("OK test_gate_failed_cell_shows_marked_digest_fallback passed")


def test_gate_failed_cell_with_fallback_text_shows_verbatim_claims(tmp_path):
    # s7: records carrying fallback_text degrade to readable verbatim claims,
    # not the "N items across M themes" digest bookkeeping (George, s6c).
    failed = _summary_record(
        gate="failed citation gate: 1 uncited sentence(s)",
        fallback_text="Launch of product line A [C0004, C0005]",
    )
    wb = _write_workbook(tmp_path, [failed])
    cell = wb["AI Summary"].cell(row=2, column=2).value
    assert cell.startswith("Launch of product line A [C0004, C0005]")
    assert "[fallback: verbatim claims — citation gate failed; see Summary Log]" in cell
    assert "items across" not in cell
    print("OK test_gate_failed_cell_with_fallback_text_shows_verbatim_claims passed")


def test_call_failed_cell_marked(tmp_path):
    failed = _summary_record(gate="call failed: timeout", summary="", raw_response="",
                             cited_ids=[], error="timeout")
    wb = _write_workbook(tmp_path, [failed])
    cell = wb["AI Summary"].cell(row=2, column=2).value
    assert "[fallback: deterministic digest — call failed; see Summary Log]" in cell
    print("OK test_call_failed_cell_marked passed")


def test_no_summaries_no_new_sheets(tmp_path):
    wb = _write_workbook(tmp_path, None)
    assert "AI Summary" not in wb.sheetnames
    assert "Summary Log" not in wb.sheetnames
    print("OK test_no_summaries_no_new_sheets passed")


def test_summary_log_audit_fields(tmp_path):
    wb = _write_workbook(tmp_path, [_summary_record()])
    ws = wb["Summary Log"]
    header = [c.value for c in ws[1]]
    row = dict(zip(header, [c.value for c in ws[2]]))
    assert row["Gate"] == "pass"
    assert row["Faithfulness"] == "not-assessed"  # judge's write target
    assert row["System Fingerprint"] == "fp_test"
    assert row["Prompt Version"] == "s1"
    assert row["Prompt"] == "PROMPT"
    assert row["Raw Response"] == _GOOD_SUMMARY
    assert row["Model"] == "gpt-4.1-mini"
    print("OK test_summary_log_audit_fields passed")


def test_summary_log_fallback_faithfulness_values(tmp_path):
    records = [
        _summary_record(gate="failed citation gate: x"),
    ]
    wb = _write_workbook(tmp_path, records)
    ws = wb["Summary Log"]
    header = [c.value for c in ws[1]]
    row = dict(zip(header, [c.value for c in ws[2]]))
    assert row["Faithfulness"] == "fallback (failed citation gate)"
    print("OK test_summary_log_fallback_faithfulness_values passed")
