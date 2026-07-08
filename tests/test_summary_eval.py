"""
Offline tests for the Tier-2 judge plumbing (diagnostics/summary_judge.py)
and the corruption generators (diagnostics/summary_eval.py) — no Azure, no
key: everything here is deterministic parsing/generation logic whose
correctness the faithfulness eval depends on.
"""
from diagnostics.summary_eval import (
    corrupt_delete_top_sentence,
    corrupt_inject_fact,
    corrupt_reattach_citation,
    corrupt_swap_entity,
    corrupt_swap_number,
    digest_judgeable_text,
    parse_prompt_themes,
)
import openpyxl

from diagnostics.summary_judge import (
    annotate_matrix_cell,
    build_judge_prompt,
    judge_summary,
    load_claim_texts,
    parse_verdicts,
    sentences_with_claims,
    strip_matrix_annotations,
    verdict_to_cell,
)
from src.summarize import _cell_prompt, _split_sentences, mechanical_gate

# ── Fixture: a gate-passed record as load_passed_summaries returns it ─────────

_SUMMARY = (
    "Acme obtained 3 regulatory clearances [C0001], [C0002]. "
    "It launched products in Japan [C0004], [C0005]. "
    "A partnership was signed [C0006]."
)

_PROMPT = (
    'You are summarizing verified extracted claims about Acme for the question "Recent news".\n'
    "\n"
    'Theme "Regulatory clearance" (3 claims):\n'
    "[C0001] Regulatory clearance for assay Z\n"
    "[C0002] Regulatory approval in Europe\n"
    "\n"
    'Theme "Product launch" (2 claims):\n'
    "[C0004] Launch of product line A\n"
    "[C0005] New product launch in Japan\n"
    "\n"
    'Theme "Partnership" (1 claims):\n'
    "[C0006] Partnership with Acme Corp"
)

_RECORD = {
    "entity": "Acme",
    "question": "Recent news",
    "summary": _SUMMARY,
    "input_ids": {"C0001", "C0002", "C0004", "C0005", "C0006"},
    "prompt": _PROMPT,
}

_CLAIM_TEXTS = {
    "C0001": "Regulatory clearance for assay Z",
    "C0002": "Regulatory approval in Europe",
    "C0004": "Launch of product line A",
    "C0005": "New product launch in Japan",
    "C0006": "Partnership with Acme Corp",
}


# ── Judge plumbing ────────────────────────────────────────────────────────────

def test_parse_verdicts_strict():
    assert parse_verdicts('{"1": "faithful", "2": "unsupported"}', 2) == {
        1: "faithful", 2: "unsupported"}
    # Fenced JSON accepted (mini often fences).
    assert parse_verdicts('```json\n{"1": "faithful"}\n```', 1) == {1: "faithful"}
    # Anything partial/off-vocabulary is a judge FAILURE, never a partial pass.
    assert parse_verdicts('{"1": "faithful"}', 2) is None
    assert parse_verdicts('{"1": "mostly ok"}', 1) is None
    assert parse_verdicts('{"one": "faithful"}', 1) is None
    assert parse_verdicts("not json at all", 1) is None
    assert parse_verdicts('["faithful"]', 1) is None
    print("OK test_parse_verdicts_strict passed")


def test_sentences_with_claims_resolves_ids():
    items = sentences_with_claims(_SUMMARY, _CLAIM_TEXTS)
    assert len(items) == 3
    assert items[0]["cited_ids"] == ["C0001", "C0002"]
    assert items[0]["claims"]["C0001"] == _CLAIM_TEXTS["C0001"]
    assert items[2]["n"] == 3
    print("OK test_sentences_with_claims_resolves_ids passed")


def test_judge_prompt_contains_sentences_and_claims():
    items = sentences_with_claims(_SUMMARY, _CLAIM_TEXTS)
    prompt = build_judge_prompt("Acme", "Recent news", items)
    assert 'Sentence 1: "' in prompt
    assert "[C0001] Regulatory clearance for assay Z" in prompt
    assert "strict JSON" in prompt
    print("OK test_judge_prompt_contains_sentences_and_claims passed")


def test_uncited_sentence_flagged_without_a_call():
    # client=None proves no Azure call happens on this path.
    summary = "Cited sentence [C0001]. Totally uncited sentence."
    verdicts, fp, error = judge_summary(None, "Acme", "Q", summary, _CLAIM_TEXTS)
    assert error is None and fp is None
    assert verdicts == {1: "faithful", 2: "unsupported"}
    assert verdict_to_cell(verdicts) == "1 flagged sentence(s)"
    assert verdict_to_cell(None) == "not-assessed"
    assert verdict_to_cell({1: "faithful"}) == "faithful"
    print("OK test_uncited_sentence_flagged_without_a_call passed")


def test_digest_judgeable_text_strips_all_template_arithmetic():
    line = ('84 items across 7 themes. Top: "Anticoagulant Monitoring Inc. kit" '
            '(19 items) [C0915]; "Blood Gas" (12 items) [C0532].')
    judged = digest_judgeable_text(line)
    assert judged == 'Top: "Anticoagulant Monitoring Inc. kit" [C0915]; "Blood Gas" [C0532].'
    # No counts left anywhere — the judge must never see uncited arithmetic.
    assert "items" not in judged
    # Below-threshold lines (no citations) are not judgeable positives.
    assert digest_judgeable_text("3 items (below grouping threshold — see Grouped Themes).") is None
    print("OK test_digest_judgeable_text_strips_all_template_arithmetic passed")


def test_annotate_matrix_cell_flags_the_right_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Summary"
    ws.append(["Entity — AI-synthesized prose (...)", "Recent news", "R&D location"])
    ws.append(["Acme", "Prose about Acme [C0001].", "No data found"])
    ws.append(["Bruker", "Prose about Bruker [C0002].", "No data found"])

    annotate_matrix_cell(wb, "Bruker", "Recent news", "1 flagged sentence(s)")

    flagged = ws.cell(row=3, column=2)
    assert "[faithfulness: 1 flagged sentence(s) — see Summary Log]" in flagged.value
    assert flagged.fill.start_color.rgb.endswith("FFE0B2")
    # Neighbours untouched.
    assert ws.cell(row=2, column=2).value == "Prose about Acme [C0001]."
    assert ws.cell(row=3, column=3).value == "No data found"
    # Unknown entity/question or missing sheet: silently no-op, never raises.
    annotate_matrix_cell(wb, "Nobody", "Recent news", "x")
    annotate_matrix_cell(openpyxl.Workbook(), "Acme", "Recent news", "x")
    print("OK test_annotate_matrix_cell_flags_the_right_cell passed")


def test_load_claim_texts_reads_claim_column_not_claim_id():
    # Regression: Provenance has "Claim ID" (col 1) BEFORE "Claim" (col 5);
    # a prefix match on "Claim" hit "Claim ID" first, so every claim text
    # loaded as its own ID and the judge flagged everything unsupported
    # (the 12/71-faithful run, 2026-07-07).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provenance"
    ws.append(["Claim ID", "Entity", "Source URL", "Question", "Claim", "Verbatim Quote"])
    ws.append(["C0001", "Acme", "http://x", "Company type", "own-product", "quote"])
    ws.append(["C0002", "Acme", "http://x", "R&D location", "Ettlingen, Germany", "quote"])

    texts = load_claim_texts(wb)
    assert texts == {"C0001": "own-product", "C0002": "Ettlingen, Germany"}
    print("OK test_load_claim_texts_reads_claim_column_not_claim_id passed")


def test_strip_matrix_annotations_reverses_annotate():
    # --rejudge must never stack a fresh verdict on top of a stale one.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Summary"
    ws.append(["Entity — AI-synthesized prose (...)", "Recent news"])
    ws.append(["Acme", "Prose about Acme [C0001]."])
    annotate_matrix_cell(wb, "Acme", "Recent news", "2 flagged sentence(s)")
    assert "[faithfulness:" in ws.cell(row=2, column=2).value

    strip_matrix_annotations(wb)
    cell = ws.cell(row=2, column=2)
    assert cell.value == "Prose about Acme [C0001]."
    assert cell.fill.patternType is None  # orange fill cleared
    # Missing sheet: silently no-op, never raises.
    strip_matrix_annotations(openpyxl.Workbook())
    print("OK test_strip_matrix_annotations_reverses_annotate passed")


def test_parse_prompt_themes_roundtrip_with_real_prompt_builder():
    # Themes recovered from a prompt BUILT BY _cell_prompt must match the
    # shown member IDs — the corruption leg's coverage scoring depends on it.
    claim_index = {
        ("E", "Q", "alpha claim one"): ("C0001", 2),
        ("E", "Q", "alpha claim two"): ("C0002", 3),
        ("E", "Q", "beta claim one"): ("C0003", 4),
    }
    groups = [
        {"entity": "E", "question": "Q", "theme": "Alpha claim one",
         "n_items": 2, "values": ["Alpha claim one", "Alpha claim two"], "sources": 1},
        {"entity": "E", "question": "Q", "theme": "Beta claim one",
         "n_items": 1, "values": ["Beta claim one"], "sources": 1},
    ]
    prompt, input_ids, top_sets = _cell_prompt("E", "Q", groups, claim_index)
    parsed = parse_prompt_themes(prompt)
    assert [(label, ids) for label, ids in parsed] == top_sets
    assert input_ids == {"C0001", "C0002", "C0003"}
    print("OK test_parse_prompt_themes_roundtrip_with_real_prompt_builder passed")


# ── Corruption generators ─────────────────────────────────────────────────────

def test_swap_number_avoids_citation_digits():
    corrupted, bad = corrupt_swap_number(_RECORD, ["Acme", "Bruker"])[0:2]
    assert corrupted != _SUMMARY
    assert "obtained 10 regulatory" in corrupted  # 3 + 7
    # Citation IDs untouched.
    assert "[C0001]" in corrupted and "[C0010]" not in corrupted
    assert bad == {1}
    print("OK test_swap_number_avoids_citation_digits passed")


def test_swap_entity_needs_another_entity():
    corrupted, bad = corrupt_swap_entity(_RECORD, ["Acme", "Bruker"])
    assert "Bruker obtained" in corrupted
    assert bad == {1}
    assert corrupt_swap_entity(_RECORD, ["Acme"]) is None  # nothing to swap to
    print("OK test_swap_entity_needs_another_entity passed")


def test_inject_fact_cites_real_id_and_flags_last_sentence():
    corrupted, bad = corrupt_inject_fact(_RECORD, [])
    sentences = _split_sentences(corrupted)
    assert bad == {len(sentences)}
    assert "$9 billion [C0001]" in sentences[-1]
    # The injected sentence passes the MECHANICAL gate (real ID, cited) —
    # exactly the failure class only the judge can catch.
    themes = parse_prompt_themes(_RECORD["prompt"])[:3]
    reasons, _, _ = mechanical_gate(corrupted, _RECORD["input_ids"], themes)
    assert reasons == []
    print("OK test_inject_fact_cites_real_id_and_flags_last_sentence passed")


def test_reattach_citation_swaps_first_two_citation_sets():
    corrupted, bad = corrupt_reattach_citation(_RECORD, [])
    assert bad == {1, 2}
    sentences = _split_sentences(corrupted)
    assert "[C0004, C0005]" in sentences[0]  # sentence 1 now carries set 2
    assert "[C0001, C0002]" in sentences[1]
    print("OK test_reattach_citation_swaps_first_two_citation_sets passed")


def test_delete_top_sentence_caught_by_coverage_gate_only():
    deleted = corrupt_delete_top_sentence(_RECORD)
    assert "[C0001]" not in deleted and "[C0004]" in deleted
    themes = parse_prompt_themes(_RECORD["prompt"])[:3]
    reasons, _, _ = mechanical_gate(deleted, _RECORD["input_ids"], themes)
    assert any("top theme not cited" in r and "Regulatory clearance" in r for r in reasons)
    print("OK test_delete_top_sentence_caught_by_coverage_gate_only passed")
