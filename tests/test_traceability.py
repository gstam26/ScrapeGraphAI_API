"""
Traceability chain tests (Advisory requirement, 2026-07-03): every theme and
digest line must trace back to verified claims. Chain under test:

    Digest -> Grouped Themes -> Provenance (claim IDs + internal hyperlinks)
           -> source URL (external hyperlink)

Fully offline: builds a small PipelineResult + claim_groups by hand, writes a
real workbook to tmp_path, reads it back with openpyxl.
"""
import openpyxl

from models import ColumnSpec, ExtractedCell, ExtractedRow, PipelineResult, SourceQuote
from src.io_excel import write_output_excel

_URL = "https://example.com/news"
_CLAIMS = [
    "regulatory clearance for assay Z",
    "regulatory approval in Europe",
    "launch of product line A",
    "new product launch in Japan",
]


def _result() -> PipelineResult:
    evidence = [SourceQuote(value=c, quote=c, source_url=_URL, verified=True) for c in _CLAIMS]
    cell = ExtractedCell(entity="TestCo", source_url=_URL, column="Recent news",
                         value=list(_CLAIMS), evidence=evidence, verified=True)
    return PipelineResult(rows=[ExtractedRow(entity="TestCo", cells=[cell], all_cells=[cell])])


def _diag() -> dict:
    return {
        "claim_groups": [
            {"entity": "TestCo", "question": "Recent news",
             "theme": _CLAIMS[0], "n_items": 2, "values": _CLAIMS[:2], "sources": 1},
            {"entity": "TestCo", "question": "Recent news",
             "theme": _CLAIMS[2], "n_items": 2, "values": _CLAIMS[2:], "sources": 1},
        ],
    }


def _write(tmp_path) -> str:
    out = str(tmp_path / "trace.xlsx")
    write_output_excel(_result(), [ColumnSpec(name="Recent news")], out, diag=_diag())
    return out


def test_provenance_claim_ids_sequential_and_unique(tmp_path):
    wb = openpyxl.load_workbook(_write(tmp_path))
    ws = wb["Provenance"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Claim ID"
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert ids == [f"C{i:04d}" for i in range(1, len(_CLAIMS) + 1)]
    wb.close()
    print("OK test_provenance_claim_ids_sequential_and_unique passed")


def test_grouped_themes_bullets_carry_claim_ids(tmp_path):
    wb = openpyxl.load_workbook(_write(tmp_path))
    ws = wb["Grouped Themes"]
    values_cell = str(ws.cell(row=2, column=5).value)
    assert "[C0001]" in values_cell and "[C0002]" in values_cell
    ids_cell = str(ws.cell(row=2, column=6).value)
    assert ids_cell == "C0001, C0002"
    wb.close()
    print("OK test_grouped_themes_bullets_carry_claim_ids passed")


def test_theme_cell_hyperlinks_to_its_provenance_row(tmp_path):
    wb = openpyxl.load_workbook(_write(tmp_path))
    ws = wb["Grouped Themes"]
    link = ws.cell(row=2, column=3).hyperlink  # theme = _CLAIMS[0] = Provenance row 2
    assert link is not None and link.target == "#Provenance!A2"
    link2 = ws.cell(row=3, column=3).hyperlink  # theme = _CLAIMS[2] = Provenance row 4
    assert link2 is not None and link2.target == "#Provenance!A4"
    wb.close()
    print("OK test_theme_cell_hyperlinks_to_its_provenance_row passed")


def test_digest_sheet_template_text_and_link(tmp_path):
    wb = openpyxl.load_workbook(_write(tmp_path))
    assert "Digest" in wb.sheetnames
    # Deliverable-facing order: Summary, Matrix, Digest, Provenance, Grouped Themes.
    names = wb.sheetnames
    assert names.index("Digest") == names.index("Matrix") + 1
    ws = wb["Digest"]
    assert [c.value for c in ws[1]] == ["Entity", "Question", "Items", "Themes", "Digest"]
    digest = str(ws.cell(row=2, column=5).value)
    assert "4 items across 2 themes" in digest
    assert _CLAIMS[0] in digest and "[C0001]" in digest, "top theme label + claim ID must be cited"
    link = ws.cell(row=2, column=2).hyperlink
    assert link is not None and link.target == "#'Grouped Themes'!A2"
    wb.close()
    print("OK test_digest_sheet_template_text_and_link passed")


def test_provenance_source_urls_are_hyperlinks(tmp_path):
    wb = openpyxl.load_workbook(_write(tmp_path))
    ws = wb["Provenance"]
    c = ws.cell(row=2, column=3)
    assert c.hyperlink is not None and c.hyperlink.target == _URL
    wb.close()
    print("OK test_provenance_source_urls_are_hyperlinks passed")


def test_no_digest_or_links_without_groups(tmp_path):
    out = str(tmp_path / "plain.xlsx")
    write_output_excel(_result(), [ColumnSpec(name="Recent news")], out, diag={})
    wb = openpyxl.load_workbook(out)
    assert "Digest" not in wb.sheetnames and "Grouped Themes" not in wb.sheetnames
    # Claim IDs + source-URL links still present — Provenance traceability
    # doesn't depend on grouping.
    ws = wb["Provenance"]
    assert ws.cell(row=2, column=1).value == "C0001"
    assert ws.cell(row=2, column=3).hyperlink is not None
    wb.close()
    print("OK test_no_digest_or_links_without_groups passed")


def test_claim_index_anchors_on_verified_occurrence(tmp_path):
    """Standing decision (2026-07-06): grouping/digest citations resolve to
    VERIFIED Provenance rows. A claim extracted on two pages — unverified on
    the first, verified on the second — must anchor its theme hyperlink and
    Claim ID on the verified row, not simply the first occurrence (the
    audit's gap #1 failure mode)."""
    claim = _CLAIMS[0]
    cell_a = ExtractedCell(
        entity="TestCo", source_url="https://a.example.com", column="Recent news",
        value=[claim],
        evidence=[SourceQuote(value=claim, quote="mangled", source_url="https://a.example.com",
                              verified=False)],
    )
    cell_b = ExtractedCell(
        entity="TestCo", source_url="https://b.example.com", column="Recent news",
        value=[claim],
        evidence=[SourceQuote(value=claim, quote=claim, source_url="https://b.example.com",
                              verified=True)],
    )
    result = PipelineResult(rows=[ExtractedRow(entity="TestCo", cells=[cell_b],
                                               all_cells=[cell_a, cell_b])])
    diag = {"claim_groups": [
        {"entity": "TestCo", "question": "Recent news",
         "theme": claim, "n_items": 1, "values": [claim], "sources": 1},
    ]}
    out = str(tmp_path / "verified_anchor.xlsx")
    write_output_excel(result, [ColumnSpec(name="Recent news")], out, diag=diag)

    wb = openpyxl.load_workbook(out)
    prov = wb["Provenance"]
    # Row 2 = unverified occurrence (C0001), row 3 = verified (C0002).
    assert prov.cell(row=2, column=10).value is False
    assert prov.cell(row=3, column=10).value is True

    ws = wb["Grouped Themes"]
    link = ws.cell(row=2, column=3).hyperlink
    assert link is not None and link.target == "#Provenance!A3", \
        "theme must anchor on the verified occurrence, not the first one"
    assert str(ws.cell(row=2, column=6).value) == "C0002"
    assert "[C0002]" in str(ws.cell(row=2, column=5).value)
    wb.close()
    print("OK test_claim_index_anchors_on_verified_occurrence passed")


def test_unverified_provenance_rows_flagged_orange(tmp_path):
    """Unverified claims stay in Provenance flagged for analyst review — the
    Verified=False cell carries the orange review fill. (The previous check
    compared against 'FALSE' but openpyxl yields Python bools, str() 'False',
    so the flag never rendered.)"""
    claim = _CLAIMS[0]
    cell = ExtractedCell(
        entity="TestCo", source_url=_URL, column="Recent news",
        value=[claim],
        evidence=[SourceQuote(value=claim, quote="mangled", source_url=_URL, verified=False)],
    )
    result = PipelineResult(rows=[ExtractedRow(entity="TestCo", cells=[cell], all_cells=[cell])])
    out = str(tmp_path / "flagged.xlsx")
    write_output_excel(result, [ColumnSpec(name="Recent news")], out, diag={})

    wb = openpyxl.load_workbook(out)
    ws = wb["Provenance"]
    c = ws.cell(row=2, column=10)  # Verified column
    assert c.value is False
    assert "FFE0B2" in str(c.fill.start_color.rgb), "Verified=False must carry the orange review flag"
    wb.close()
    print("OK test_unverified_provenance_rows_flagged_orange passed")


def test_grouped_themes_claim_ids_not_truncated_by_display_cap(tmp_path):
    """Regression for the audit finding (2026-07-05): the 'Claim IDs' column
    was sliced to MATRIX_MAX_DISPLAY_ITEMS in lockstep with the display
    bullets, so a theme with more members than the display cap (HORIBA's
    328-item Recent news cell in the validation run, e.g.) silently dropped
    the claim IDs for every item past the cap — exactly the items the
    "+N more — see Provenance" overflow note tells a consultant to go look
    up, with no ID left to search Provenance by. Claim IDs must list every
    member regardless of the bullet-display cap.
    """
    from config import MATRIX_MAX_DISPLAY_ITEMS

    n = MATRIX_MAX_DISPLAY_ITEMS + 5
    claims = [f"news item {i:03d}" for i in range(n)]
    evidence = [SourceQuote(value=c, quote=c, source_url=_URL, verified=True) for c in claims]
    cell = ExtractedCell(entity="TestCo", source_url=_URL, column="Recent news",
                         value=list(claims), evidence=evidence, verified=True)
    result = PipelineResult(rows=[ExtractedRow(entity="TestCo", cells=[cell], all_cells=[cell])])
    diag = {"claim_groups": [
        {"entity": "TestCo", "question": "Recent news",
         "theme": claims[0], "n_items": n, "values": claims, "sources": 1},
    ]}

    out = str(tmp_path / "overflow.xlsx")
    write_output_excel(result, [ColumnSpec(name="Recent news")], out, diag=diag)
    wb = openpyxl.load_workbook(out)
    ws = wb["Grouped Themes"]
    ids_cell = str(ws.cell(row=2, column=6).value)
    values_cell = str(ws.cell(row=2, column=5).value)

    # Values (bullets) stay capped with the overflow marker...
    assert "[+5 more items" in values_cell
    # ...but Claim IDs lists all n members, not just the displayed cap.
    ids = [x.strip() for x in ids_cell.split(",")]
    assert len(ids) == n, f"expected {n} claim IDs, got {len(ids)}"
    assert ids == [f"C{i:04d}" for i in range(1, n + 1)]
    wb.close()
    print("OK test_grouped_themes_claim_ids_not_truncated_by_display_cap passed")
